"""Memory and resource regression tests for core workflows.

These tests use generated fixtures and generous thresholds to stay stable
across environments while catching unbounded growth or leaks.
"""

from __future__ import annotations

import gc
import tracemalloc
from pathlib import Path

import pytest

from modelops_core.ai.context_builder import build_context_bundle
from modelops_core.exports.export_service import export_model_csv
from modelops_core.fixtures.fixture_generator import generate_fixture_repo
from modelops_core.imports.dataset_profiler import profile_csv, profile_xlsx
from modelops_core.index import build_index
from modelops_core.index.lineage_edges import export_lineage_jsonl
from modelops_core.trace.trace_service import trace_object

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _modelops_memory_diff(func, iterations: int = 10) -> int:
    """Run *func* *iterations* times and return net memory growth (bytes).

    Only counts allocations whose traceback includes ``modelops_core``.
    """
    gc.collect()
    tracemalloc.start()
    before = tracemalloc.take_snapshot()

    for _ in range(iterations):
        func()
        gc.collect()

    after = tracemalloc.take_snapshot()
    tracemalloc.stop()

    diff = after.compare_to(before, "lineno")
    total = 0
    for stat in diff:
        if stat.size_diff > 0:
            tb = "\n".join(str(frame) for frame in stat.traceback)
            if "modelops_core" in tb:
                total += stat.size_diff
    return total


# ---------------------------------------------------------------------------
# 1. Repeated-run memory checks
# ---------------------------------------------------------------------------


class TestRepeatedBuildIndex:
    """Ensure index rebuilds do not leak memory."""

    def test_repeated_build_index_no_unbounded_growth(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="medium")

        def _run() -> None:
            build_index(repo_root=tmp_path, export_jsonl=True, allow_invalid=True)

        # Warm-up to stabilise caches
        _run()
        gc.collect()

        growth = _modelops_memory_diff(_run, iterations=10)
        # Generous ceiling: 30 MB over 10 runs on a medium fixture
        assert growth < 30 * 1024 * 1024, (
            f"build_index memory grew by {growth} bytes over 10 runs; "
            "possible leak or unbounded cache accumulation"
        )


class TestRepeatedTrace:
    """Ensure trace queries do not leak memory."""

    def test_repeated_trace_no_unbounded_growth(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="medium")
        build_index(repo_root=tmp_path, allow_invalid=True)
        db_path = tmp_path / "generated" / "modelops.db"

        def _run() -> None:
            trace_object(db_path, "ATTR-001", max_depth=3, direction="both")

        _run()
        gc.collect()

        growth = _modelops_memory_diff(_run, iterations=20)
        # Generous ceiling: 10 MB over 20 trace calls
        assert growth < 10 * 1024 * 1024, (
            f"trace_object memory grew by {growth} bytes over 20 runs; "
            "possible leak or unbounded cache accumulation"
        )


# ---------------------------------------------------------------------------
# 2. Bounded processing patterns
# ---------------------------------------------------------------------------


class TestProfilerBounds:
    """Dataset profiler must respect max_rows and max_file_size limits."""

    def test_profile_csv_respects_max_rows(self, tmp_path: Path) -> None:
        csv_path = tmp_path / "wide.csv"
        rows = ["col1,col2"] + [f"v{i},v{i}" for i in range(100)]
        csv_path.write_text("\n".join(rows), encoding="utf-8")

        profile = profile_csv(csv_path, dataset_id="wide", max_rows=10)
        assert profile.status.success is True
        assert profile.row_count == 10
        assert profile.status.rows_processed == 10

    def test_profile_csv_rejects_oversized_file(self, tmp_path: Path) -> None:
        csv_path = tmp_path / "huge.csv"
        csv_path.write_text("col1\nvalue\n", encoding="utf-8")
        # Force a tiny limit so the file exceeds it
        profile = profile_csv(
            csv_path,
            dataset_id="huge",
            max_file_size=1,
        )
        assert profile.status.success is False
        assert profile.status.truncated is True
        assert "exceeds limit" in (profile.status.reason or "").lower()

    def test_profile_xlsx_respects_max_rows(self, tmp_path: Path) -> None:
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        xlsx_path = tmp_path / "wide.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(["col1", "col2"])
        for i in range(100):
            ws.append([f"v{i}", f"v{i}"])
        wb.save(str(xlsx_path))
        wb.close()

        profile = profile_xlsx(xlsx_path, dataset_id="wide", max_rows=10)
        assert profile.status.success is True
        sheet = profile.sheets[0]
        assert sheet.row_count == 10
        assert sheet.status.rows_processed == 10

    def test_profile_csv_sampling(self, tmp_path: Path) -> None:
        csv_path = tmp_path / "sampled.csv"
        rows = ["col1,col2"] + [f"v{i},v{i}" for i in range(100)]
        csv_path.write_text("\n".join(rows), encoding="utf-8")

        profile = profile_csv(csv_path, dataset_id="sampled", sample_interval=10)
        assert profile.status.success is True
        assert profile.status.sampled is True
        assert profile.status.sample_interval == 10
        assert profile.row_count == 100
        assert profile.status.rows_processed == 10
        assert "every 10th row" in (profile.status.reason or "")

    def test_profile_xlsx_sampling(self, tmp_path: Path) -> None:
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        xlsx_path = tmp_path / "sampled.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(["col1", "col2"])
        for i in range(100):
            ws.append([f"v{i}", f"v{i}"])
        wb.save(str(xlsx_path))
        wb.close()

        profile = profile_xlsx(xlsx_path, dataset_id="sampled", sample_interval=10)
        assert profile.status.success is True
        assert profile.status.sampled is True
        sheet = profile.sheets[0]
        assert sheet.row_count == 100
        assert sheet.status.rows_processed == 10
        assert "every 10th row" in (sheet.status.reason or "")

    def test_profile_csv_no_sampling_by_default(self, tmp_path: Path) -> None:
        csv_path = tmp_path / "full.csv"
        rows = ["col1,col2"] + [f"v{i},v{i}" for i in range(20)]
        csv_path.write_text("\n".join(rows), encoding="utf-8")

        profile = profile_csv(csv_path, dataset_id="full")
        assert profile.status.sampled is False
        assert profile.row_count == 20
        assert profile.status.rows_processed == 20


class TestTraceBounds:
    """Trace traversal must respect max_depth."""

    def test_trace_respects_max_depth(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="medium")
        build_index(repo_root=tmp_path, allow_invalid=True)
        db_path = tmp_path / "generated" / "modelops.db"

        result_deep = trace_object(db_path, "ATTR-001", max_depth=5, direction="both")
        result_shallow = trace_object(db_path, "ATTR-001", max_depth=1, direction="both")

        assert result_shallow.nodes  # at least some nodes exist
        assert len(result_shallow.nodes) <= len(result_deep.nodes)
        assert all(n.depth <= 1 for n in result_shallow.nodes)


class TestContextBuilderBounds:
    """Context builder must respect object and relationship budgets."""

    def test_context_bundle_respects_max_objects(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="large")
        build_index(repo_root=tmp_path, allow_invalid=True)
        db_path = tmp_path / "generated" / "modelops.db"

        bundle = build_context_bundle(
            db_path=db_path,
            workflow="chat-to-model",
            target_object_id="ATTR-001",
            max_objects=5,
        )
        assert len(bundle.included_objects) <= 5
        assert any("Excluded" in w for w in bundle.warnings)

    def test_context_bundle_respects_max_relationships(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="large")
        build_index(repo_root=tmp_path, allow_invalid=True)
        db_path = tmp_path / "generated" / "modelops.db"

        bundle = build_context_bundle(
            db_path=db_path,
            workflow="chat-to-model",
            target_object_id="ATTR-001",
            max_relationships=3,
        )
        assert len(bundle.relationship_refs) <= 3
        assert any("Excluded" in w for w in bundle.warnings)


# ---------------------------------------------------------------------------
# 3. Large fixture handling
# ---------------------------------------------------------------------------


class TestLargeFixtureHandling:
    """Core workflows must complete without error on large generated fixtures."""

    def test_large_fixture_build_index(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="large")
        summary = build_index(repo_root=tmp_path, export_jsonl=True, allow_invalid=True)
        # Fixture has synthetic gaps; we only care that the build completes
        assert summary is not None

    def test_large_fixture_trace(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="large")
        build_index(repo_root=tmp_path, allow_invalid=True)
        db_path = tmp_path / "generated" / "modelops.db"

        # Trace several object types to exercise different graph regions
        for obj_id in ("ATTR-001", "FEP-001", "MAP-001", "DOMAIN-001"):
            result = trace_object(db_path, obj_id, max_depth=3, direction="both")
            assert result.root_object_id == obj_id

    def test_large_fixture_lineage_export(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="large")
        build_index(repo_root=tmp_path, allow_invalid=True)
        db_path = tmp_path / "generated" / "modelops.db"
        output = tmp_path / "generated" / "lineage_edges.jsonl"

        export_lineage_jsonl(db_path, output)
        assert output.exists()

    def test_large_fixture_context_bundle(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="large")
        build_index(repo_root=tmp_path, allow_invalid=True)
        db_path = tmp_path / "generated" / "modelops.db"

        bundle = build_context_bundle(
            db_path=db_path,
            workflow="chat-to-model",
            target_object_id="ATTR-001",
        )
        assert len(bundle.included_objects) > 0
        assert bundle.bundle_id

    def test_large_fixture_csv_export(self, tmp_path: Path) -> None:
        generate_fixture_repo(tmp_path, profile="large")
        build_index(repo_root=tmp_path, allow_invalid=True)
        export_model_csv(tmp_path / "model")
        csv_dir = tmp_path / "generated" / "exports" / "csv"
        assert any(csv_dir.glob("*.csv"))


# ---------------------------------------------------------------------------
# 4. Actionable diagnostics
# ---------------------------------------------------------------------------


class TestActionableDiagnostics:
    """When limits are exceeded, output must include clear reasons."""

    def test_profile_csv_actionable_reason(self, tmp_path: Path) -> None:
        csv_path = tmp_path / "big.csv"
        csv_path.write_text("col\nval\n", encoding="utf-8")
        profile = profile_csv(csv_path, dataset_id="big", max_file_size=1)
        assert profile.status.reason is not None
        assert "1" in profile.status.reason  # limit value is mentioned

    def test_profile_xlsx_actionable_reason(self, tmp_path: Path) -> None:
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        xlsx_path = tmp_path / "big.xlsx"
        wb = Workbook()
        wb.active.append(["col"])
        wb.active.append(["val"])
        wb.save(str(xlsx_path))
        wb.close()

        profile = profile_xlsx(xlsx_path, dataset_id="big", max_file_size=1)
        assert profile.status.reason is not None
        assert "1" in profile.status.reason
