"""Tests for normalized schema inspection and CLI output."""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

from typer.testing import CliRunner

from modelops_core.cli import app
from modelops_core.imports.schema_inspection import inspect_schema_file

runner = CliRunner()


def _write_json_schema(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "$schema": "https://json-schema.org/draft/2020-12/schema",
                "$id": "https://example.com/product.schema.json",
                "title": "Product",
                "description": "Product payload",
                "type": "object",
                "required": ["id", "name"],
                "properties": {
                    "id": {"type": "string", "description": "Product identifier"},
                    "name": {"type": "string", "maxLength": 255},
                    "status": {"type": "string", "enum": ["active", "draft"]},
                    "tags": {"type": "array", "items": {"type": "string"}},
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def _write_json_payload(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "id": "1000001",
                "name": "Contoso Retail",
                "status": "active",
                "isBlocked": False,
                "salesAreas": [
                    {
                        "salesOrg": "1000",
                        "distributionChannel": "10",
                    }
                ],
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def _write_field_catalog_csv(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "Entity,Field Name,Description,Data Type,Required,Length,Allowed Values",
                "Customer,CustomerId,Customer identifier,string,yes,10,",
                "Customer,Status,Customer status,string,no,,active;blocked",
                "SalesArea,SalesOrg,Sales organization,string,yes,4,",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def _write_field_catalog_xlsx(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CustomerFields"
    worksheet.append(
        ["Entity", "Field Path", "Description", "Data Type", "Required", "Length", "Enum"]
    )
    worksheet.append(
        ["Customer", "CustomerId", "Customer identifier", "string", "yes", 10, ""]
    )
    worksheet.append(
        ["Customer", "Status", "Customer status", "string", "no", "", "active|blocked"]
    )
    worksheet.append(["SalesArea", "SalesOrg", "Sales organization", "string", "yes", 4, ""])
    workbook.save(path)


def _write_sap_mapping_workbook(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Customer Mapping"
    worksheet.append(["Customer sales-area mapping"])
    worksheet.append(
        [
            "Source Field",
            "Source System",
            "Target Table",
            "Target Field",
            "Data Type",
            "Required",
            "Rule",
            "Owner",
            "Description",
        ]
    )
    worksheet.append(
        [
            "customer_group",
            "Legacy CRM",
            "KNVV",
            "KDGRP",
            "CHAR",
            "yes",
            "copy",
            "MDM",
            "Customer Group",
        ]
    )
    worksheet.append(
        [
            "Source Field",
            "Source System",
            "Target Table",
            "Target Field",
            "Data Type",
            "Required",
            "Rule",
            "Owner",
            "Description",
        ]
    )
    worksheet.append(
        [
            "sales_office",
            "Legacy CRM",
            "KNVV",
            "VKBUR",
            "CHAR",
            "no",
            "lookup",
            "MDM",
            "Sales Office",
        ]
    )
    validation = workbook.create_sheet("Validation Results")
    validation.append(["Validation summary"])
    validation.append(["Rule", "Status", "Comment"])
    validation.append(["Customer mapping checked", "pass", "OK"])
    workbook.save(path)


def _write_migration_cockpit_workbook(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Customer"
    worksheet.append(
        [
            "Migration Object",
            "Structure",
            "Field Name",
            "Description",
            "Data Type",
            "Required",
            "Length",
            "Allowed Values",
        ]
    )
    worksheet.append(
        [
            "Customer",
            "KNVV",
            "KDGRP",
            "Customer Group",
            "CHAR",
            "yes",
            2,
            "01;02",
        ]
    )
    worksheet.append(
        [
            "Customer",
            "KNVV",
            "VKBUR",
            "Sales Office",
            "CHAR",
            "no",
            4,
            "",
        ]
    )
    workbook.save(path)


def _write_xml_payload(path: Path) -> None:
    path.write_text(
        """
<CustomerMessage xmlns="urn:customer:payload" actionCode="CREATE">
  <CustomerId>1000001</CustomerId>
  <Status>active</Status>
  <SalesAreas>
    <SalesArea>
      <SalesOrg>1000</SalesOrg>
      <DistributionChannel>10</DistributionChannel>
    </SalesArea>
    <SalesArea>
      <SalesOrg>2000</SalesOrg>
      <DistributionChannel>20</DistributionChannel>
    </SalesArea>
  </SalesAreas>
</CustomerMessage>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_idoc_payload(path: Path) -> None:
    path.write_text(
        """
<DEBMAS07>
  <IDOC BEGIN="1">
    <EDI_DC40 SEGMENT="1">
      <TABNAM>EDI_DC40</TABNAM>
      <MESTYP>DEBMAS</MESTYP>
      <IDOCTYP>DEBMAS07</IDOCTYP>
    </EDI_DC40>
    <E1KNA1M SEGMENT="1">
      <KUNNR>0001000001</KUNNR>
      <KTOKD>Z001</KTOKD>
      <E1KNVVM SEGMENT="1">
        <VKORG>1000</VKORG>
        <VTWEG>10</VTWEG>
      </E1KNVVM>
      <E1KNVVM SEGMENT="1">
        <VKORG>2000</VKORG>
        <VTWEG>20</VTWEG>
      </E1KNVVM>
    </E1KNA1M>
  </IDOC>
</DEBMAS07>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_cds_metadata_csv(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                (
                    "Artifact Name,Artifact Kind,Element Name,Element Kind,Data Type,Length,"
                    "Is Key,Association Target,Min Occurs,Max Occurs,Description,Namespace"
                ),
                (
                    "I_BUSINESSPARTNER,VIEW,BusinessPartner,ELEMENT,CHAR,10,yes,,1,1,"
                    "Business Partner,API_BUSINESS_PARTNER"
                ),
                (
                    "I_BUSINESSPARTNER,VIEW,BusinessPartnerCategory,ELEMENT,CHAR,1,no,,0,1,"
                    "Business Partner Category,API_BUSINESS_PARTNER"
                ),
                (
                    "I_BUSINESSPARTNER,VIEW,to_BusinessPartnerAddress,ASSOCIATION,,,,"
                    "I_BUSINESSPARTNERADDRESS,0,*,Address association,API_BUSINESS_PARTNER"
                ),
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def _write_cds_metadata_json(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "rows": [
                    {
                        "Artifact Name": "I_BUSINESSPARTNER",
                        "Artifact Kind": "VIEW",
                        "Element Name": "BusinessPartner",
                        "Element Kind": "ELEMENT",
                        "Data Type": "CHAR",
                        "Length": 10,
                        "Is Key": "X",
                        "Min Occurs": 1,
                        "Max Occurs": 1,
                        "Description": "Business Partner",
                        "Namespace": "API_BUSINESS_PARTNER",
                    },
                    {
                        "Artifact Name": "I_BUSINESSPARTNER",
                        "Artifact Kind": "VIEW",
                        "Element Name": "to_BusinessPartnerRole",
                        "Element Kind": "ASSOCIATION",
                        "Association Target": "I_BUSINESSPARTNERROLE",
                        "Min Occurs": 0,
                        "Max Occurs": "*",
                        "Description": "Roles",
                        "Namespace": "API_BUSINESS_PARTNER",
                    },
                ]
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def _write_iflow_artifact(path: Path) -> None:
    path.write_text(
        """
<bpmn2:definitions xmlns:bpmn2="http://www.omg.org/spec/BPMN/20100524/MODEL"
    id="com.example.bp.sync" name="BusinessPartnerSync">
  <bpmn2:process id="BusinessPartnerSyncProcess" name="BusinessPartnerSyncProcess">
    <bpmn2:startEvent id="StartEvent_1" name="HTTPS Sender" componentType="https"
        endpointAddress="/business-partner/sync" />
    <bpmn2:serviceTask id="Receiver_1" name="S4 Receiver" adapterType="odata-v2" />
  </bpmn2:process>
</bpmn2:definitions>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_integration_suite_zip(path: Path) -> None:
    iflow_xml = """
<bpmn2:definitions xmlns:bpmn2="http://www.omg.org/spec/BPMN/20100524/MODEL"
    id="com.example.bp.package" name="PackageBusinessPartnerSync">
  <bpmn2:process id="PackageBusinessPartnerSyncProcess" name="PackageBusinessPartnerSyncProcess">
    <bpmn2:startEvent id="StartEvent_1" name="SOAP Sender" componentType="soap"
        endpointAddress="/package/business-partner" />
    <bpmn2:serviceTask id="Receiver_1" name="ERP Receiver" adapterType="idoc" />
  </bpmn2:process>
</bpmn2:definitions>
""".strip()
    manifest = "\n".join(
        [
            "Manifest-Version: 1.0",
            "Bundle-Name: BP Sync Package",
            "Bundle-SymbolicName: com.example.bp.package",
            "Bundle-Version: 1.0.0",
            "Bundle-Vendor: Example Corp",
            "",
        ]
    )
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("META-INF/MANIFEST.MF", manifest)
        archive.writestr(
            "src/main/resources/scenarioflows/integrationflow/main.iflw",
            iflow_xml,
        )
        archive.writestr(
            "src/main/resources/wsdl/businessPartner.wsdl",
            "<definitions />\n",
        )
        archive.writestr(
            "src/main/resources/xsd/businessPartner.xsd",
            "<schema />\n",
        )


def _write_we60_html_bundle(path: Path) -> None:
    path.write_text(
        """
<html>
  <head><title>WE60 Documentation for IDoc Type DEBMAS07</title></head>
  <body>
    <h1>IDoc type DEBMAS07</h1>
    <table>
      <tr>
        <th>Segment type</th>
        <th>Description</th>
        <th>Parent segment</th>
        <th>Min</th>
        <th>Max</th>
      </tr>
      <tr>
        <td>EDI_DC40</td>
        <td>Control record</td>
        <td></td>
        <td>1</td>
        <td>1</td>
      </tr>
      <tr>
        <td>E1KNA1M</td>
        <td>Customer master general data</td>
        <td></td>
        <td>1</td>
        <td>1</td>
      </tr>
      <tr>
        <td>E1KNVVM</td>
        <td>Customer sales area data</td>
        <td>E1KNA1M</td>
        <td>0</td>
        <td>999999</td>
      </tr>
    </table>
    <ul>
      <li><a href="EDI_DC40.html">EDI_DC40</a></li>
      <li><a href="E1KNA1M.html">E1KNA1M</a></li>
      <li><a href="E1KNVVM.html">E1KNVVM</a></li>
    </ul>
  </body>
</html>
""".strip()
        + "\n",
        encoding="utf-8",
    )
    (path.parent / "EDI_DC40.html").write_text(
        """
<html>
  <head><title>Segment EDI_DC40</title></head>
  <body>
    <h2>Segment EDI_DC40</h2>
    <table>
      <tr>
        <th>Field name</th>
        <th>Description</th>
        <th>Data type</th>
        <th>Length</th>
      </tr>
      <tr>
        <td>IDOCTYP</td>
        <td>Basic type</td>
        <td>CHAR</td>
        <td>30</td>
      </tr>
      <tr>
        <td>MESTYP</td>
        <td>Message type</td>
        <td>CHAR</td>
        <td>30</td>
      </tr>
    </table>
  </body>
</html>
""".strip()
        + "\n",
        encoding="utf-8",
    )
    (path.parent / "E1KNA1M.html").write_text(
        """
<html>
  <head><title>Segment E1KNA1M</title></head>
  <body>
    <h2>Segment E1KNA1M</h2>
    <table>
      <tr>
        <th>Field name</th>
        <th>Description</th>
        <th>Data type</th>
        <th>Length</th>
      </tr>
      <tr>
        <td>KUNNR</td>
        <td>Customer number</td>
        <td>CHAR</td>
        <td>10</td>
      </tr>
      <tr>
        <td>KTOKD</td>
        <td>Account group</td>
        <td>CHAR</td>
        <td>4</td>
      </tr>
    </table>
  </body>
</html>
""".strip()
        + "\n",
        encoding="utf-8",
    )
    (path.parent / "E1KNVVM.html").write_text(
        """
<html>
  <head><title>Segment E1KNVVM</title></head>
  <body>
    <h2>Segment E1KNVVM</h2>
    <table>
      <tr>
        <th>Field name</th>
        <th>Description</th>
        <th>Data type</th>
        <th>Length</th>
        <th>Min</th>
        <th>Max</th>
      </tr>
      <tr>
        <td>VKORG</td>
        <td>Sales organization</td>
        <td>CHAR</td>
        <td>4</td>
        <td>1</td>
        <td>1</td>
      </tr>
      <tr>
        <td>VTWEG</td>
        <td>Distribution channel</td>
        <td>CHAR</td>
        <td>2</td>
        <td>0</td>
        <td>1</td>
      </tr>
    </table>
  </body>
</html>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_openapi(path: Path) -> None:
    path.write_text(
        """
openapi: 3.0.3
info:
  title: Product API
  version: 2026-07
servers:
  - url: https://api.example.com/catalog
paths:
  /products:
    get:
      operationId: listProducts
      summary: List products
      responses:
        "200":
          description: Product list
          content:
            application/json:
              schema:
                $ref: "#/components/schemas/Product"
    post:
      operationId: createProduct
      summary: Create product
      requestBody:
        required: true
        content:
          application/json:
            schema:
              $ref: "#/components/schemas/Product"
      responses:
        "201":
          description: Created product
          content:
            application/json:
              schema:
                $ref: "#/components/schemas/Product"
components:
  schemas:
    Product:
      type: object
      required: [id, name]
      properties:
        id:
          type: string
        name:
          type: string
        price:
          type: number
          description: Sales price
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_edmx(path: Path) -> None:
    path.write_text(
        """
<edmx:Edmx Version="4.0"
    xmlns:edmx="http://docs.oasis-open.org/odata/ns/edmx"
    xmlns:edm="http://docs.oasis-open.org/odata/ns/edm">
  <edmx:DataServices>
    <edm:Schema Namespace="API_BUSINESS_PARTNER">
      <edm:EntityType Name="A_BusinessPartner">
        <edm:Key>
          <edm:PropertyRef Name="BusinessPartner"/>
        </edm:Key>
        <edm:Property Name="BusinessPartner" Type="Edm.String" Nullable="false" MaxLength="10" />
        <edm:Property Name="BusinessPartnerCategory" Type="Edm.String" Nullable="true"
          MaxLength="1" />
        <edm:NavigationProperty Name="to_BusinessPartnerAddress"
          Type="Collection(API_BUSINESS_PARTNER.A_BusinessPartnerAddress)" />
      </edm:EntityType>
    </edm:Schema>
  </edmx:DataServices>
</edmx:Edmx>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_xsd(path: Path) -> None:
    path.write_text(
        """
<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"
    targetNamespace="urn:customer" version="1.0">
  <xs:simpleType name="StatusType">
    <xs:restriction base="xs:string">
      <xs:enumeration value="active" />
      <xs:enumeration value="blocked" />
    </xs:restriction>
  </xs:simpleType>
  <xs:complexType name="CustomerType">
    <xs:sequence>
      <xs:element name="CustomerId" type="xs:string" minOccurs="1" />
      <xs:element name="Status" type="StatusType" minOccurs="0" />
      <xs:element name="Tags" minOccurs="0" maxOccurs="unbounded">
        <xs:simpleType>
          <xs:restriction base="xs:string">
            <xs:maxLength value="20" />
          </xs:restriction>
        </xs:simpleType>
      </xs:element>
    </xs:sequence>
    <xs:attribute name="languageCode" type="xs:string" use="required" />
  </xs:complexType>
  <xs:element name="CustomerMessage" type="CustomerType" />
</xs:schema>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_wsdl(path: Path) -> None:
    path.write_text(
        """
<wsdl:definitions name="CustomerService"
    targetNamespace="urn:customer:service"
    xmlns:wsdl="http://schemas.xmlsoap.org/wsdl/"
    xmlns:xs="http://www.w3.org/2001/XMLSchema"
    xmlns:soap="http://schemas.xmlsoap.org/wsdl/soap/"
    xmlns:tns="urn:customer:service">
  <wsdl:types>
    <xs:schema targetNamespace="urn:customer:service">
      <xs:element name="CreateCustomerRequest">
        <xs:complexType>
          <xs:sequence>
            <xs:element name="CustomerId" type="xs:string" minOccurs="1" />
            <xs:element name="Status" minOccurs="0">
              <xs:simpleType>
                <xs:restriction base="xs:string">
                  <xs:enumeration value="active" />
                  <xs:enumeration value="blocked" />
                </xs:restriction>
              </xs:simpleType>
            </xs:element>
          </xs:sequence>
        </xs:complexType>
      </xs:element>
      <xs:element name="CreateCustomerResponse">
        <xs:complexType>
          <xs:sequence>
            <xs:element name="BusinessPartner" type="xs:string" minOccurs="1" />
          </xs:sequence>
        </xs:complexType>
      </xs:element>
    </xs:schema>
  </wsdl:types>
  <wsdl:message name="CreateCustomerInput">
    <wsdl:part name="parameters" element="tns:CreateCustomerRequest" />
  </wsdl:message>
  <wsdl:message name="CreateCustomerOutput">
    <wsdl:part name="parameters" element="tns:CreateCustomerResponse" />
  </wsdl:message>
  <wsdl:portType name="CustomerPortType">
    <wsdl:operation name="CreateCustomer">
      <wsdl:input message="tns:CreateCustomerInput" />
      <wsdl:output message="tns:CreateCustomerOutput" />
    </wsdl:operation>
  </wsdl:portType>
  <wsdl:binding name="CustomerBinding" type="tns:CustomerPortType">
    <soap:binding style="document" transport="http://schemas.xmlsoap.org/soap/http" />
    <wsdl:operation name="CreateCustomer">
      <soap:operation soapAction="urn:createCustomer" />
    </wsdl:operation>
  </wsdl:binding>
</wsdl:definitions>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def test_inspect_json_schema_normalizes_fields(tmp_path: Path) -> None:
    schema_path = tmp_path / "product.schema.json"
    _write_json_schema(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "json_schema"
    assert document.source_identity == "Product"
    assert document.namespace == "https://example.com/product.schema.json"
    assert len(document.entities) == 1
    field_paths = {field.field_path: field for field in document.fields}
    assert field_paths["id"].required is True
    assert field_paths["id"].cardinality == "1..1"
    assert field_paths["status"].enumerations == ["active", "draft"]
    assert field_paths["tags"].data_type == "array"
    assert field_paths["tags[]"].data_type == "string"


def test_inspect_json_payload_normalizes_nested_fields(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-payload.json"
    _write_json_payload(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "json_payload"
    assert document.source_identity == "customer-payload"
    assert {entity.name for entity in document.entities} == {"customer-payload"}
    field_paths = {field.field_path: field for field in document.fields}
    assert field_paths["id"].data_type == "string"
    assert field_paths["id"].required is True
    assert field_paths["isBlocked"].data_type == "boolean"
    assert field_paths["salesAreas"].data_type == "array"
    assert field_paths["salesAreas"].cardinality == "1..*"
    assert field_paths["salesAreas[]"].data_type == "object"
    assert field_paths["salesAreas[].salesOrg"].data_type == "string"
    assert field_paths["salesAreas[].distributionChannel"].data_type == "string"


def test_inspect_xml_payload_normalizes_nested_fields(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-payload.xml"
    _write_xml_payload(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "xml_payload"
    assert document.source_identity == "CustomerMessage"
    assert document.namespace == "urn:customer:payload"
    assert {entity.name for entity in document.entities} == {"CustomerMessage"}
    field_paths = {field.field_path: field for field in document.fields}
    assert field_paths["@actionCode"].data_type == "string"
    assert field_paths["CustomerId"].data_type == "integer"
    assert field_paths["Status"].data_type == "string"
    assert field_paths["SalesAreas"].data_type == "object"
    assert field_paths["SalesAreas.SalesArea"].data_type == "array"
    assert field_paths["SalesAreas.SalesArea[]"].data_type == "object"
    assert field_paths["SalesAreas.SalesArea[].SalesOrg"].data_type == "integer"
    assert (
        field_paths["SalesAreas.SalesArea[].DistributionChannel"].data_type == "integer"
    )


def test_inspect_idoc_payload_normalizes_segments_and_fields(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-idoc.xml"
    _write_idoc_payload(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "idoc_payload"
    assert document.source_identity == "DEBMAS07"
    assert {entity.name for entity in document.entities} == {"EDI_DC40", "E1KNA1M", "E1KNVVM"}
    entity_kinds = {entity.name: entity.kind for entity in document.entities}
    assert entity_kinds["EDI_DC40"] == "idoc_control_record"
    assert entity_kinds["E1KNA1M"] == "idoc_segment"
    field_map = {(field.entity_name, field.field_path): field for field in document.fields}
    assert field_map[("EDI_DC40", "IDOCTYP")].data_type == "string"
    assert field_map[("E1KNA1M", "KUNNR")].data_type == "integer"
    assert field_map[("E1KNA1M", "KUNNR")].annotations["idoc_segment_path"] == "IDOC.E1KNA1M"
    assert field_map[("E1KNVVM", "VKORG")].data_type == "integer"
    assert field_map[("E1KNVVM", "VKORG")].annotations["idoc_parent_segment"] == "E1KNA1M"
    assert any("representative observed segments" in warning for warning in document.warnings)


def test_inspect_cds_metadata_csv_normalizes_artifacts_and_elements(tmp_path: Path) -> None:
    schema_path = tmp_path / "cds-metadata.csv"
    _write_cds_metadata_csv(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "cds_metadata_export"
    assert document.source_identity == "I_BUSINESSPARTNER"
    assert document.namespace == "API_BUSINESS_PARTNER"
    assert {entity.name for entity in document.entities} == {"I_BUSINESSPARTNER"}
    field_map = {(field.entity_name, field.field_path): field for field in document.fields}
    assert field_map[("I_BUSINESSPARTNER", "BusinessPartner")].required is True
    assert field_map[("I_BUSINESSPARTNER", "BusinessPartner")].cardinality == "1..1"
    assert field_map[("I_BUSINESSPARTNER", "BusinessPartner")].is_key is True
    assert (
        field_map[("I_BUSINESSPARTNER", "to_BusinessPartnerAddress")].associations
        == ["I_BUSINESSPARTNERADDRESS"]
    )
    assert (
        field_map[("I_BUSINESSPARTNER", "to_BusinessPartnerAddress")].annotations[
            "element_kind"
        ]
        == "ASSOCIATION"
    )


def test_inspect_cds_metadata_json_normalizes_row_exports(tmp_path: Path) -> None:
    schema_path = tmp_path / "cds-metadata.json"
    _write_cds_metadata_json(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "cds_metadata_export"
    assert document.source_identity == "I_BUSINESSPARTNER"
    field_map = {(field.entity_name, field.field_path): field for field in document.fields}
    assert field_map[("I_BUSINESSPARTNER", "BusinessPartner")].length == 10
    assert field_map[("I_BUSINESSPARTNER", "to_BusinessPartnerRole")].required is False
    assert field_map[("I_BUSINESSPARTNER", "to_BusinessPartnerRole")].cardinality == "0..*"


def test_inspect_iflow_artifact_normalizes_operation(tmp_path: Path) -> None:
    schema_path = tmp_path / "main.iflw"
    _write_iflow_artifact(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "integration_flow_artifact"
    assert document.source_identity == "BusinessPartnerSync"
    assert {entity.name for entity in document.entities} == {"BusinessPartnerSync"}
    assert len(document.operations) == 1
    operation = document.operations[0]
    assert operation.operation_id == "BusinessPartnerSync"
    assert operation.protocol == "https"
    assert operation.path == "/business-partner/sync"


def test_inspect_integration_suite_zip_normalizes_package_and_resources(tmp_path: Path) -> None:
    schema_path = tmp_path / "bp-sync-package.zip"
    _write_integration_suite_zip(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "integration_suite_package"
    assert document.source_identity == "com.example.bp.package"
    assert document.source_version == "1.0.0"
    assert document.namespace == "Example Corp"
    assert {entity.name for entity in document.entities} == {
        "com.example.bp.package",
        "PackageBusinessPartnerSync",
    }
    assert len(document.operations) == 1
    assert document.operations[0].protocol == "soap"
    field_map = {field.field_path: field for field in document.fields}
    assert field_map["wsdl/businessPartner.wsdl"].data_type == "wsdl_resource"
    assert field_map["xsd/businessPartner.xsd"].data_type == "xsd_resource"


def test_inspect_we60_html_normalizes_segments_and_linked_fields(tmp_path: Path) -> None:
    schema_path = tmp_path / "DEBMAS07.html"
    _write_we60_html_bundle(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "we60_html_documentation"
    assert document.source_identity == "DEBMAS07"
    assert {entity.name for entity in document.entities} == {"EDI_DC40", "E1KNA1M", "E1KNVVM"}
    field_map = {(field.entity_name, field.field_path): field for field in document.fields}
    assert field_map[("EDI_DC40", "IDOCTYP")].data_type == "CHAR"
    assert field_map[("EDI_DC40", "IDOCTYP")].length == 30
    assert field_map[("E1KNA1M", "KUNNR")].description == "Customer number"
    assert field_map[("E1KNVVM", "VKORG")].required is True
    assert field_map[("E1KNVVM", "VKORG")].cardinality == "1..1"
    assert field_map[("E1KNVVM", "VTWEG")].required is False
    assert field_map[("E1KNVVM", "VTWEG")].annotations["we60_source_page"] == "E1KNVVM.html"


def test_inspect_csv_field_catalog_normalizes_entities_and_fields(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-field-catalog.csv"
    _write_field_catalog_csv(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "field_catalog"
    assert document.source_identity == "customer-field-catalog"
    assert {entity.name for entity in document.entities} == {"Customer", "SalesArea"}
    field_map = {(field.entity_name, field.field_path): field for field in document.fields}
    assert field_map[("Customer", "CustomerId")].required is True
    assert field_map[("Customer", "CustomerId")].length == 10
    assert field_map[("Customer", "Status")].enumerations == ["active", "blocked"]
    assert field_map[("SalesArea", "SalesOrg")].required is True


def test_inspect_xlsx_field_catalog_normalizes_entities_and_fields(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-field-catalog.xlsx"
    _write_field_catalog_xlsx(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "field_catalog"
    assert document.source_identity == "customer-field-catalog"
    assert {entity.name for entity in document.entities} == {"Customer", "SalesArea"}
    field_map = {(field.entity_name, field.field_path): field for field in document.fields}
    assert field_map[("Customer", "CustomerId")].required is True
    assert field_map[("Customer", "CustomerId")].length == 10
    assert field_map[("Customer", "Status")].enumerations == ["active", "blocked"]
    assert field_map[("SalesArea", "SalesOrg")].required is True


def test_inspect_sap_mapping_workbook_normalizes_target_fields(tmp_path: Path) -> None:
    schema_path = tmp_path / "sap-customer-mapping.xlsx"
    _write_sap_mapping_workbook(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "sap_mapping_workbook"
    assert document.source_identity == "sap-customer-mapping"
    assert {entity.name for entity in document.entities} == {"KNVV"}
    field_map = {(field.entity_name, field.field_path): field for field in document.fields}
    assert field_map[("KNVV", "KDGRP")].required is True
    assert field_map[("KNVV", "KDGRP")].data_type == "CHAR"
    assert field_map[("KNVV", "KDGRP")].annotations["source_field"] == "customer_group"
    assert field_map[("KNVV", "KDGRP")].annotations["mapping_rule"] == "copy"
    assert field_map[("KNVV", "VKBUR")].required is False
    assert field_map[("KNVV", "VKBUR")].annotations["source_field"] == "sales_office"


def test_inspect_migration_cockpit_workbook_normalizes_structure_fields(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "migration-cockpit-template.xlsx"
    _write_migration_cockpit_workbook(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "sap_migration_cockpit_template"
    assert document.source_identity == "migration-cockpit-template"
    assert {entity.name for entity in document.entities} == {"KNVV"}
    field_map = {(field.entity_name, field.field_path): field for field in document.fields}
    assert field_map[("KNVV", "KDGRP")].required is True
    assert field_map[("KNVV", "KDGRP")].length == 2
    assert field_map[("KNVV", "KDGRP")].enumerations == ["01", "02"]
    assert field_map[("KNVV", "KDGRP")].annotations["migration_object"] == "Customer"
    assert field_map[("KNVV", "VKBUR")].required is False


def test_inspect_openapi_normalizes_entities_and_operations(tmp_path: Path) -> None:
    schema_path = tmp_path / "product-api.yaml"
    _write_openapi(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "openapi"
    assert document.source_identity == "Product API"
    assert document.source_version == "3.0.3"
    assert document.namespace == "https://api.example.com/catalog"
    assert {entity.name for entity in document.entities} == {"Product"}
    assert len(document.operations) == 2
    operation_ids = {operation.operation_id for operation in document.operations}
    assert operation_ids == {"listProducts", "createProduct"}
    assert any(
        field.entity_name == "Product" and field.field_path == "price"
        for field in document.fields
    )


def test_cli_schema_inspect_json_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "product.schema.json"
    _write_json_schema(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "json_schema"
    assert payload["source_identity"] == "Product"
    assert payload["fields"]


def test_cli_schema_inspect_json_payload_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-payload.json"
    _write_json_payload(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "json_payload"
    assert payload["fields"]


def test_cli_schema_inspect_xml_payload_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-payload.xml"
    _write_xml_payload(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "xml_payload"
    assert payload["fields"]


def test_cli_schema_inspect_idoc_payload_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-idoc.xml"
    _write_idoc_payload(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "idoc_payload"
    assert payload["fields"]


def test_cli_schema_inspect_cds_metadata_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "cds-metadata.json"
    _write_cds_metadata_json(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "cds_metadata_export"
    assert payload["fields"]


def test_cli_schema_inspect_iflow_artifact_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "main.iflw"
    _write_iflow_artifact(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "integration_flow_artifact"
    assert payload["operations"]


def test_cli_schema_inspect_integration_suite_zip_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "bp-sync-package.zip"
    _write_integration_suite_zip(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "integration_suite_package"
    assert payload["operations"]


def test_cli_schema_inspect_we60_html_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "DEBMAS07.html"
    _write_we60_html_bundle(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "we60_html_documentation"
    assert payload["fields"]


def test_cli_schema_inspect_csv_field_catalog_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-field-catalog.csv"
    _write_field_catalog_csv(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "field_catalog"
    assert payload["fields"]


def test_cli_schema_inspect_xlsx_field_catalog_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-field-catalog.xlsx"
    _write_field_catalog_xlsx(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "field_catalog"
    assert payload["fields"]


def test_cli_schema_inspect_sap_mapping_workbook_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "sap-customer-mapping.xlsx"
    _write_sap_mapping_workbook(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "sap_mapping_workbook"
    assert payload["fields"]


def test_cli_schema_inspect_migration_cockpit_workbook_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "migration-cockpit-template.xlsx"
    _write_migration_cockpit_workbook(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "sap_migration_cockpit_template"
    assert payload["fields"]


def test_cli_schema_inspect_openapi_human_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "product-api.yaml"
    _write_openapi(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path)])

    assert result.exit_code == 0, result.output
    assert "Schema inspection:" in result.output
    assert "Format:" in result.output
    assert "Operations:" in result.output


def test_inspect_edmx_normalizes_entity_types_and_properties(tmp_path: Path) -> None:
    schema_path = tmp_path / "api_business_partner.edmx"
    _write_edmx(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "edmx"
    assert document.source_identity == "API_BUSINESS_PARTNER"
    assert document.source_version == "4.0"
    assert document.namespace == "API_BUSINESS_PARTNER"
    assert {entity.name for entity in document.entities} == {"A_BusinessPartner"}
    field_paths = {field.field_path: field for field in document.fields}
    assert field_paths["BusinessPartner"].required is True
    assert field_paths["BusinessPartner"].is_key is True
    assert field_paths["BusinessPartnerCategory"].length == 1
    assert field_paths["to_BusinessPartnerAddress"].associations == [
        "Collection(API_BUSINESS_PARTNER.A_BusinessPartnerAddress)"
    ]


def test_inspect_xsd_normalizes_elements_and_restrictions(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer.xsd"
    _write_xsd(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "xsd"
    assert document.source_identity == "urn:customer"
    assert document.source_version == "1.0"
    assert document.namespace == "urn:customer"
    assert {entity.name for entity in document.entities} == {"CustomerMessage"}
    field_paths = {field.field_path: field for field in document.fields}
    assert field_paths["CustomerId"].required is True
    assert field_paths["CustomerId"].cardinality == "1..1"
    assert field_paths["Status"].enumerations == ["active", "blocked"]
    assert field_paths["Tags"].cardinality == "0..*"
    assert field_paths["Tags"].length == 20
    assert field_paths["@languageCode"].required is True


def test_inspect_wsdl_normalizes_messages_and_operations(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-service.wsdl"
    _write_wsdl(schema_path)

    document = inspect_schema_file(schema_path)

    assert document.source_format == "wsdl"
    assert document.source_identity == "CustomerService"
    assert document.namespace == "urn:customer:service"
    assert len(document.operations) == 1
    assert {entity.name for entity in document.entities} >= {
        "CreateCustomerRequest",
        "CreateCustomerResponse",
        "CreateCustomerInput",
        "CreateCustomerOutput",
    }
    operation = document.operations[0]
    assert operation.operation_id == "CreateCustomer"
    assert operation.protocol == "soap"
    assert operation.request_body_schema == "CreateCustomerInput"
    assert operation.response_schemas == [
        {"status_code": "output", "schema": "CreateCustomerOutput"}
    ]
    field_paths = {
        (field.entity_name, field.field_path): field for field in document.fields
    }
    assert field_paths[("CreateCustomerInput", "parameters")].associations == [
        "tns:CreateCustomerRequest"
    ]
    assert field_paths[("CreateCustomerInput", "parameters.CustomerId")].required is True
    assert field_paths[("CreateCustomerInput", "parameters.Status")].enumerations == [
        "active",
        "blocked",
    ]


def test_cli_schema_inspect_xsd_json_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer.xsd"
    _write_xsd(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "xsd"
    assert payload["entities"]


def test_cli_schema_inspect_wsdl_json_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-service.wsdl"
    _write_wsdl(schema_path)

    result = runner.invoke(app, ["schema", "inspect", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["source_format"] == "wsdl"
    assert payload["operations"]
