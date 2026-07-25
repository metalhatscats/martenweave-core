"""Tests for governed schema import proposals."""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

from typer.testing import CliRunner

from modelops_core.cli import app
from modelops_core.config import resolve_generated_path
from modelops_core.imports.schema_import_service import inspect_to_proposal, write_schema_proposal
from modelops_core.patching.apply_service import apply_patch_proposal
from modelops_core.patching.patch_proposal_service import transition_patch_proposal_status
from modelops_core.reports.source_registry_service import SourceRegistryService
from modelops_core.trace.trace_service import trace_object

runner = CliRunner()


def _write_json_schema(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "$schema": "https://json-schema.org/draft/2020-12/schema",
                "$id": "https://example.com/customer.schema.json",
                "title": "Customer",
                "description": "Customer payload",
                "type": "object",
                "required": ["id", "name"],
                "properties": {
                    "id": {"type": "string", "description": "Customer identifier"},
                    "name": {"type": "string"},
                    "status": {"type": "string", "enum": ["active", "blocked"]},
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def _write_json_payload(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "id": "1000001",
                "name": "Contoso Retail",
                "status": "active",
                "salesAreas": [
                    {
                        "salesOrg": "1000",
                        "distributionChannel": "10",
                    }
                ],
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def _write_field_catalog_csv(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "Entity,Field Name,Description,Data Type,Required,Length,Allowed Values",
                "Customer,CustomerId,Customer identifier,string,yes,10,",
                "Customer,Status,Customer status,string,no,,active;blocked",
                "SalesArea,SalesOrg,Sales organization,string,yes,4,",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def _write_field_catalog_xlsx(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CustomerFields"
    worksheet.append(
        ["Entity", "Field Path", "Description", "Data Type", "Required", "Length", "Enum"]
    )
    worksheet.append(["Customer", "CustomerId", "Customer identifier", "string", "yes", 10, ""])
    worksheet.append(
        ["Customer", "Status", "Customer status", "string", "no", "", "active|blocked"]
    )
    worksheet.append(["SalesArea", "SalesOrg", "Sales organization", "string", "yes", 4, ""])
    workbook.save(path)


def _write_sap_mapping_workbook(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Customer Mapping"
    worksheet.append(["Customer sales-area mapping"])
    worksheet.append(
        [
            "Source Field",
            "Source System",
            "Target Table",
            "Target Field",
            "Data Type",
            "Required",
            "Rule",
            "Owner",
            "Description",
        ]
    )
    worksheet.append(
        [
            "customer_group",
            "Legacy CRM",
            "KNVV",
            "KDGRP",
            "CHAR",
            "yes",
            "copy",
            "MDM",
            "Customer Group",
        ]
    )
    worksheet.append(
        [
            "Source Field",
            "Source System",
            "Target Table",
            "Target Field",
            "Data Type",
            "Required",
            "Rule",
            "Owner",
            "Description",
        ]
    )
    worksheet.append(
        [
            "sales_office",
            "Legacy CRM",
            "KNVV",
            "VKBUR",
            "CHAR",
            "no",
            "lookup",
            "MDM",
            "Sales Office",
        ]
    )
    workbook.save(path)


def _write_migration_cockpit_workbook(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Customer"
    worksheet.append(
        [
            "Migration Object",
            "Structure",
            "Field Name",
            "Description",
            "Data Type",
            "Required",
            "Length",
            "Allowed Values",
        ]
    )
    worksheet.append(
        [
            "Customer",
            "KNVV",
            "KDGRP",
            "Customer Group",
            "CHAR",
            "yes",
            2,
            "01;02",
        ]
    )
    worksheet.append(
        [
            "Customer",
            "KNVV",
            "VKBUR",
            "Sales Office",
            "CHAR",
            "no",
            4,
            "",
        ]
    )
    workbook.save(path)


def _write_xml_payload(path: Path) -> None:
    path.write_text(
        """
<CustomerMessage xmlns="urn:customer:payload" actionCode="CREATE">
  <CustomerId>1000001</CustomerId>
  <Status>active</Status>
  <SalesAreas>
    <SalesArea>
      <SalesOrg>1000</SalesOrg>
      <DistributionChannel>10</DistributionChannel>
    </SalesArea>
  </SalesAreas>
</CustomerMessage>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_idoc_payload(path: Path) -> None:
    path.write_text(
        """
<DEBMAS07>
  <IDOC BEGIN="1">
    <EDI_DC40 SEGMENT="1">
      <TABNAM>EDI_DC40</TABNAM>
      <MESTYP>DEBMAS</MESTYP>
      <IDOCTYP>DEBMAS07</IDOCTYP>
    </EDI_DC40>
    <E1KNA1M SEGMENT="1">
      <KUNNR>0001000001</KUNNR>
      <KTOKD>Z001</KTOKD>
      <E1KNVVM SEGMENT="1">
        <VKORG>1000</VKORG>
        <VTWEG>10</VTWEG>
      </E1KNVVM>
    </E1KNA1M>
  </IDOC>
</DEBMAS07>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_cds_metadata_json(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "rows": [
                    {
                        "Artifact Name": "I_BUSINESSPARTNER",
                        "Artifact Kind": "VIEW",
                        "Element Name": "BusinessPartner",
                        "Element Kind": "ELEMENT",
                        "Data Type": "CHAR",
                        "Length": 10,
                        "Is Key": "X",
                        "Min Occurs": 1,
                        "Max Occurs": 1,
                        "Description": "Business Partner",
                        "Namespace": "API_BUSINESS_PARTNER",
                    },
                    {
                        "Artifact Name": "I_BUSINESSPARTNER",
                        "Artifact Kind": "VIEW",
                        "Element Name": "to_BusinessPartnerRole",
                        "Element Kind": "ASSOCIATION",
                        "Association Target": "I_BUSINESSPARTNERROLE",
                        "Min Occurs": 0,
                        "Max Occurs": "*",
                        "Description": "Roles",
                        "Namespace": "API_BUSINESS_PARTNER",
                    },
                ]
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def _write_iflow_artifact(path: Path) -> None:
    path.write_text(
        """
<bpmn2:definitions xmlns:bpmn2="http://www.omg.org/spec/BPMN/20100524/MODEL"
    id="com.example.bp.sync" name="BusinessPartnerSync">
  <bpmn2:process id="BusinessPartnerSyncProcess" name="BusinessPartnerSyncProcess">
    <bpmn2:startEvent id="StartEvent_1" name="HTTPS Sender" componentType="https"
        endpointAddress="/business-partner/sync" />
    <bpmn2:serviceTask id="Receiver_1" name="S4 Receiver" adapterType="odata-v2" />
  </bpmn2:process>
</bpmn2:definitions>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_integration_suite_zip(path: Path) -> None:
    iflow_xml = """
<bpmn2:definitions xmlns:bpmn2="http://www.omg.org/spec/BPMN/20100524/MODEL"
    id="com.example.bp.package" name="PackageBusinessPartnerSync">
  <bpmn2:process id="PackageBusinessPartnerSyncProcess" name="PackageBusinessPartnerSyncProcess">
    <bpmn2:startEvent id="StartEvent_1" name="SOAP Sender" componentType="soap"
        endpointAddress="/package/business-partner" />
    <bpmn2:serviceTask id="Receiver_1" name="ERP Receiver" adapterType="idoc" />
  </bpmn2:process>
</bpmn2:definitions>
""".strip()
    manifest = "\n".join(
        [
            "Manifest-Version: 1.0",
            "Bundle-Name: BP Sync Package",
            "Bundle-SymbolicName: com.example.bp.package",
            "Bundle-Version: 1.0.0",
            "Bundle-Vendor: Example Corp",
            "",
        ]
    )
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("META-INF/MANIFEST.MF", manifest)
        archive.writestr(
            "src/main/resources/scenarioflows/integrationflow/main.iflw",
            iflow_xml,
        )
        archive.writestr(
            "src/main/resources/wsdl/businessPartner.wsdl",
            "<definitions />\n",
        )


def _write_we60_html_bundle(path: Path) -> None:
    path.write_text(
        """
<html>
  <head><title>WE60 Documentation for IDoc Type DEBMAS07</title></head>
  <body>
    <h1>IDoc type DEBMAS07</h1>
    <table>
      <tr>
        <th>Segment type</th>
        <th>Description</th>
        <th>Parent segment</th>
      </tr>
      <tr>
        <td>EDI_DC40</td>
        <td>Control record</td>
        <td></td>
      </tr>
      <tr>
        <td>E1KNA1M</td>
        <td>Customer master general data</td>
        <td></td>
      </tr>
    </table>
    <ul>
      <li><a href="EDI_DC40.html">EDI_DC40</a></li>
      <li><a href="E1KNA1M.html">E1KNA1M</a></li>
    </ul>
  </body>
</html>
""".strip()
        + "\n",
        encoding="utf-8",
    )
    (path.parent / "EDI_DC40.html").write_text(
        """
<html>
  <head><title>Segment EDI_DC40</title></head>
  <body>
    <table>
      <tr>
        <th>Field name</th>
        <th>Description</th>
        <th>Data type</th>
        <th>Length</th>
      </tr>
      <tr>
        <td>IDOCTYP</td>
        <td>Basic type</td>
        <td>CHAR</td>
        <td>30</td>
      </tr>
    </table>
  </body>
</html>
""".strip()
        + "\n",
        encoding="utf-8",
    )
    (path.parent / "E1KNA1M.html").write_text(
        """
<html>
  <head><title>Segment E1KNA1M</title></head>
  <body>
    <table>
      <tr>
        <th>Field name</th>
        <th>Description</th>
        <th>Data type</th>
        <th>Length</th>
      </tr>
      <tr>
        <td>KUNNR</td>
        <td>Customer number</td>
        <td>CHAR</td>
        <td>10</td>
      </tr>
    </table>
  </body>
</html>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_openapi(path: Path) -> None:
    path.write_text(
        """
openapi: 3.0.3
info:
  title: Customer API
  version: 2026-07
servers:
  - url: https://api.example.com/customer
paths:
  /customers:
    get:
      operationId: listCustomers
      summary: List customers
      responses:
        "200":
          description: Customer list
          content:
            application/json:
              schema:
                $ref: "#/components/schemas/Customer"
    post:
      operationId: createCustomer
      summary: Create customer
      requestBody:
        required: true
        content:
          application/json:
            schema:
              $ref: "#/components/schemas/Customer"
      responses:
        "201":
          description: Customer created
          content:
            application/json:
              schema:
                $ref: "#/components/schemas/Customer"
components:
  schemas:
    Customer:
      type: object
      required: [id]
      properties:
        id:
          type: string
        email:
          type: string
          format: email
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_edmx(path: Path) -> None:
    path.write_text(
        """
<edmx:Edmx Version="4.0"
    xmlns:edmx="http://docs.oasis-open.org/odata/ns/edmx"
    xmlns:edm="http://docs.oasis-open.org/odata/ns/edm">
  <edmx:DataServices>
    <edm:Schema Namespace="API_BUSINESS_PARTNER">
      <edm:EntityType Name="A_BusinessPartner">
        <edm:Key>
          <edm:PropertyRef Name="BusinessPartner"/>
        </edm:Key>
        <edm:Property Name="BusinessPartner" Type="Edm.String" Nullable="false" />
        <edm:Property Name="BusinessPartnerCategory" Type="Edm.String" Nullable="true" />
      </edm:EntityType>
    </edm:Schema>
  </edmx:DataServices>
</edmx:Edmx>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_xsd(path: Path) -> None:
    path.write_text(
        """
<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"
    targetNamespace="urn:customer" version="1.0">
  <xs:simpleType name="StatusType">
    <xs:restriction base="xs:string">
      <xs:enumeration value="active" />
      <xs:enumeration value="blocked" />
    </xs:restriction>
  </xs:simpleType>
  <xs:complexType name="CustomerType">
    <xs:sequence>
      <xs:element name="CustomerId" type="xs:string" minOccurs="1" />
      <xs:element name="Status" type="StatusType" minOccurs="0" />
    </xs:sequence>
  </xs:complexType>
  <xs:element name="CustomerMessage" type="CustomerType" />
</xs:schema>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_wsdl(path: Path) -> None:
    path.write_text(
        """
<wsdl:definitions name="CustomerService"
    targetNamespace="urn:customer:service"
    xmlns:wsdl="http://schemas.xmlsoap.org/wsdl/"
    xmlns:xs="http://www.w3.org/2001/XMLSchema"
    xmlns:soap="http://schemas.xmlsoap.org/wsdl/soap/"
    xmlns:tns="urn:customer:service">
  <wsdl:types>
    <xs:schema targetNamespace="urn:customer:service">
      <xs:element name="CreateCustomerRequest">
        <xs:complexType>
          <xs:sequence>
            <xs:element name="CustomerId" type="xs:string" minOccurs="1" />
            <xs:element name="Status" minOccurs="0">
              <xs:simpleType>
                <xs:restriction base="xs:string">
                  <xs:enumeration value="active" />
                  <xs:enumeration value="blocked" />
                </xs:restriction>
              </xs:simpleType>
            </xs:element>
          </xs:sequence>
        </xs:complexType>
      </xs:element>
      <xs:element name="CreateCustomerResponse">
        <xs:complexType>
          <xs:sequence>
            <xs:element name="BusinessPartner" type="xs:string" minOccurs="1" />
          </xs:sequence>
        </xs:complexType>
      </xs:element>
    </xs:schema>
  </wsdl:types>
  <wsdl:message name="CreateCustomerInput">
    <wsdl:part name="parameters" element="tns:CreateCustomerRequest" />
  </wsdl:message>
  <wsdl:message name="CreateCustomerOutput">
    <wsdl:part name="parameters" element="tns:CreateCustomerResponse" />
  </wsdl:message>
  <wsdl:portType name="CustomerPortType">
    <wsdl:operation name="CreateCustomer">
      <wsdl:input message="tns:CreateCustomerInput" />
      <wsdl:output message="tns:CreateCustomerOutput" />
    </wsdl:operation>
  </wsdl:portType>
  <wsdl:binding name="CustomerBinding" type="tns:CustomerPortType">
    <soap:binding style="document" transport="http://schemas.xmlsoap.org/soap/http" />
    <wsdl:operation name="CreateCustomer">
      <soap:operation soapAction="urn:createCustomer" />
    </wsdl:operation>
  </wsdl:binding>
</wsdl:definitions>
""".strip()
        + "\n",
        encoding="utf-8",
    )


def test_inspect_to_proposal_json_schema_creates_reviewable_operations(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer.schema.json"
    _write_json_schema(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.proposal["type"] == "PatchProposal"
    assert result.proposal["id"].startswith("PP-SCHEMA-")
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types
    assert "ValueList" in object_types
    assert "customer.schema.json" in result.proposal["source_evidence"].lower()


def test_inspect_to_proposal_json_payload_creates_reviewable_objects(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-payload.json"
    _write_json_payload(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "json_payload"
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types


def test_inspect_to_proposal_xml_payload_creates_reviewable_objects(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-payload.xml"
    _write_xml_payload(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "xml_payload"
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types


def test_inspect_to_proposal_idoc_payload_creates_reviewable_objects(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-idoc.xml"
    _write_idoc_payload(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "idoc_payload"
    assert result.proposal["validation_status"] == "valid"
    endpoint_types = {
        op["after"]["endpoint_type"]
        for op in result.proposal["operations"]
        if op["object_type"] == "FieldEndpoint"
    }
    assert "idoc_segment_field" in endpoint_types
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types


def test_inspect_to_proposal_cds_metadata_creates_reviewable_objects(tmp_path: Path) -> None:
    schema_path = tmp_path / "cds-metadata.json"
    _write_cds_metadata_json(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "cds_metadata_export"
    assert result.proposal["validation_status"] == "valid"
    endpoint_types = {
        op["after"]["endpoint_type"]
        for op in result.proposal["operations"]
        if op["object_type"] == "FieldEndpoint"
    }
    assert "cds_metadata_field" in endpoint_types
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types


def test_inspect_to_proposal_iflow_artifact_creates_interface_objects(tmp_path: Path) -> None:
    schema_path = tmp_path / "main.iflw"
    _write_iflow_artifact(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "integration_flow_artifact"
    assert result.proposal["validation_status"] == "valid"
    object_types = [op["object_type"] for op in result.proposal["operations"]]
    assert "BusinessEntity" in object_types
    assert "Interface" in object_types
    assert "InterfaceEndpoint" in object_types
    assert "MessageType" in object_types


def test_inspect_to_proposal_integration_suite_zip_creates_interface_and_resource_objects(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "bp-sync-package.zip"
    _write_integration_suite_zip(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "integration_suite_package"
    assert result.proposal["validation_status"] == "valid"
    endpoint_types = {
        op["after"]["endpoint_type"]
        for op in result.proposal["operations"]
        if op["object_type"] == "FieldEndpoint"
    }
    assert "schema_field" not in endpoint_types
    assert "Interface" in [op["object_type"] for op in result.proposal["operations"]]
    assert "InterfaceEndpoint" in [op["object_type"] for op in result.proposal["operations"]]


def test_inspect_to_proposal_we60_html_creates_reviewable_objects(tmp_path: Path) -> None:
    schema_path = tmp_path / "DEBMAS07.html"
    _write_we60_html_bundle(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "we60_html_documentation"
    assert result.proposal["validation_status"] == "valid"
    endpoint_types = {
        op["after"]["endpoint_type"]
        for op in result.proposal["operations"]
        if op["object_type"] == "FieldEndpoint"
    }
    assert "we60_documentation_field" in endpoint_types
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types


def test_inspect_to_proposal_csv_field_catalog_creates_reviewable_objects(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "customer-field-catalog.csv"
    _write_field_catalog_csv(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "field_catalog"
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types
    assert "ValueList" in object_types


def test_inspect_to_proposal_xlsx_field_catalog_creates_reviewable_objects(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "customer-field-catalog.xlsx"
    _write_field_catalog_xlsx(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "field_catalog"
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types
    assert "ValueList" in object_types


def test_inspect_to_proposal_sap_mapping_workbook_creates_reviewable_objects(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "sap-customer-mapping.xlsx"
    _write_sap_mapping_workbook(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "sap_mapping_workbook"
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types


def test_inspect_to_proposal_migration_cockpit_workbook_creates_reviewable_objects(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "migration-cockpit-template.xlsx"
    _write_migration_cockpit_workbook(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "sap_migration_cockpit_template"
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types
    assert "ValueList" in object_types


def test_inspect_to_proposal_openapi_creates_interface_objects(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer-api.yaml"
    _write_openapi(schema_path)

    result = inspect_to_proposal(schema_path)

    object_types = [op["object_type"] for op in result.proposal["operations"]]
    assert "Interface" in object_types
    assert "InterfaceEndpoint" in object_types
    assert "BusinessEntity" in object_types
    assert "MessageType" in object_types
    assert "SchemaNode" in object_types
    assert result.proposal["affected_objects"]

    message_type_ids = {
        op["object_id"]
        for op in result.proposal["operations"]
        if op["object_type"] == "MessageType"
    }
    endpoint_ops = [
        op for op in result.proposal["operations"] if op["object_type"] == "InterfaceEndpoint"
    ]
    assert endpoint_ops
    create_customer = next(
        op
        for op in endpoint_ops
        if op["after"]["method"] == "POST" and op["after"]["path"] == "/customers"
    )
    assert create_customer["after"]["request_message_type"] in message_type_ids
    assert set(create_customer["after"]["response_message_types"]) <= message_type_ids
    assert create_customer["after"]["request_message_type"]
    assert create_customer["after"]["response_message_types"]
    assert create_customer["after"]["message_exchange_pattern"] == "request_response"


def test_cli_schema_import_preview_json_output(tmp_path: Path) -> None:
    schema_path = tmp_path / "customer.schema.json"
    _write_json_schema(schema_path)

    result = runner.invoke(app, ["schema", "import", str(schema_path), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["proposal_id"].startswith("PP-SCHEMA-")
    assert payload["proposal_path"] is None
    assert payload["validation_status"] == "valid"
    assert payload["proposal"]["affected_objects"]
    assert payload["proposal"]["operations"]


def test_cli_schema_import_as_proposal_writes_artifact(sample_repo: Path, tmp_path: Path) -> None:
    schema_path = tmp_path / "customer.schema.json"
    _write_json_schema(schema_path)

    result = runner.invoke(
        app,
        [
            "schema",
            "import",
            str(schema_path),
            "--repo",
            str(sample_repo),
            "--as-proposal",
            "--json",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    proposal_path = Path(payload["proposal_path"])
    assert proposal_path.exists()
    assert proposal_path.read_text(encoding="utf-8").startswith("---\n")
    service = SourceRegistryService(sample_repo)
    latest = service.get_latest_by_id(payload["source_id"])
    assert latest is not None
    assert latest.source_type == "schema_import"
    assert latest.file_hash == payload["checksum"]
    assert latest.metadata["source_format"] == "json_schema"
    assert latest.metadata["source_version"] == "https://json-schema.org/draft/2020-12/schema"
    assert latest.metadata["namespace"] == "https://example.com/customer.schema.json"
    assert latest.metadata["parser_version"]
    assert latest.metadata["retrieved_at"]
    assert latest.metadata["source_url"] is None
    assert latest.metadata["license_note"] is None
    assert latest.metadata["usage_note"] is None


def test_cli_schema_import_as_proposal_preserves_optional_provenance(
    sample_repo: Path, tmp_path: Path
) -> None:
    schema_path = tmp_path / "customer-api.yaml"
    _write_openapi(schema_path)

    result = runner.invoke(
        app,
        [
            "schema",
            "import",
            str(schema_path),
            "--repo",
            str(sample_repo),
            "--as-proposal",
            "--source-url",
            "https://api.example.com/docs/customer-api",
            "--license-note",
            "Example partner API terms apply.",
            "--usage-note",
            "Imported from a local export approved for internal modeling only.",
            "--json",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    service = SourceRegistryService(sample_repo)
    latest = service.get_latest_by_id(payload["source_id"])
    assert latest is not None
    assert latest.metadata["source_format"] == "openapi"
    assert latest.metadata["source_version"] == "3.0.3"
    assert latest.metadata["namespace"] == "https://api.example.com/customer"
    assert latest.metadata["source_url"] == "https://api.example.com/docs/customer-api"
    assert latest.metadata["license_note"] == "Example partner API terms apply."
    assert (
        latest.metadata["usage_note"]
        == "Imported from a local export approved for internal modeling only."
    )


def test_inspect_to_proposal_edmx_creates_entity_attribute_and_field_endpoint(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "api_business_partner.edmx"
    _write_edmx(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "edmx"
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types


def test_inspect_to_proposal_xsd_creates_entity_attribute_and_value_list(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "customer.xsd"
    _write_xsd(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "xsd"
    assert result.proposal["validation_status"] == "valid"
    object_types = {op["object_type"] for op in result.proposal["operations"]}
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types
    assert "ValueList" in object_types


def test_inspect_to_proposal_wsdl_creates_interface_and_message_objects(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "customer-service.wsdl"
    _write_wsdl(schema_path)

    result = inspect_to_proposal(schema_path)

    assert result.inspected.source_format == "wsdl"
    assert result.proposal["validation_status"] == "valid"
    object_types = [op["object_type"] for op in result.proposal["operations"]]
    assert "Interface" in object_types
    assert "InterfaceEndpoint" in object_types
    assert "BusinessEntity" in object_types
    assert "MessageType" in object_types
    assert "SchemaNode" in object_types
    assert "ValueList" in object_types
    assert result.proposal["affected_objects"]

    message_type_ids = {
        op["object_id"]
        for op in result.proposal["operations"]
        if op["object_type"] == "MessageType"
    }
    endpoint_ops = [
        op for op in result.proposal["operations"] if op["object_type"] == "InterfaceEndpoint"
    ]
    assert endpoint_ops
    assert endpoint_ops[0]["after"]["method"] == "CALL"
    assert endpoint_ops[0]["after"]["request_message_type"] in message_type_ids
    assert set(endpoint_ops[0]["after"]["response_message_types"]) <= message_type_ids
    assert endpoint_ops[0]["after"]["request_message_type"]
    assert endpoint_ops[0]["after"]["response_message_types"]
    assert endpoint_ops[0]["after"]["message_exchange_pattern"] == "request_response"


def test_inspect_to_proposal_json_payload_creates_message_and_schema_node_objects(
    tmp_path: Path,
) -> None:
    schema_path = tmp_path / "customer-payload.json"
    _write_json_payload(schema_path)

    result = inspect_to_proposal(schema_path)

    object_types = [op["object_type"] for op in result.proposal["operations"]]
    assert "BusinessEntity" in object_types
    assert "Attribute" in object_types
    assert "FieldEndpoint" in object_types
    assert "MessageType" in object_types
    assert "SchemaNode" in object_types


def test_schema_import_proposal_apply_builds_traceable_interface_lineage(
    sample_repo: Path, tmp_path: Path
) -> None:
    schema_path = tmp_path / "customer-api.yaml"
    _write_openapi(schema_path)

    result = inspect_to_proposal(schema_path)
    proposal_path = write_schema_proposal(result, repo_model_path=sample_repo / "model")
    transition_patch_proposal_status(proposal_path, "accepted", reviewer="alice")

    apply_result = apply_patch_proposal(
        sample_repo / "model",
        result.proposal["id"],
        skip_risk_check=True,
    )

    assert apply_result.application_status == "applied"
    assert apply_result.index_rebuilt

    endpoint_id = next(
        op["object_id"]
        for op in result.proposal["operations"]
        if op["object_type"] == "InterfaceEndpoint" and op["after"].get("method") == "POST"
    )
    trace_result = trace_object(
        resolve_generated_path(sample_repo) / "modelops.db",
        endpoint_id,
        max_depth=4,
        direction="both",
    )

    node_types = {node.object_type for node in trace_result.nodes}
    assert "MessageType" in node_types
    assert "SchemaNode" in node_types
    assert "FieldEndpoint" in node_types
