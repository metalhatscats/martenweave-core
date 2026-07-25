"""Golden scenario-lab checks for evidence-first SAP migration workflows."""

from __future__ import annotations

from pathlib import Path

import pytest

from modelops_core.evidence_ingestion import ingest_evidence
from modelops_core.pilot.preflight import inspect_file


def _multi_domain_mapping(path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = openpyxl.Workbook()
    bp = workbook.active
    bp.title = "BP mapping"
    bp.append(["Legacy BP field", "S/4 BP field", "Rule", "Owner"])
    bp.append(["BUT000.PARTNER", "BUT000.PARTNER", "Direct", "MDM Lead"])
    bp["B2"].hyperlink = "https://example.com/business-partner"
    bp.column_dimensions["D"].hidden = True
    bp.append([None, None, None, None])
    bp.merge_cells("A4:D4")
    bp["A2"].comment = openpyxl.comments.Comment("Confirm survivor rule", "SAP consultant")
    customer = workbook.create_sheet("Customer mapping")
    customer.append(["Source_Column", "Target Field", "Transform Rule"])
    customer.append(["KDGRP_LEGACY", "KNVV.KDGRP", "=UPPER(A2)"])
    customer.row_dimensions[2].hidden = True
    supplier = workbook.create_sheet("Supplier mapping")
    supplier.append(["Old Field", "New Field", "Scope"])
    supplier.append(["LFA1.KTOKK", "LFA1.KTOKK", "In scope"])
    supplier_table = Table(displayName="SupplierScope", ref="A1:C2")
    supplier_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    supplier.add_table(supplier_table)
    reference = workbook.create_sheet("Obsolete values")
    reference.sheet_state = "hidden"
    reference.append(["Obsolete code", "Disposition"])
    reference.append(["Z9", "Out of scope"])
    workbook.defined_names.add(
        DefinedName("BP_SCOPE", attr_text="'BP mapping'!$A$1:$D$2")
    )
    workbook.save(path)


def test_scenario_lab_interprets_multisheet_sap_mapping(tmp_path: Path) -> None:
    mapping = tmp_path / "bp-customer-supplier.xlsx"
    _multi_domain_mapping(mapping)

    inspection = inspect_file(mapping)

    assert inspection["status"] == "warning"
    assert [sheet["name"] for sheet in inspection["sheets"]] == [
        "BP mapping",
        "Customer mapping",
        "Supplier mapping",
        "Obsolete values",
    ]
    assert inspection["sheets"][0]["columns"] == [
        "Legacy BP field",
        "S/4 BP field",
        "Rule",
        "Owner",
    ]
    assert inspection["sheets"][0]["hidden_column_count"] == 1
    assert inspection["sheets"][0]["hidden_columns"] == ["D"]
    assert inspection["sheets"][0]["hyperlink_count"] == 1
    assert inspection["sheets"][0]["hyperlink_cells"] == ["B2"]
    assert inspection["sheets"][1]["hidden_row_count"] == 1
    assert inspection["sheets"][1]["hidden_rows"] == [2]
    assert inspection["sheets"][2]["tables"] == [
        {"sheet": "Supplier mapping", "name": "SupplierScope", "ref": "A1:C2"}
    ]
    assert inspection["sheets"][-1]["included"] is False
    assert inspection["formula_count"] == 1
    assert inspection["comment_count"] == 1
    assert inspection["hyperlink_count"] == 1
    assert inspection["hyperlink_cells"] == {"BP mapping": ["B2"]}
    assert inspection["hidden_rows"] == {"Customer mapping": [2]}
    assert inspection["hidden_columns"] == {"BP mapping": ["D"]}
    assert inspection["defined_names"] == [
        {"name": "BP_SCOPE", "target": "'BP mapping'!$A$1:$D$2", "type": "RANGE"}
    ]
    assert inspection["excel_tables"] == [
        {"sheet": "Supplier mapping", "name": "SupplierScope", "ref": "A1:C2"}
    ]
    assert inspection["merged_ranges"] == {"BP mapping": ["A4:D4"]}
    assert any("Hidden sheet" in warning for warning in inspection["warnings"])
    assert any("Hidden row" in warning for warning in inspection["warnings"])
    assert any("Hidden column" in warning for warning in inspection["warnings"])
    assert any("hyperlink" in warning.lower() for warning in inspection["warnings"])
    assert any("defined name" in warning.lower() for warning in inspection["warnings"])
    assert any("Excel table" in warning for warning in inspection["warnings"])


def test_scenario_lab_repeated_evidence_ingestion_is_deterministic(
    sample_repo: Path, tmp_path: Path
) -> None:
    evidence = tmp_path / "migration-findings.md"
    evidence.write_text(
        "- Missing owner for ATTR-CUST-SALES-CUSTOMER-GROUP\n"
        "- Missing mapping for FEP-S4-KNVV-KDGRP\n"
        "- Decision needed for obsolete source field\n",
        encoding="utf-8",
    )

    first = ingest_evidence(evidence, sample_repo / "model")
    second = ingest_evidence(evidence, sample_repo / "model")

    assert first.source_sha256 == second.source_sha256
    assert first.proposal["id"] == second.proposal["id"]
    assert [op["object_id"] for op in first.proposal["operations"]] == [
        op["object_id"] for op in second.proposal["operations"]
    ]
    assert first.finding_count == 3
