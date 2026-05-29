"""Tests for model export service and CLI."""

from __future__ import annotations

from pathlib import Path

import pytest
from typer.testing import CliRunner

from modelops_core.cli import app
from modelops_core.exports import export_model_csv, export_model_jsonl, export_model_xlsx

runner = CliRunner()


def test_export_model_csv(temp_model_dir: Path) -> None:
    result = export_model_csv(temp_model_dir)
    assert len(result) > 0
    for f in result:
        assert f.exists()
        assert f.suffix == ".csv"


def test_export_model_csv_columns(temp_model_dir: Path) -> None:
    export_model_csv(temp_model_dir)
    csv_dir = temp_model_dir.parent / "generated" / "exports" / "csv"
    domain_csv = csv_dir / "masterdatadomain.csv"
    assert domain_csv.exists()
    content = domain_csv.read_text()
    assert "id" in content
    assert "type" in content
    assert "status" in content
    assert "DOMAIN-TEST" in content


def test_export_model_xlsx(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    path = export_model_xlsx(temp_model_dir)
    assert path.exists()
    assert path.suffix == ".xlsx"


def test_export_model_xlsx_sheets(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    path = export_model_xlsx(temp_model_dir)
    wb = load_workbook(path)
    sheet_names = wb.sheetnames
    assert len(sheet_names) > 0
    assert any("masterdatadomain" in s.lower() for s in sheet_names)
    wb.close()


def test_cli_export_model_csv(temp_model_dir: Path) -> None:
    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["export-model", "--repo", repo, "--format", "csv"])
    assert result.exit_code == 0
    assert "Exported" in result.output
    assert ".csv" in result.output


def test_cli_export_model_xlsx(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["export-model", "--repo", repo, "--format", "xlsx"])
    assert result.exit_code == 0
    assert "Exported" in result.output
    assert ".xlsx" in result.output


def test_export_model_jsonl(temp_model_dir: Path) -> None:
    result = export_model_jsonl(temp_model_dir)
    assert len(result) > 0
    for f in result:
        assert f.exists()
        assert f.suffix == ".jsonl"


def test_export_model_jsonl_content(temp_model_dir: Path) -> None:
    import json

    export_model_jsonl(temp_model_dir)
    jsonl_dir = temp_model_dir.parent / "generated" / "exports" / "jsonl"
    domain_jsonl = jsonl_dir / "masterdatadomain.jsonl"
    assert domain_jsonl.exists()
    lines = domain_jsonl.read_text().strip().split("\n")
    assert len(lines) >= 1
    obj = json.loads(lines[0])
    assert "id" in obj
    assert "type" in obj
    assert "status" in obj
    assert "source_file" in obj


def test_cli_export_model_json(temp_model_dir: Path) -> None:
    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["export-model", "--repo", repo, "--format", "json"])
    assert result.exit_code == 0
    assert "Exported" in result.output
    assert ".jsonl" in result.output


def test_cli_export_model_unknown_format(temp_model_dir: Path) -> None:
    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["export-model", "--repo", repo, "--format", "pdf"])
    assert result.exit_code == 1
    assert "Unknown format" in result.output


def test_export_model_csv_includes_valuelist(sample_repo: Path) -> None:
    model_path = sample_repo / "model"
    export_model_csv(model_path)
    csv_dir = sample_repo / "generated" / "exports" / "csv"
    vl_csv = csv_dir / "valuelist.csv"
    assert vl_csv.exists()
    content = vl_csv.read_text()
    assert "VLIST-LEGACY-CUST-GROUP" in content
    assert "VLIST-S4-CUST-GROUP" in content
    assert "entries" in content


def test_export_model_csv_includes_valuemapping(sample_repo: Path) -> None:
    model_path = sample_repo / "model"
    export_model_csv(model_path)
    csv_dir = sample_repo / "generated" / "exports" / "csv"
    vm_csv = csv_dir / "valuemapping.csv"
    assert vm_csv.exists()
    content = vm_csv.read_text()
    assert "VMAP-CUST-GROUP-LEGACY-TO-S4" in content
    assert "source_value_list" in content
    assert "target_value_list" in content


def test_export_model_xlsx_includes_lov_sheets(sample_repo: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    model_path = sample_repo / "model"
    path = export_model_xlsx(model_path)
    wb = load_workbook(path)
    sheet_names = [s.lower() for s in wb.sheetnames]
    assert "valuelist" in sheet_names
    assert "valuemapping" in sheet_names
    wb.close()


# ---------------------------------------------------------------------------
# Business-review export tests (#48)
# ---------------------------------------------------------------------------


def test_export_model_xlsx_business_review_has_readme(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    path = export_model_xlsx(temp_model_dir, business_review=True)
    wb = load_workbook(path)
    sheet_names = [s.lower() for s in wb.sheetnames]
    assert "read me" in sheet_names
    wb.close()


def test_export_model_xlsx_business_review_has_reviewer_notes(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    path = export_model_xlsx(temp_model_dir, business_review=True)
    wb = load_workbook(path)
    sheet_name = next(s for s in wb.sheetnames if s.lower() == "masterdatadomain")
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    assert "reviewer_notes" in headers
    wb.close()


def test_export_model_xlsx_business_review_styled_headers(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    path = export_model_xlsx(temp_model_dir, business_review=True)
    wb = load_workbook(path)
    sheet_name = next(s for s in wb.sheetnames if s.lower() == "masterdatadomain")
    ws = wb[sheet_name]
    header_cell = ws.cell(row=1, column=1)
    assert header_cell.font.bold is True
    wb.close()


def test_cli_export_model_xlsx_business_review(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    repo = str(temp_model_dir.parent)
    result = runner.invoke(
        app, ["export-model", "--repo", repo, "--format", "xlsx", "--business-review"]
    )
    assert result.exit_code == 0
    assert "business-review" in result.output
    assert ".xlsx" in result.output


# --json flag tests -----------------------------------------------------------


def test_cli_export_model_csv_json(temp_model_dir: Path) -> None:
    import json

    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["export-model", "--repo", repo, "--format", "csv", "--json"])
    assert result.exit_code == 0
    data = json.loads(result.output)
    assert data["format"] == "csv"
    assert isinstance(data["files"], list)
    assert len(data["files"]) > 0
    assert data["business_review"] is False


def test_cli_export_model_xlsx_json(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    import json

    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["export-model", "--repo", repo, "--format", "xlsx", "--json"])
    assert result.exit_code == 0
    data = json.loads(result.output)
    assert data["format"] == "xlsx"
    assert "file" in data
    assert data["business_review"] is False


def test_cli_export_model_json_json(temp_model_dir: Path) -> None:
    import json

    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["export-model", "--repo", repo, "--format", "json", "--json"])
    assert result.exit_code == 0
    data = json.loads(result.output)
    assert data["format"] == "json"
    assert isinstance(data["files"], list)
    assert len(data["files"]) > 0
    assert data["business_review"] is False


def test_cli_export_model_unknown_format_json(temp_model_dir: Path) -> None:
    import json

    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["export-model", "--repo", repo, "--format", "pdf", "--json"])
    assert result.exit_code == 1
    data = json.loads(result.output)
    assert "error" in data
