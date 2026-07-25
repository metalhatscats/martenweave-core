"""End-to-end regression corpus for irregular SAP mapping workbooks.

The workbooks are generated rather than checked in so the corpus stays small,
synthetic, and easy to extend.  Each scenario represents a structure commonly
seen in mapping workbooks received from migration teams.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from modelops_core.imports.dataset_profiler import profile_xlsx
from modelops_core.pilot.bootstrap import bootstrap_assessment
from modelops_core.pilot.preflight import run_preflight
from modelops_core.run.migration_assessment import (
    _profile_mapping_workbook,
    generate_migration_assessment,
)

SCENARIOS = (
    "shifted_headers",
    "merged_title",
    "hidden_sheet",
    "formula_cells",
    "multiple_tables",
    "multilingual_headers",
    "inconsistent_names",
    "colour_status",
    "large_sheet",
    "split_source_target",
)


def _write_workbook(path: Path, scenario: str) -> None:
    """Write one synthetic workbook with a deliberately awkward structure."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Customer Mapping"

    headers = ["Source Field", "Source Table", "Target Table", "Target Field", "Owner", "Status"]
    row = ["CUSTOMER_GROUP", "KNA1", "KNVV", "KDGRP", "", "active"]

    if scenario == "multilingual_headers":
        headers = [
            "Поле источника",
            "Таблица источника",
            "Таблица цели",
            "Поле цели",
            "Владелец",
            "Статус",
        ]
    elif scenario == "inconsistent_names":
        headers = ["Legacy field", "Source table", "SAP table", "New field", "Responsible", "State"]

    if scenario in {"shifted_headers", "merged_title"}:
        sheet.append(["Customer to S/4HANA Mapping"])
        if scenario == "merged_title":
            sheet.merge_cells("A1:F1")
    sheet.append(headers)
    sheet.append(row)

    if scenario == "hidden_sheet":
        hidden = workbook.create_sheet("Legacy notes")
        hidden.append(["Source Field", "Target Field", "Owner"])
        hidden.append(["SHOULD_NOT_BE_READ", "NOPE", ""])
        hidden.sheet_state = "hidden"
    elif scenario == "formula_cells":
        sheet.cell(sheet.max_row, 6).value = '="active"'
    elif scenario == "multiple_tables":
        sheet.append([])
        sheet.append(["Supplier mapping"])
        sheet.append(headers)
        sheet.append(["SUPPLIER_ID", "LFA1", "LFA1", "LIFNR", "supplier.team", "active"])
    elif scenario == "colour_status":
        status_cell = sheet.cell(sheet.max_row, 6)
        status_cell.value = ""
        status_cell.fill = PatternFill(fill_type="solid", fgColor="92D050")
    elif scenario == "large_sheet":
        for index in range(5_100):
            sheet.append([f"CUSTOMER_{index}", "KNA1", "KNVV", "KDGRP", "migration.team", "active"])
    elif scenario == "split_source_target":
        sheet.title = "Legacy extract"
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(["Source Field", "Source Table", "Owner"])
        sheet.append(["CUSTOMER_GROUP", "KNA1", "migration.team"])
        target = workbook.create_sheet("S4 target fields")
        target.append(["Target Table", "Target Field", "Owner"])
        target.append(["KNVV", "KDGRP", "migration.team"])

    workbook.save(path)
    workbook.close()


@pytest.mark.parametrize("scenario", SCENARIOS)
def test_irregular_workbook_corpus_completes_pilot_path(
    sample_repo: Path, tmp_path: Path, scenario: str
) -> None:
    """Preflight, proposal-only bootstrap, and assessment survive each layout."""
    workbook = tmp_path / f"{scenario}.xlsx"
    _write_workbook(workbook, scenario)

    preflight = run_preflight(workbook, [], [], [], tmp_path / "preflight")
    assert preflight.overall_status in {"allowed", "warning"}
    assert (tmp_path / "preflight" / "workbook_manifest.json").exists()
    if scenario == "colour_status":
        assert any("colour-only statuses" in warning for warning in preflight.files[0]["warnings"])

    profile = _profile_mapping_workbook(workbook)
    if scenario != "split_source_target":
        assert len(profile.missing_owner_rows) == 1
    if scenario == "hidden_sheet":
        assert all(row["sheet"] != "Legacy notes" for row in profile.missing_owner_rows)
    if scenario == "multiple_tables":
        assert profile.total_rows == 2
    if scenario == "large_sheet":
        assert profile.total_rows == 5_101

    bootstrap = bootstrap_assessment(
        workbook,
        f"{scenario} pilot",
        tmp_path / "bootstrap",
    )
    assert bootstrap.proposal_path.exists()
    profiled_sheets = profile_xlsx(workbook, "mapping").sheets
    if scenario in {"shifted_headers", "merged_title"}:
        assert profiled_sheets[0].header_row == 2
        assert profiled_sheets[0].columns[0].name == "Source Field"
    if scenario == "hidden_sheet":
        assert all(sheet.sheet_name != "Legacy notes" for sheet in profiled_sheets)

    assessment = generate_migration_assessment(
        sample_repo,
        workbook,
        None,
        [],
        tmp_path / "assessment",
    )
    assert all(stage.status != "failed" for stage in assessment.stage_statuses)
    assert (tmp_path / "assessment" / "findings.json").exists()
