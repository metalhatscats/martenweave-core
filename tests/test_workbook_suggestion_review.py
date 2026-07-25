"""Tests for governed workbook suggestion review workbook roundtrip."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from typer.testing import CliRunner

from modelops_core.cli import app
from modelops_core.pilot.structural_scan import scan_workbook_structure
from modelops_core.pilot.workbook_suggestion_review import (
    import_workbook_suggestion_review_xlsx,
    write_workbook_suggestion_feedback_json,
    write_workbook_suggestion_feedback_markdown,
    write_workbook_suggestion_review_xlsx,
)
from modelops_core.pilot.workbook_suggestions import generate_workbook_suggestions

runner = CliRunner()

MAPPING_WORKBOOK = Path(__file__).parent / "fixtures" / "pilot" / "sap_customer_mapping.xlsx"


@pytest.mark.skipif(not MAPPING_WORKBOOK.exists(), reason="pilot fixture missing")
def test_write_workbook_suggestion_review_xlsx_protects_identity_columns(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    suggestion_set = generate_workbook_suggestions(scan_workbook_structure(MAPPING_WORKBOOK))
    review_path = tmp_path / "workbook_suggestion_review.xlsx"

    write_workbook_suggestion_review_xlsx(suggestion_set, review_path)

    workbook = load_workbook(review_path)
    assert "Read Me" in workbook.sheetnames
    sheet = workbook["Suggestions"]
    headers = [cell.value for cell in sheet[1]]
    suggestion_id_col = headers.index("suggestion_id") + 1
    reviewer_decision_col = headers.index("reviewer_decision") + 1
    reviewer_notes_col = headers.index("reviewer_notes") + 1

    assert sheet.protection.sheet is True
    assert sheet.cell(2, suggestion_id_col).protection.locked is True
    assert sheet.cell(2, reviewer_decision_col).protection.locked is False
    assert sheet.cell(2, reviewer_notes_col).protection.locked is False
    assert sheet.data_validations.dataValidation
    workbook.close()


@pytest.mark.skipif(not MAPPING_WORKBOOK.exists(), reason="pilot fixture missing")
def test_import_workbook_suggestion_review_xlsx_roundtrip(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    suggestion_set = generate_workbook_suggestions(scan_workbook_structure(MAPPING_WORKBOOK))
    review_path = tmp_path / "workbook_suggestion_review.xlsx"
    feedback_dir = tmp_path / "feedback"

    write_workbook_suggestion_review_xlsx(suggestion_set, review_path)

    workbook = load_workbook(review_path)
    sheet = workbook["Suggestions"]
    headers = [cell.value for cell in sheet[1]]
    sheet.cell(2, headers.index("reviewer_decision") + 1).value = "accepted"
    sheet.cell(2, headers.index("reviewer_notes") + 1).value = "Correct mapping interpretation."
    sheet.cell(3, headers.index("reviewer_decision") + 1).value = "rejected"
    workbook.save(review_path)
    workbook.close()

    feedback = import_workbook_suggestion_review_xlsx(review_path)
    json_path = write_workbook_suggestion_feedback_json(
        feedback,
        feedback_dir / "workbook_suggestion_feedback.json",
    )
    markdown_path = write_workbook_suggestion_feedback_markdown(
        feedback,
        feedback_dir / "workbook_suggestion_feedback.md",
    )

    assert feedback.decision_counts()["accepted"] == 1
    assert feedback.decision_counts()["rejected"] == 1
    assert feedback.feedback_records[0].suggestion_id.startswith("WSUG-")
    assert json.loads(json_path.read_text(encoding="utf-8"))["decision_counts"]["accepted"] == 1
    assert "Workbook suggestion review feedback" in markdown_path.read_text(encoding="utf-8")


@pytest.mark.skipif(not MAPPING_WORKBOOK.exists(), reason="pilot fixture missing")
def test_cli_import_workbook_suggestion_review_writes_feedback_artifacts(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    suggestion_set = generate_workbook_suggestions(scan_workbook_structure(MAPPING_WORKBOOK))
    review_path = tmp_path / "workbook_suggestion_review.xlsx"
    out_dir = tmp_path / "feedback"

    write_workbook_suggestion_review_xlsx(suggestion_set, review_path)
    workbook = load_workbook(review_path)
    sheet = workbook["Suggestions"]
    headers = [cell.value for cell in sheet[1]]
    sheet.cell(2, headers.index("reviewer_decision") + 1).value = "accepted"
    workbook.save(review_path)
    workbook.close()

    result = runner.invoke(
        app,
        [
            "import-workbook-suggestion-review",
            "--from",
            str(review_path),
            "--out",
            str(out_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    assert (out_dir / "workbook_suggestion_feedback.json").exists()
    assert (out_dir / "workbook_suggestion_feedback.md").exists()
    assert "Canonical model files were not changed" in result.output
