"""Tests for import-model-sheet service and CLI."""

from __future__ import annotations

from pathlib import Path

import pytest
from typer.testing import CliRunner

from modelops_core.cli import app
from modelops_core.exports import export_model_csv, export_model_xlsx
from modelops_core.imports.model_sheet_import_service import (
    import_model_sheet_csv,
    import_model_sheet_xlsx,
)

runner = CliRunner()

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def test_import_csv_no_changes(temp_model_dir: Path) -> None:
    export_model_csv(temp_model_dir)
    csv_dir = temp_model_dir.parent / "generated" / "exports" / "csv"
    proposal = import_model_sheet_csv(csv_dir, temp_model_dir)
    assert proposal["id"].startswith("PP-IMPORT-")
    assert len(proposal["operations"]) == 0


def test_import_csv_detects_update(temp_model_dir: Path) -> None:
    export_model_csv(temp_model_dir)
    csv_dir = temp_model_dir.parent / "generated" / "exports" / "csv"

    # Modify a CSV file
    domain_csv = csv_dir / "masterdatadomain.csv"
    content = domain_csv.read_text()
    content = content.replace("Test Domain", "Test Domain Updated")
    domain_csv.write_text(content)

    proposal = import_model_sheet_csv(csv_dir, temp_model_dir)
    assert len(proposal["operations"]) == 1
    op = proposal["operations"][0]
    assert op["op"] == "update_object"
    assert op["object_id"] == "DOMAIN-TEST"
    assert op["target_path"] == "name"
    assert op["after"] == "Test Domain Updated"


def test_import_csv_detects_new_object(temp_model_dir: Path) -> None:
    export_model_csv(temp_model_dir)
    csv_dir = temp_model_dir.parent / "generated" / "exports" / "csv"

    # Append a new row to domain CSV
    domain_csv = csv_dir / "masterdatadomain.csv"
    with domain_csv.open("a", newline="", encoding="utf-8") as f:
        f.write("NEW-DOMAIN,MasterDataDomain,draft,New Domain,,,\n")

    proposal = import_model_sheet_csv(csv_dir, temp_model_dir)
    create_ops = [op for op in proposal["operations"] if op["op"] == "create_object"]
    assert len(create_ops) == 1
    assert create_ops[0]["object_id"] == "NEW-DOMAIN"


def test_import_csv_duplicate_id_warning(temp_model_dir: Path) -> None:
    export_model_csv(temp_model_dir)
    csv_dir = temp_model_dir.parent / "generated" / "exports" / "csv"

    # Append duplicate row
    domain_csv = csv_dir / "masterdatadomain.csv"
    with domain_csv.open("a", newline="", encoding="utf-8") as f:
        f.write("DOMAIN-TEST,MasterDataDomain,draft,Dup,,,\n")

    proposal = import_model_sheet_csv(csv_dir, temp_model_dir)
    assert any("Duplicate ID" in w for w in proposal.get("warnings", []))


def test_import_xlsx_detects_update(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    export_model_xlsx(temp_model_dir)
    xlsx_path = temp_model_dir.parent / "generated" / "exports" / "model.xlsx"

    wb = load_workbook(xlsx_path)
    # Find sheet case-insensitively
    sheet_name = next(s for s in wb.sheetnames if s.lower() == "masterdatadomain")
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "DOMAIN-TEST":
            row[3].value = "Test Domain Updated"
            break
    wb.save(xlsx_path)
    wb.close()

    proposal = import_model_sheet_xlsx(xlsx_path, temp_model_dir)
    update_ops = [op for op in proposal["operations"] if op["op"] == "update_object"]
    assert len(update_ops) == 1
    assert update_ops[0]["after"] == "Test Domain Updated"


def test_cli_import_model_sheet_csv(temp_model_dir: Path) -> None:
    export_model_csv(temp_model_dir)
    csv_dir = temp_model_dir.parent / "generated" / "exports" / "csv"
    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["import-model-sheet", str(csv_dir), "--repo", repo])
    assert result.exit_code == 0
    assert "PatchProposal" in result.output


def test_cli_import_model_sheet_xlsx(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    export_model_xlsx(temp_model_dir)
    xlsx_path = temp_model_dir.parent / "generated" / "exports" / "model.xlsx"
    repo = str(temp_model_dir.parent)
    result = runner.invoke(app, ["import-model-sheet", str(xlsx_path), "--repo", repo])
    assert result.exit_code == 0
    assert "PatchProposal" in result.output


def test_cli_import_model_sheet_invalid_input(temp_model_dir: Path) -> None:
    repo = str(temp_model_dir.parent)
    bad_path = temp_model_dir.parent / "not_a_file.txt"
    bad_path.write_text("nope")
    result = runner.invoke(app, ["import-model-sheet", str(bad_path), "--repo", repo])
    assert result.exit_code == 1
    assert "must be a CSV directory or an .xlsx workbook" in result.output


# ---------------------------------------------------------------------------
# Hardened Excel import tests (#47)
# ---------------------------------------------------------------------------


def test_import_xlsx_structured_workbook(sample_repo: Path) -> None:
    pytest.importorskip("openpyxl")
    xlsx_path = FIXTURES_DIR / "product_model.xlsx"
    proposal = import_model_sheet_xlsx(xlsx_path, sample_repo / "model")
    assert proposal["id"].startswith("PP-IMPORT-")
    create_ops = [op for op in proposal["operations"] if op["op"] == "create_object"]
    assert len(create_ops) >= 4  # domain, entity, attributes, field endpoints, value list
    assert any(op["object_id"] == "DOMAIN-PRODUCT" for op in create_ops)
    assert any(op["object_id"] == "ATTR-PRODUCT-ID" for op in create_ops)


def test_import_xlsx_broken_reference_warning(sample_repo: Path) -> None:
    pytest.importorskip("openpyxl")
    xlsx_path = FIXTURES_DIR / "product_model_broken_refs.xlsx"
    proposal = import_model_sheet_xlsx(xlsx_path, sample_repo / "model")
    assert any(
        "Broken reference" in w and "DOMAIN-NONEXISTENT" in w for w in proposal.get("warnings", [])
    )
    assert any(
        "Broken reference" in w and "ENTITY-NONEXISTENT" in w for w in proposal.get("warnings", [])
    )


def test_import_xlsx_duplicate_id_warning(sample_repo: Path) -> None:
    pytest.importorskip("openpyxl")
    xlsx_path = FIXTURES_DIR / "product_model_duplicate_ids.xlsx"
    proposal = import_model_sheet_xlsx(xlsx_path, sample_repo / "model")
    assert any("Duplicate ID" in w for w in proposal.get("warnings", []))


def test_import_xlsx_formula_warning(sample_repo: Path) -> None:
    pytest.importorskip("openpyxl")
    xlsx_path = FIXTURES_DIR / "product_model_with_formulas.xlsx"
    proposal = import_model_sheet_xlsx(xlsx_path, sample_repo / "model")
    assert any("Formula detected" in w for w in proposal.get("warnings", []))


def test_import_xlsx_unknown_sheet_type_warning(sample_repo: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("UnknownTypeSheet")
    ws.append(["id", "type", "status", "name"])
    ws.append(["OBJ-001", "UnknownType", "draft", "Test"])
    xlsx_path = sample_repo.parent / "unknown_type.xlsx"
    wb.save(xlsx_path)
    wb.close()

    proposal = import_model_sheet_xlsx(xlsx_path, sample_repo / "model")
    assert any("does not match a known object type" in w for w in proposal.get("warnings", []))


# ---------------------------------------------------------------------------
# Business-review roundtrip tests (#48)
# ---------------------------------------------------------------------------


def test_import_xlsx_business_review_roundtrip_no_changes(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")

    export_model_xlsx(temp_model_dir, business_review=True)
    xlsx_path = temp_model_dir.parent / "generated" / "exports" / "model.xlsx"

    proposal = import_model_sheet_xlsx(xlsx_path, temp_model_dir)
    assert proposal["id"].startswith("PP-IMPORT-")
    # reviewer_notes is a meta column and should not produce operations
    assert len(proposal["operations"]) == 0


def test_import_xlsx_business_review_roundtrip_detects_update(temp_model_dir: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    export_model_xlsx(temp_model_dir, business_review=True)
    xlsx_path = temp_model_dir.parent / "generated" / "exports" / "model.xlsx"

    wb = load_workbook(xlsx_path)
    sheet_name = next(s for s in wb.sheetnames if s.lower() == "masterdatadomain")
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "DOMAIN-TEST":
            row[3].value = "Test Domain Updated"
            break
    wb.save(xlsx_path)
    wb.close()

    proposal = import_model_sheet_xlsx(xlsx_path, temp_model_dir)
    update_ops = [op for op in proposal["operations"] if op["op"] == "update_object"]
    assert len(update_ops) == 1
    assert update_ops[0]["after"] == "Test Domain Updated"
    assert update_ops[0]["target_path"] == "name"
