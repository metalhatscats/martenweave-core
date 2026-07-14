"""Tests for the one-command SAP migration assessment workflow."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from typer.testing import CliRunner

from modelops_core.cli import app
from modelops_core.run.migration_assessment import (
    MappingWorkbookProfile,
    _profile_mapping_workbook,
    generate_migration_assessment,
)

runner = CliRunner()


def _write_minimal_mapping_workbook(path: Path) -> None:
    """Create a minimal mapping workbook for tests."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise pytest.skip("openpyxl not installed") from exc

    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover
        raise RuntimeError("openpyxl did not create an active worksheet")
    ws.title = "Customer_Mappings"
    ws.append(
        [
            "source_field",
            "source_system",
            "target_table",
            "target_field",
            "owner",
            "status",
            "notes",
        ]
    )
    ws.append(
        [
            "customer_group",
            "Legacy CRM",
            "KNVV",
            "KDGRP",
            "sales_data_team",
            "active",
            "Customer group mapping",
        ]
    )
    ws.append(
        [
            "customer_group",
            "Legacy CRM",
            "KNVV",
            "KDGRP",
            "",
            "active",
            "Duplicate row missing owner",
        ]
    )
    wb.save(path)
    wb.close()


def _write_mapping_workbook_with_formula(path: Path) -> None:
    """Create a mapping workbook containing a formula."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise pytest.skip("openpyxl not installed") from exc

    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover
        raise RuntimeError("openpyxl did not create an active worksheet")
    ws.title = "Mappings"
    ws.append(["source_field", "target_field", "owner"])
    ws.append(["field_a", "FIELD_A", "owner_a"])
    ws["D2"] = "=A2"
    wb.save(path)
    wb.close()


def test_profile_mapping_workbook_metadata(tmp_path: Path) -> None:
    mapping = tmp_path / "mapping.xlsx"
    _write_minimal_mapping_workbook(mapping)

    profile = _profile_mapping_workbook(mapping)

    assert isinstance(profile, MappingWorkbookProfile)
    assert profile.file_path == str(mapping)
    assert len(profile.file_hash) == 64
    assert profile.sheet_names == ["Customer_Mappings"]
    assert profile.total_rows == 2
    assert "source_field" in profile.column_names
    assert "target_field" in profile.column_names


def test_profile_mapping_workbook_detects_missing_owner(tmp_path: Path) -> None:
    mapping = tmp_path / "mapping.xlsx"
    _write_minimal_mapping_workbook(mapping)

    profile = _profile_mapping_workbook(mapping)

    assert len(profile.missing_owner_rows) == 1
    row = profile.missing_owner_rows[0]
    assert row["sheet"] == "Customer_Mappings"
    assert row["row"] == 3


def test_profile_mapping_workbook_detects_duplicates(tmp_path: Path) -> None:
    mapping = tmp_path / "mapping.xlsx"
    _write_minimal_mapping_workbook(mapping)

    profile = _profile_mapping_workbook(mapping)

    assert len(profile.duplicate_rows) == 1
    dup = profile.duplicate_rows[0]
    assert dup["sheet"] == "Customer_Mappings"
    assert dup["duplicate_of"]["row"] == 2


def test_profile_mapping_workbook_detects_formulas(tmp_path: Path) -> None:
    mapping = tmp_path / "mapping_with_formula.xlsx"
    _write_mapping_workbook_with_formula(mapping)

    profile = _profile_mapping_workbook(mapping)

    assert len(profile.formula_warnings) == 1
    assert "Formula in sheet 'Mappings' cell" in profile.formula_warnings[0]


def test_generate_migration_assessment_happy_path(
    sample_repo: Path, tmp_path: Path
) -> None:
    mapping = tmp_path / "mapping.xlsx"
    _write_minimal_mapping_workbook(mapping)
    out = tmp_path / "assessment"

    manifest = generate_migration_assessment(
        repo_root=sample_repo,
        mapping_path=mapping,
        dataset_path=None,
        evidence_paths=[],
        out_dir=out,
    )

    assert manifest.repo_name
    assert manifest.repo_path == str(sample_repo)
    assert (out / "manifest.json").exists()
    assert (out / "mapping_profile.json").exists()
    assert (out / "01_readiness_scorecard.md").exists()
    assert (out / "02_gap_report.md").exists()
    assert (out / "03_high_risk_fields.md").exists()
    assert (out / "04_impact_reports").exists()
    assert (out / "05_business_review.xlsx").exists()
    assert (out / "06_recommendations.md").exists()
    assert (out / "review_pack").exists()

    status_names = {s.name: s.status for s in manifest.stage_statuses}
    assert status_names["validation"] == "success"
    assert status_names["index"] == "success"
    assert status_names["mapping_profile"] == "success"
    assert status_names["dataset_readiness"] == "skipped"
    assert status_names["assessment_package"] == "success"
    assert status_names["review_pack"] == "success"

    # Manifest inputs
    manifest_data = json.loads((out / "manifest.json").read_text(encoding="utf-8"))
    assert manifest_data["inputs"]["mapping"] == str(mapping)
    assert manifest_data["inputs"]["dataset"] is None
    assert manifest_data["inputs"]["evidence"] == []
    assert any(
        a["path"] == "manifest.json" for a in manifest_data["generated_artifacts"]
    )


def test_generate_migration_assessment_with_dataset(
    sample_repo: Path, tmp_path: Path
) -> None:
    mapping = tmp_path / "mapping.xlsx"
    _write_minimal_mapping_workbook(mapping)
    dataset = sample_repo / "data" / "samples" / "customer_messy.csv"
    out = tmp_path / "assessment"

    manifest = generate_migration_assessment(
        repo_root=sample_repo,
        mapping_path=mapping,
        dataset_path=dataset,
        evidence_paths=[],
        out_dir=out,
    )

    status_names = {s.name: s.status for s in manifest.stage_statuses}
    assert status_names["dataset_readiness"] == "success"
    assert (out / "dataset_readiness" / "readiness.json").exists()
    assert (out / "dataset_readiness" / "readiness.md").exists()


def test_cli_migration_assessment_happy_path(sample_repo: Path, tmp_path: Path) -> None:
    mapping = tmp_path / "mapping.xlsx"
    _write_minimal_mapping_workbook(mapping)
    out = tmp_path / "assessment"

    result = runner.invoke(
        app,
        [
            "run",
            "migration-assessment",
            "--repo",
            str(sample_repo),
            "--mapping",
            str(mapping),
            "--out",
            str(out),
        ],
    )

    assert result.exit_code == 0, result.output
    assert "Migration assessment complete" in result.output
    assert (out / "manifest.json").exists()


def test_cli_migration_assessment_json_output(
    sample_repo: Path, tmp_path: Path
) -> None:
    mapping = tmp_path / "mapping.xlsx"
    _write_minimal_mapping_workbook(mapping)
    out = tmp_path / "assessment"

    result = runner.invoke(
        app,
        [
            "run",
            "migration-assessment",
            "--repo",
            str(sample_repo),
            "--mapping",
            str(mapping),
            "--out",
            str(out),
            "--json",
        ],
    )

    assert result.exit_code == 0, result.output
    data = json.loads(result.output)
    assert data["repo_name"]
    assert data["inputs"]["mapping"] == str(mapping)
    assert any(s["name"] == "assessment_package" for s in data["stage_statuses"])


def test_cli_migration_assessment_missing_mapping_fails(
    sample_repo: Path, tmp_path: Path
) -> None:
    out = tmp_path / "assessment"
    result = runner.invoke(
        app,
        [
            "run",
            "migration-assessment",
            "--repo",
            str(sample_repo),
            "--mapping",
            str(tmp_path / "does_not_exist.xlsx"),
            "--out",
            str(out),
        ],
    )

    assert result.exit_code == 1
    assert "Mapping workbook not found" in result.output
