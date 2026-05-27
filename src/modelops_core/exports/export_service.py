"""Export canonical model objects to CSV, XLSX, and JSONL."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from modelops_core.errors import ResourceLimitExceeded
from modelops_core.repository import parse_file, scan_repository

_COMMON_COLUMNS = ("id", "type", "status", "name", "title", "domain", "description")

# Columns that are generated / read-only in a business-review workbook.
_READ_ONLY_COLUMNS = {"id", "type", "source_file"}


def _flatten_value(value: Any) -> str:
    """Flatten a frontmatter value to a string safe for tabular output."""
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    return str(value)


def _collect_objects(repo_model_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Scan and parse all canonical objects, grouped by type."""
    files = scan_repository(repo_model_path)
    objects_by_type: dict[str, list[dict[str, Any]]] = {}

    for file_path in files:
        parsed = parse_file(file_path)
        if parsed.parser_error or not parsed.frontmatter:
            continue
        fm = dict(parsed.frontmatter)
        obj_type = str(fm.get("type", "Unknown"))
        fm["source_file"] = Path(parsed.source_path).name
        objects_by_type.setdefault(obj_type, []).append(fm)

    return objects_by_type


def _build_columns(objects: list[dict[str, Any]], business_review: bool = False) -> list[str]:
    """Determine column order for a set of objects."""
    extra_keys: set[str] = set()
    for obj in objects:
        extra_keys.update(obj.keys())
    extra_keys.discard("source_file")
    for col in _COMMON_COLUMNS:
        extra_keys.discard(col)
    columns = list(_COMMON_COLUMNS) + sorted(extra_keys) + ["source_file"]
    if business_review:
        columns.append("reviewer_notes")
    return columns


def export_model_csv(
    repo_model_path: Path,
    output_dir: Path | None = None,
    max_objects: int | None = None,
) -> list[Path]:
    """Export canonical objects to one CSV file per type.

    Args:
        repo_model_path: Path to the model directory.
        output_dir: Directory to write CSV files. Defaults to
            ``repo_model_path.parent / "generated" / "exports" / "csv"``.
        max_objects: Maximum objects per type. If exceeded, raises
            ``ResourceLimitExceeded``.

    Returns:
        List of written file paths.

    Raises:
        ResourceLimitExceeded: If any object type exceeds the limit.
    """
    if output_dir is None:
        output_dir = repo_model_path.parent / "generated" / "exports" / "csv"
    output_dir.mkdir(parents=True, exist_ok=True)

    objects_by_type = _collect_objects(repo_model_path)
    if max_objects is not None:
        for obj_type, objects in objects_by_type.items():
            if len(objects) > max_objects:
                raise ResourceLimitExceeded(
                    resource="max_export_objects",
                    message=(
                        f"Type '{obj_type}' has {len(objects)} objects, "
                        f"exceeding max_export_objects limit of {max_objects}. "
                        f"Increase the limit or filter by type."
                    ),
                )
    written: list[Path] = []

    for obj_type, objects in objects_by_type.items():
        safe_name = obj_type.replace(" ", "_").lower()
        file_path = output_dir / f"{safe_name}.csv"
        columns = _build_columns(objects)

        with file_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            for obj in objects:
                row = {col: _flatten_value(obj.get(col)) for col in columns}
                writer.writerow(row)

        written.append(file_path)

    return written


def export_model_jsonl(
    repo_model_path: Path,
    output_dir: Path | None = None,
    max_objects: int | None = None,
) -> list[Path]:
    """Export canonical objects to one JSONL file per type.

    Each line is a JSON object containing all frontmatter fields,
    the body/content, and the source file name.

    Args:
        repo_model_path: Path to the model directory.
        output_dir: Directory to write JSONL files. Defaults to
            ``repo_model_path.parent / "generated" / "exports" / "jsonl"``.
        max_objects: Maximum objects per type. If exceeded, raises
            ``ResourceLimitExceeded``.

    Returns:
        List of written file paths.

    Raises:
        ResourceLimitExceeded: If any object type exceeds the limit.
    """
    if output_dir is None:
        output_dir = repo_model_path.parent / "generated" / "exports" / "jsonl"
    output_dir.mkdir(parents=True, exist_ok=True)

    files = scan_repository(repo_model_path)
    objects_by_type: dict[str, list[dict[str, Any]]] = {}

    for file_path in files:
        parsed = parse_file(file_path)
        if parsed.parser_error or not parsed.frontmatter:
            continue
        fm = dict(parsed.frontmatter)
        obj_type = str(fm.get("type", "Unknown"))
        fm["source_file"] = Path(parsed.source_path).name
        if parsed.body:
            fm["body"] = parsed.body
        objects_by_type.setdefault(obj_type, []).append(fm)

    if max_objects is not None:
        for obj_type, objects in objects_by_type.items():
            if len(objects) > max_objects:
                raise ResourceLimitExceeded(
                    resource="max_export_objects",
                    message=(
                        f"Type '{obj_type}' has {len(objects)} objects, "
                        f"exceeding max_export_objects limit of {max_objects}. "
                        f"Increase the limit or filter by type."
                    ),
                )

    written: list[Path] = []

    for obj_type, objects in objects_by_type.items():
        safe_name = obj_type.replace(" ", "_").lower()
        file_path = output_dir / f"{safe_name}.jsonl"
        # Deterministic ordering by id
        objects_sorted = sorted(objects, key=lambda o: str(o.get("id", "")))

        with file_path.open("w", encoding="utf-8") as f:
            for obj in objects_sorted:
                f.write(json.dumps(obj, default=str, sort_keys=True) + "\n")

        written.append(file_path)

    return written


def export_model_xlsx(
    repo_model_path: Path,
    output_path: Path | None = None,
    max_objects: int | None = None,
    business_review: bool = False,
) -> Path:
    """Export canonical objects to one XLSX workbook with a sheet per type.

    Args:
        repo_model_path: Path to the model directory.
        output_path: Path for the workbook. Defaults to
            ``repo_model_path.parent / "generated" / "exports" / "model.xlsx"``.
        max_objects: Maximum objects per sheet. If exceeded, raises
            ``ResourceLimitExceeded``.
        business_review: When ``True``, produce a styled workbook with a
            cover sheet, data-validation dropdowns, alternating row colours,
            frozen panes, auto-filters, and a ``reviewer_notes`` column.

    Returns:
        Path to the written workbook.

    Raises:
        ResourceLimitExceeded: If any object type exceeds the limit.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for XLSX export. Install it with: pip install openpyxl"
        ) from exc

    if output_path is None:
        output_path = repo_model_path.parent / "generated" / "exports" / "model.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    objects_by_type = _collect_objects(repo_model_path)
    if max_objects is not None:
        for obj_type, objects in objects_by_type.items():
            if len(objects) > max_objects:
                raise ResourceLimitExceeded(
                    resource="max_export_objects",
                    message=(
                        f"Type '{obj_type}' has {len(objects)} objects, "
                        f"exceeding max_export_objects limit of {max_objects}. "
                        f"Increase the limit or filter by type."
                    ),
                )

    wb = Workbook()
    # Remove default sheet; we'll add one per type
    wb.remove(wb.active)

    # Build cover / summary sheet when in business-review mode
    if business_review:
        _add_readme_sheet(wb, objects_by_type)

    # Gather distinct status values across all objects for dropdown
    all_status_values = sorted(
        {
            str(obj.get("status", ""))
            for objs in objects_by_type.values()
            for obj in objs
            if obj.get("status")
        }
    )
    status_dv = None
    if business_review and all_status_values:
        status_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(all_status_values)}"',
            allow_blank=True,
        )
        status_dv.error = "Please select a value from the list."
        status_dv.errorTitle = "Invalid Status"

    for obj_type, objects in objects_by_type.items():
        safe_name = obj_type.replace(" ", "_")[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=safe_name)
        columns = _build_columns(objects, business_review=business_review)

        if business_review:
            _style_business_review_sheet(
                ws,
                columns,
                objects,
                status_dv,
            )
        else:
            # Plain export (legacy)
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)

            for row_idx, obj in enumerate(objects, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=_flatten_value(obj.get(col_name)),
                    )

            for col_idx, col_name in enumerate(columns, start=1):
                max_len = max(len(col_name), 12)
                letter = get_column_letter(col_idx)
                ws.column_dimensions[letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    return output_path


def _add_readme_sheet(
    wb: Any,
    objects_by_type: dict[str, list[dict[str, Any]]],
) -> None:
    """Add a 'Read Me' cover sheet with instructions and object counts."""
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title="Read Me", index=0)
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 15

    title_font = Font(bold=True, size=14)
    heading_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    rows = [
        ("Model Review Workbook", title_font),
        ("", normal_font),
        (
            "This workbook is a generated review view over the canonical model files. "
            "It is NOT the source of truth. Changes made here must be imported back "
            "via 'modelops import-model-sheet' to produce a PatchProposal for human review.",
            normal_font,
        ),
        ("", normal_font),
        ("Editable columns:", heading_font),
        (
            "  - status, name, title, domain, description, and any other business columns.",
            normal_font,
        ),
        ("", normal_font),
        ("Read-only / generated columns (do not edit):", heading_font),
        ("  - id, type, source_file", normal_font),
        ("", normal_font),
        ("Review column:", heading_font),
        (
            "  - reviewer_notes: free-text comments that are NOT imported "
            "into canonical files.",
            normal_font,
        ),
        ("", normal_font),
        ("Object counts by type:", heading_font),
    ]

    for idx, (text, font) in enumerate(rows, start=1):
        cell = ws.cell(row=idx, column=1, value=text)
        cell.font = font
        cell.alignment = wrap_align

    row_offset = len(rows) + 1
    for obj_type in sorted(objects_by_type.keys()):
        count = len(objects_by_type[obj_type])
        ws.cell(row=row_offset, column=1, value=f"  - {obj_type}").font = normal_font
        ws.cell(row=row_offset, column=2, value=count).font = normal_font
        row_offset += 1


def _style_business_review_sheet(
    ws: Any,
    columns: list[str],
    objects: list[dict[str, Any]],
    status_dv: Any | None,
) -> None:
    """Apply business-review styling to a single worksheet."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    readonly_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, obj in enumerate(objects, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            if col_name == "reviewer_notes":
                value = ""
            else:
                value = _flatten_value(obj.get(col_name))

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align
            cell.border = thin_border

            # Alternating row colour
            if row_idx % 2 == 0:
                cell.fill = alt_fill

            # Read-only columns
            if col_name in _READ_ONLY_COLUMNS:
                cell.fill = readonly_fill
                cell.font = Font(italic=True, color="666666")

            # Reviewer notes column styling
            if col_name == "reviewer_notes":
                cell.fill = review_fill
                cell.font = Font(italic=True, color="333333")

    # Auto-filter on header row
    last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(objects) + 1}"

    # Freeze header row
    ws.freeze_panes = "A2"

    # Data validation for status column
    if status_dv is not None and "status" in columns:
        status_col_idx = columns.index("status") + 1
        status_col_letter = ws.cell(row=1, column=status_col_idx).column_letter
        dv_range = f"{status_col_letter}2:{status_col_letter}{len(objects) + 1}"
        ws.add_data_validation(status_dv)
        status_dv.add(dv_range)

    # Column widths
    for col_idx, col_name in enumerate(columns, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        if col_name in ("description", "reviewer_notes"):
            ws.column_dimensions[letter].width = 50
        elif col_name == "id":
            ws.column_dimensions[letter].width = 35
        else:
            max_len = max(len(col_name), 12)
            for obj in objects:
                val_len = len(_flatten_value(obj.get(col_name, "")))
                max_len = max(max_len, val_len)
            ws.column_dimensions[letter].width = min(max_len + 2, 45)
