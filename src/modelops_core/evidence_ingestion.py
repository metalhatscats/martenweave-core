"""Deterministic evidence-to-proposal ingestion for local review workflows."""

from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from modelops_core.patching.patch_proposal_service import render_patch_proposal_markdown
from modelops_core.patching.patch_validator import validate_patch_proposal
from modelops_core.validation.result import ValidationSeverity

_MAX_FINDINGS = 25
_ID_PATTERN = re.compile(r"[A-Z][A-Z0-9]*(?:-[A-Z0-9]+)*")
_NOTE_SIGNALS = (
    ("missing owner", "missing_owner"),
    ("missing mapping", "missing_mapping"),
    ("validation", "validation_issue"),
    ("unresolved", "unresolved_question"),
    ("decision", "decision_note"),
)


class EvidenceIngestionError(ValueError):
    """Raised when evidence cannot be safely converted into a review proposal."""


@dataclass(frozen=True)
class EvidenceIngestionResult:
    proposal: dict[str, Any]
    finding_count: int
    source_sha256: str


def _file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _proposal_id(file_hash: str) -> str:
    return f"PP-EVIDENCE-{file_hash[:16].upper()}"


def _issue_id(file_hash: str, index: int) -> str:
    return f"ISS-EVIDENCE-{file_hash[:10].upper()}-{index:03d}"


def _normalise_text(value: object) -> str:
    return " ".join(str(value).split())


def _severity(value: str) -> str:
    normalized = value.strip().lower()
    if normalized in {"critical", "high", "error", "blocker"}:
        return "high"
    if normalized in {"low", "info", "informational"}:
        return "low"
    return "medium"


def _object_ids(text: str) -> list[str]:
    return sorted({candidate for candidate in _ID_PATTERN.findall(text) if len(candidate) >= 3})


def _first_header(headers: dict[str, Any], candidates: tuple[str, ...]) -> Any | None:
    return next((headers[key] for key in candidates if key in headers), None)


def _issue_operation(
    *,
    issue_id: str,
    name: str,
    issue_type: str,
    severity: str,
    detail: str,
) -> dict[str, Any]:
    return {
        "op": "create_object",
        "object_id": issue_id,
        "object_type": "Issue",
        "after": {
            "id": issue_id,
            "type": "Issue",
            "status": "open",
            "schema_version": "1.0",
            "name": name,
            "issue_type": issue_type,
            "severity": severity,
            "recommended_action": detail,
        },
        "reason": detail,
    }


def _findings_from_note(path: Path, file_hash: str) -> list[dict[str, Any]]:
    note = path.read_text(encoding="utf-8")
    findings: list[dict[str, Any]] = []
    for line in note.splitlines():
        detail = _normalise_text(line.lstrip("#-*> "))
        if not detail:
            continue
        lowered = detail.lower()
        matching_signal = next((signal for signal in _NOTE_SIGNALS if signal[0] in lowered), None)
        if matching_signal is None:
            continue
        _, issue_type = matching_signal
        affected = _object_ids(detail)
        name = detail[:120]
        if affected:
            name = f"{name} ({', '.join(affected)})"[:160]
        findings.append(
            _issue_operation(
                issue_id=_issue_id(file_hash, len(findings) + 1),
                name=name,
                issue_type=issue_type,
                severity="medium",
                detail=detail,
            )
        )
        if len(findings) == _MAX_FINDINGS:
            break
    if not findings:
        raise EvidenceIngestionError(
            "No supported candidate finding was found in the note. Include a missing owner, "
            "missing mapping, validation issue, decision note, or unresolved question."
        )
    return findings


def _findings_from_csv(path: Path, file_hash: str) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise EvidenceIngestionError("Validation report has no header row.")
        headers = {header.lower().strip(): header for header in reader.fieldnames if header}
        message_key = _first_header(
            headers, ("message", "description", "finding", "issue", "error")
        )
        if message_key is None:
            raise EvidenceIngestionError(
                "Validation report needs a message, description, finding, issue, or error column."
            )
        severity_key = _first_header(headers, ("severity", "level", "priority"))
        type_key = _first_header(headers, ("type", "code", "rule", "rule_id"))
        findings: list[dict[str, Any]] = []
        for row in reader:
            detail = _normalise_text(row.get(message_key, ""))
            if not detail:
                continue
            issue_type = _normalise_text(row.get(type_key, "")) if type_key else "validation_issue"
            issue_type = issue_type.lower().replace(" ", "_")[:60] or "validation_issue"
            severity = _severity(_normalise_text(row.get(severity_key, "")) if severity_key else "")
            findings.append(
                _issue_operation(
                    issue_id=_issue_id(file_hash, len(findings) + 1),
                    name=detail[:160],
                    issue_type=issue_type,
                    severity=severity,
                    detail=detail,
                )
            )
            if len(findings) == _MAX_FINDINGS:
                break
    if not findings:
        raise EvidenceIngestionError("Validation report contains no non-empty finding rows.")
    return findings


def _findings_from_xlsx(path: Path, file_hash: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is a package requirement
        raise EvidenceIngestionError("XLSX ingestion requires openpyxl.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            raise EvidenceIngestionError("Validation workbook has no header row.")
        headers = [_normalise_text(value) for value in header_row]
        if not any(headers):
            raise EvidenceIngestionError("Validation workbook has no header row.")
        # Normalize the first populated worksheet through the same strict CSV extractor
        # without persisting a transformed input beside user evidence.
        header_map = {
            header.lower().strip(): index for index, header in enumerate(headers) if header
        }
        message_index = _first_header(
            header_map, ("message", "description", "finding", "issue", "error")
        )
        if message_index is None:
            raise EvidenceIngestionError(
                "Validation workbook needs a message, description, finding, issue, or error column."
            )
        severity_index = _first_header(header_map, ("severity", "level", "priority"))
        type_index = _first_header(header_map, ("type", "code", "rule", "rule_id"))
        findings: list[dict[str, Any]] = []
        for row in rows:
            detail = _normalise_text(row[message_index] if message_index < len(row) else "")
            if not detail:
                continue
            raw_type = row[type_index] if type_index is not None and type_index < len(row) else ""
            raw_severity = ""
            if severity_index is not None and severity_index < len(row):
                raw_severity = row[severity_index]
            issue_type = _normalise_text(raw_type).lower().replace(" ", "_")[:60]
            findings.append(
                _issue_operation(
                    issue_id=_issue_id(file_hash, len(findings) + 1),
                    name=detail[:160],
                    issue_type=issue_type or "validation_issue",
                    severity=_severity(_normalise_text(raw_severity)),
                    detail=detail,
                )
            )
            if len(findings) == _MAX_FINDINGS:
                break
    finally:
        workbook.close()
    if not findings:
        raise EvidenceIngestionError("Validation workbook contains no non-empty finding rows.")
    return findings


def ingest_evidence(path: Path, repo_model_path: Path) -> EvidenceIngestionResult:
    """Create a deterministic, validated PatchProposal from one local evidence file."""
    source = path.resolve()
    if not source.is_file():
        raise EvidenceIngestionError(f"Evidence file not found: {path}")
    file_hash = _file_hash(source)
    suffix = source.suffix.lower()
    if suffix in {".md", ".txt"}:
        operations = _findings_from_note(source, file_hash)
        source_kind = "note"
    elif suffix == ".csv":
        operations = _findings_from_csv(source, file_hash)
        source_kind = "validation_report_csv"
    elif suffix == ".xlsx":
        operations = _findings_from_xlsx(source, file_hash)
        source_kind = "validation_report_xlsx"
    else:
        raise EvidenceIngestionError("Evidence must be Markdown, text, CSV, or XLSX.")

    proposal = {
        "id": _proposal_id(file_hash),
        "type": "PatchProposal",
        "status": "pending_review",
        "schema_version": "1.0",
        "name": f"Evidence ingestion: {source.name}",
        "title": f"Evidence ingestion proposal from {source.name}",
        "created_by": "system",
        "generated_by": "deterministic_evidence_ingestion",
        "source_evidence": (
            f"{source_kind}: {source.name}; SHA-256 {file_hash}; "
            "local evidence remains an input and requires human review"
        ),
        "operations": operations,
        "validation_status": "pending",
        "validation_results": [],
        "assumptions": [
            "Evidence is interpreted only as a candidate Issue proposal.",
            "No source row or note is treated as canonical model truth.",
        ],
        "human_checks": [
            "Confirm each candidate Issue against the original evidence.",
            "Approve the PatchProposal before applying any canonical change.",
        ],
    }
    validation = validate_patch_proposal(proposal, repo_model_path=repo_model_path)
    errors = [result for result in validation if result.severity == ValidationSeverity.ERROR]
    if errors:
        summary = "; ".join(result.message for result in errors)
        raise EvidenceIngestionError(
            f"Generated proposal failed deterministic validation: {summary}"
        )
    return EvidenceIngestionResult(
        proposal=proposal,
        finding_count=len(operations),
        source_sha256=file_hash,
    )


def write_evidence_proposal(result: EvidenceIngestionResult, out_path: Path) -> Path:
    """Write a validated proposal to an explicit path outside canonical model files."""
    target = out_path.resolve()
    if target.suffix.lower() != ".md":
        raise EvidenceIngestionError("Output proposal path must end in .md.")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(render_patch_proposal_markdown(result.proposal), encoding="utf-8")
    return target
