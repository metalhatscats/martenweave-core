"""One-command SAP migration assessment workflow.

Orchestrates validation, indexing, mapping workbook profiling, optional dataset
readiness, assessment package generation, review pack generation, and a machine-
readable manifest. Canonical files are never mutated.
"""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from modelops_core import __version__
from modelops_core.assessment.assessment_service import (
    generate_assessment_package,
    generate_review_pack,
)
from modelops_core.assessment.finding_contract import AssessmentFinding, FindingProvenance
from modelops_core.config import (
    RepoConfig,
    load_repo_config,
    resolve_generated_path,
    resolve_model_path,
)
from modelops_core.index import build_index as _build_index
from modelops_core.pilot.structural_scan import scan_workbook_structure
from modelops_core.pilot.workbook_suggestion_review import write_workbook_suggestion_review_xlsx
from modelops_core.pilot.workbook_suggestions import (
    generate_workbook_suggestions,
    write_workbook_suggestions_json,
    write_workbook_suggestions_markdown,
)
from modelops_core.repository import parse_file, scan_repository
from modelops_core.run.dataset_readiness import (
    generate_dataset_readiness_report,
    write_readiness_report,
)
from modelops_core.validation import validate_objects
from modelops_core.validation.result import ValidationSummary


@dataclass
class StageStatus:
    """Status for one assessment stage."""

    name: str
    status: str  # success | skipped | failed
    message: str = ""


def _is_field_mapping_sheet(name: str) -> bool:
    """Return True for source-to-target field mapping sheets."""
    lower = name.lower()
    return lower.endswith("_mappings") and "value" not in lower


def _is_decisions_sheet(name: str) -> bool:
    """Return True for decision registers."""
    return "decision" in name.lower()


_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "source_field": (
        "source_field",
        "source field",
        "legacy field",
        "old field",
        "field source",
        "поле источника",
        "исходное поле",
    ),
    "source_system": ("source_system", "source system", "legacy system", "система источника"),
    "source_table": ("source_table", "source table", "legacy table", "таблица источника"),
    "target_table": ("target_table", "target table", "sap table", "таблица цели"),
    "target_field": ("target_field", "target field", "new field", "field target", "поле цели"),
    "owner": ("owner", "steward", "responsible", "владелец"),
    "status": ("status", "state", "disposition", "статус"),
    "condition": ("condition", "when", "if", "условие"),
    "validation_rule": ("validation_rule", "validation rule", "validation", "правило валидации"),
    "topic": ("topic", "decision topic", "решение", "тема"),
    "decision_id": ("decision_id", "decision id", "идентификатор решения"),
    "reviewer_comment": ("reviewer_comment", "reviewer comment", "comment", "note", "remarks"),
}


def _normalize_header(value: str) -> str:
    """Normalize an Excel header without discarding non-Latin letters."""
    text = unicodedata.normalize("NFKC", value).casefold()
    return re.sub(r"[^\w]+", "", text, flags=re.UNICODE)


def _header_roles(headers: list[str]) -> dict[str, str]:
    """Resolve irregular but recognizable headers to semantic mapping roles."""
    normalized = {_normalize_header(header): header for header in headers if header}
    resolved: dict[str, str] = {}
    for role, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            header = normalized.get(_normalize_header(alias))
            if header:
                resolved[role] = header
                break
    return resolved


_SEVERITY_BY_CATEGORY: dict[str, str] = {
    "missing_owner": "medium",
    "missing_mapping": "high",
    "obsolete_field": "low",
    "validation_coverage_gap": "medium",
    "unresolved_decision": "medium",
    "conflicting_decision": "high",
    "duplicate_target": "medium",
}


def _stable_id(*parts: str) -> str:
    """Build a deterministic, URL-safe finding ID from parts."""
    cleaned = []
    for part in parts:
        text = str(part).lower().strip().replace(" ", "_")
        text = "".join(c for c in text if c.isalnum() or c in "_-").strip("_-")
        if text:
            cleaned.append(text)
    return ":".join(cleaned)


def _readiness_impact_for_severity(severity: str) -> str:
    """Map severity to a readiness impact label."""
    if severity in ("high", "critical"):
        return "blocking"
    if severity == "medium":
        return "ready_with_warnings"
    return "informational"


_RECOMMENDED_ACTIONS: dict[str, str] = {
    "missing_owner": "Assign an owner in the mapping workbook or canonical object.",
    "missing_mapping": "Add the target mapping and link it to a canonical attribute.",
    "obsolete_field": "Review and retire the obsolete field or document the exception.",
    "validation_coverage_gap": "Add a ValidationRule for the conditional mapping.",
    "unresolved_decision": "Record a Decision object with status and rationale.",
    "conflicting_decision": "Reconcile conflicting decisions and update the decision register.",
    "duplicate_target": "Consolidate duplicate target mappings into a single representation.",
}


def _affected_objects_from_location(location: dict[str, Any]) -> list[str]:
    """Extract stable identifiers from a workbook location when available."""
    objects: list[str] = []
    key_columns = location.get("key_columns") or {}
    if key_columns.get("target_table"):
        objects.append(key_columns["target_table"])
    if key_columns.get("target_field"):
        objects.append(key_columns["target_field"])
    if key_columns.get("source_field"):
        objects.append(key_columns["source_field"])
    if not objects and location.get("topic"):
        objects.append(str(location["topic"]))
    return objects


def _build_findings(
    profile: MappingWorkbookProfile, assessment_run_id: str
) -> list[dict[str, Any]]:
    """Convert mapping-workbook profile findings into stable, reviewable IDs."""
    findings: list[dict[str, Any]] = []

    for row in profile.missing_owner_rows:
        findings.append(
            {
                "id": _stable_id("mapping", "missing_owner", row["sheet"], row["row"]),
                "category": "missing_owner",
                "severity": _SEVERITY_BY_CATEGORY["missing_owner"],
                "source": "mapping_profile",
                "location": row,
                "message": f"Missing owner in '{row['sheet']}' row {row['row']}.",
            }
        )

    for row in profile.missing_mapping_rows:
        findings.append(
            {
                "id": _stable_id("mapping", "missing_mapping", row["sheet"], row["row"]),
                "category": "missing_mapping",
                "severity": _SEVERITY_BY_CATEGORY["missing_mapping"],
                "source": "mapping_profile",
                "location": row,
                "message": (f"Missing target mapping in '{row['sheet']}' row {row['row']}."),
            }
        )

    for row in profile.obsolete_rows:
        findings.append(
            {
                "id": _stable_id("mapping", "obsolete_field", row["sheet"], row["row"]),
                "category": "obsolete_field",
                "severity": _SEVERITY_BY_CATEGORY["obsolete_field"],
                "source": "mapping_profile",
                "location": row,
                "message": f"Obsolete field in '{row['sheet']}' row {row['row']}.",
            }
        )

    for row in profile.validation_coverage_gaps:
        findings.append(
            {
                "id": _stable_id("mapping", "validation_coverage_gap", row["sheet"], row["row"]),
                "category": "validation_coverage_gap",
                "severity": _SEVERITY_BY_CATEGORY["validation_coverage_gap"],
                "source": "mapping_profile",
                "location": row,
                "message": (
                    f"Conditional rule without validation coverage in "
                    f"'{row['sheet']}' row {row['row']}."
                ),
            }
        )

    for row in profile.unresolved_decisions:
        findings.append(
            {
                "id": _stable_id("decision", "unresolved", row["sheet"], row["row"]),
                "category": "unresolved_decision",
                "severity": _SEVERITY_BY_CATEGORY["unresolved_decision"],
                "source": "mapping_profile",
                "location": row,
                "message": (f"Unresolved decision in '{row['sheet']}' row {row['row']}."),
            }
        )

    for conflict in profile.conflicting_decisions:
        topic = conflict.get("topic", "unknown")
        findings.append(
            {
                "id": _stable_id("decision", "conflict", topic),
                "category": "conflicting_decision",
                "severity": _SEVERITY_BY_CATEGORY["conflicting_decision"],
                "source": "mapping_profile",
                "location": conflict,
                "message": f"Conflicting decisions on topic '{topic}'.",
            }
        )

    for row in profile.duplicate_target_rows:
        findings.append(
            {
                "id": _stable_id("mapping", "duplicate_target", row["sheet"], row["row"]),
                "category": "duplicate_target",
                "severity": _SEVERITY_BY_CATEGORY["duplicate_target"],
                "source": "mapping_profile",
                "location": row,
                "message": (
                    f"Duplicate target representation in '{row['sheet']}' row {row['row']}."
                ),
            }
        )

    return [
        AssessmentFinding(
            id=finding["id"],
            category=finding["category"],
            severity=finding["severity"],
            message=finding["message"],
            status="open",
            lifecycle_state="open",
            provenance=FindingProvenance(
                assessment_run_id=assessment_run_id,
                source_kind="mapping_profile",
                detection_mode="deterministic",
                rule_id=f"mapping_profile:{finding['category']}",
                location=finding["location"],
                evidence_refs=["mapping_profile.json"],
                affected_objects=_affected_objects_from_location(finding["location"]),
            ),
            rule_id=f"mapping_profile:{finding['category']}",
            evidence_refs=["mapping_profile.json"],
            affected_objects=_affected_objects_from_location(finding["location"]),
            recommended_action=_RECOMMENDED_ACTIONS.get(
                finding["category"], "Review the finding and record a decision."
            ),
            readiness_impact=_readiness_impact_for_severity(finding["severity"]),
        ).model_dump(mode="json")
        for finding in findings
    ]


@dataclass
class MappingWorkbookProfile:
    """Metadata profile for a mapping workbook input."""

    file_path: str
    file_hash: str
    sheet_names: list[str] = field(default_factory=list)
    total_rows: int = 0
    column_names: list[str] = field(default_factory=list)
    hidden_sheets: list[str] = field(default_factory=list)
    formula_warnings: list[str] = field(default_factory=list)
    missing_owner_rows: list[dict[str, Any]] = field(default_factory=list)
    duplicate_rows: list[dict[str, Any]] = field(default_factory=list)
    missing_mapping_rows: list[dict[str, Any]] = field(default_factory=list)
    obsolete_rows: list[dict[str, Any]] = field(default_factory=list)
    validation_coverage_gaps: list[dict[str, Any]] = field(default_factory=list)
    unresolved_decisions: list[dict[str, Any]] = field(default_factory=list)
    conflicting_decisions: list[dict[str, Any]] = field(default_factory=list)
    duplicate_target_rows: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class MigrationAssessmentManifest:
    """Machine-readable manifest for a migration assessment run."""

    martenweave_version: str
    repo_name: str
    repo_path: str
    inputs: dict[str, Any]
    stage_statuses: list[StageStatus]
    generated_artifacts: list[dict[str, Any]]
    generated_at: str
    run_id: str
    input_fingerprint: str
    input_checksums: dict[str, str]


def _file_hash(path: Path) -> str:
    """Return SHA-256 hex digest for a file."""
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _assessment_inputs_fingerprint(
    repo_root: Path,
    mapping_path: Path,
    dataset_path: Path | None,
    evidence_paths: list[Path],
    enabled_packs: list[str] | None,
) -> tuple[str, dict[str, str]]:
    """Hash every deterministic input that can change an assessment result."""
    checksums: dict[str, str] = {}
    model_path = resolve_model_path(repo_root)
    for path in sorted(scan_repository(model_path)):
        relative_path = Path(path).relative_to(model_path).as_posix()
        checksums[f"model/{relative_path}"] = _file_hash(Path(path))
    config_path = repo_root / "modelops.config.yaml"
    if config_path.exists():
        checksums["modelops.config.yaml"] = _file_hash(config_path)
    checksums["mapping"] = _file_hash(mapping_path)
    if dataset_path is not None:
        checksums["dataset"] = _file_hash(dataset_path)
    for index, path in enumerate(sorted(evidence_paths)):
        checksums[f"evidence/{index}:{path.name}"] = _file_hash(path)
    payload = {"checksums": checksums, "enabled_domain_packs": enabled_packs or []}
    fingerprint = hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()
    return fingerprint, checksums


def _profile_mapping_workbook(mapping_path: Path) -> MappingWorkbookProfile:
    """Profile visible, detected workbook sections with deterministic row checks.

    Mapping workbooks often have a title above the header, multiple sections on
    one sheet, or familiar labels rather than the export's exact snake-case
    headers.  The structural scanner is therefore the authority for boundaries;
    this profiler only interprets detected visible sections and never infers
    meaning from workbook cell values beyond the named columns.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for mapping workbook profiling.") from exc

    file_hash = _file_hash(mapping_path)
    structural_manifest = scan_workbook_structure(mapping_path)
    wb = load_workbook(mapping_path, data_only=False, read_only=True)
    sheet_names = list(wb.sheetnames)
    hidden_sheets = [sheet.name for sheet in structural_manifest.sheets if sheet.hidden]

    column_names: list[str] = []
    total_rows = 0
    missing_owner_rows: list[dict[str, Any]] = []
    duplicate_rows: list[dict[str, Any]] = []
    missing_mapping_rows: list[dict[str, Any]] = []
    obsolete_rows: list[dict[str, Any]] = []
    validation_coverage_gaps: list[dict[str, Any]] = []
    unresolved_decisions: list[dict[str, Any]] = []
    conflicting_decisions: list[dict[str, Any]] = []
    duplicate_target_rows: list[dict[str, Any]] = []

    row_tuples: list[tuple[str, int, tuple[str, ...]]] = []
    target_seen: dict[str, dict[tuple[str, str], tuple[str, int]]] = {}
    decision_topics: dict[str, list[dict[str, Any]]] = {}

    for sheet in structural_manifest.sheets:
        if not sheet.included or not sheet.tables:
            continue
        ws = wb[sheet.name]
        for table in sheet.tables:
            header_row = next(
                ws.iter_rows(
                    min_row=table.header_row,
                    max_row=table.header_row,
                    values_only=True,
                ),
                (),
            )
            headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
            roles = _header_roles(headers)
            if not roles:
                continue
            for header in headers:
                if header and header not in column_names:
                    column_names.append(header)

            is_mapping = bool(
                roles.get("source_field")
                and (roles.get("target_table") or roles.get("target_field"))
                and "value" not in sheet.name.casefold()
            )
            is_decisions = _is_decisions_sheet(sheet.name) or bool(roles.get("decision_id"))
            if is_mapping:
                target_seen.setdefault(sheet.name, {})

            for idx, row in enumerate(
                ws.iter_rows(
                    min_row=table.header_row + 1,
                    max_row=table.end_row,
                    values_only=True,
                ),
                start=table.header_row + 1,
            ):
                values = ["" if value is None else str(value).strip() for value in row]
                if not any(values):
                    continue
                total_rows += 1
                row_dict = {h: value for h, value in zip(headers, values, strict=False) if h}
                status = row_dict.get(roles.get("status", ""), "").casefold()

                owner_col = roles.get("owner")
                if owner_col and not row_dict.get(owner_col, ""):
                    key_columns = {
                        role: row_dict[column]
                        for role in ("source_field", "target_table", "target_field")
                        if (column := roles.get(role)) and row_dict.get(column)
                    }
                    missing_owner_rows.append(
                        {"sheet": sheet.name, "row": idx, "key_columns": key_columns}
                    )

                if is_mapping:
                    source_field = row_dict.get(roles.get("source_field", ""), "")
                    target_table = row_dict.get(roles.get("target_table", ""), "")
                    target_field = row_dict.get(roles.get("target_field", ""), "")

                    if status != "obsolete" and (not target_table or not target_field):
                        missing_mapping_rows.append(
                            {
                                "sheet": sheet.name,
                                "row": idx,
                                "source_field": source_field,
                                "source_system": row_dict.get(roles.get("source_system", ""), ""),
                                "source_table": row_dict.get(roles.get("source_table", ""), ""),
                                "target_table": target_table,
                                "target_field": target_field,
                            }
                        )

                    if status == "obsolete":
                        obsolete_rows.append(
                            {
                                "sheet": sheet.name,
                                "row": idx,
                                "source_field": source_field,
                                "target_table": target_table,
                                "target_field": target_field,
                                "reviewer_comment": row_dict.get(
                                    roles.get("reviewer_comment", ""), ""
                                ),
                            }
                        )

                    condition = row_dict.get(roles.get("condition", ""), "")
                    validation_rule = row_dict.get(roles.get("validation_rule", ""), "")
                    if status != "obsolete" and condition and not validation_rule:
                        validation_coverage_gaps.append(
                            {
                                "sheet": sheet.name,
                                "row": idx,
                                "source_field": source_field,
                                "condition": condition,
                            }
                        )

                    key = tuple(
                        value for value in (source_field, target_table, target_field) if value
                    )
                    if key:
                        row_tuples.append((sheet.name, idx, key))

                    tkey = (target_table, target_field)
                    if tkey[0] or tkey[1]:
                        sheet_target_seen = target_seen[sheet.name]
                        if tkey in sheet_target_seen:
                            prev_sheet, prev_idx = sheet_target_seen[tkey]
                            duplicate_target_rows.append(
                                {
                                    "sheet": sheet.name,
                                    "row": idx,
                                    "duplicate_of": {"sheet": prev_sheet, "row": prev_idx},
                                    "target_table": target_table,
                                    "target_field": target_field,
                                }
                            )
                        else:
                            sheet_target_seen[tkey] = (sheet.name, idx)

                if status == "pending":
                    unresolved_decisions.append(
                        {
                            "sheet": sheet.name,
                            "row": idx,
                            "decision_id": row_dict.get(roles.get("decision_id", ""), ""),
                            "topic": row_dict.get(roles.get("topic", ""), ""),
                        }
                    )

                topic_col = roles.get("topic")
                if is_decisions and topic_col:
                    topic = row_dict.get(topic_col, "").casefold()
                    if topic:
                        decision_topics.setdefault(topic, []).append(
                            {
                                "sheet": sheet.name,
                                "row": idx,
                                "topic": row_dict.get(topic_col, ""),
                                "decision_id": row_dict.get(roles.get("decision_id", ""), ""),
                                "status": status,
                            }
                        )

    wb.close()

    # Formula pass
    formula_warnings: list[str] = []
    try:
        wb_formula = load_workbook(mapping_path, data_only=False, read_only=True)
        for sheet_name in wb_formula.sheetnames:
            ws = wb_formula[sheet_name]
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if (
                        cell.value is not None
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_warnings.append(
                            f"Formula in sheet '{sheet_name}' cell {cell.coordinate}"
                        )
        wb_formula.close()
    except Exception:  # pragma: no cover - best effort
        pass

    # Exact-duplicate reporting (limit to first 50 for sanity)
    seen: dict[tuple, tuple[str, int]] = {}
    for sheet_name, idx, key in row_tuples:
        if key in seen:
            prev_sheet, prev_idx = seen[key]
            duplicate_rows.append(
                {
                    "sheet": sheet_name,
                    "row": idx,
                    "duplicate_of": {"sheet": prev_sheet, "row": prev_idx},
                    "key": key,
                }
            )
        else:
            seen[key] = (sheet_name, idx)
        if len(duplicate_rows) >= 50:
            break

    # Conflicting decisions: the same topic appears more than once in the Decisions sheet.
    for topic, topic_rows in decision_topics.items():
        if len(topic_rows) > 1:
            conflicting_decisions.append(
                {"topic": topic, "rows": topic_rows, "count": len(topic_rows)}
            )

    return MappingWorkbookProfile(
        file_path=str(mapping_path),
        file_hash=file_hash,
        sheet_names=sheet_names,
        total_rows=total_rows,
        column_names=column_names,
        hidden_sheets=hidden_sheets,
        formula_warnings=formula_warnings,
        missing_owner_rows=missing_owner_rows,
        duplicate_rows=duplicate_rows,
        missing_mapping_rows=missing_mapping_rows,
        obsolete_rows=obsolete_rows,
        validation_coverage_gaps=validation_coverage_gaps,
        unresolved_decisions=unresolved_decisions,
        conflicting_decisions=conflicting_decisions,
        duplicate_target_rows=duplicate_target_rows,
    )


def _validate_repo(repo_root: Path) -> tuple[ValidationSummary, RepoConfig | None]:
    """Validate canonical objects in the repository."""
    model_path = resolve_model_path(repo_root)
    files = scan_repository(model_path)
    parsed_objects = [parse_file(f) for f in files]
    config = load_repo_config(repo_root)
    enabled_packs = config.enabled_domain_packs if config else None
    summary = validate_objects(parsed_objects, enabled_packs)
    return summary, config


def _ensure_index(repo_root: Path) -> Path:
    """Build the SQLite index if missing or refresh it."""
    db_path = resolve_generated_path(repo_root) / "modelops.db"
    _build_index(repo_root, db_path=db_path, allow_invalid=True, export_jsonl=False)
    return db_path


def _write_manifest(
    manifest: MigrationAssessmentManifest,
    out_dir: Path,
) -> Path:
    """Write the manifest as JSON."""
    manifest_path = out_dir / "manifest.json"
    data = {
        "martenweave_version": manifest.martenweave_version,
        "repo_name": manifest.repo_name,
        "repo_path": manifest.repo_path,
        "generated_at": manifest.generated_at,
        "run_id": manifest.run_id,
        "input_fingerprint": manifest.input_fingerprint,
        "input_checksums": manifest.input_checksums,
        "inputs": manifest.inputs,
        "stage_statuses": [
            {"name": s.name, "status": s.status, "message": s.message}
            for s in manifest.stage_statuses
        ],
        "generated_artifacts": manifest.generated_artifacts,
    }
    manifest_path.write_text(
        json.dumps(data, indent=2, default=str, sort_keys=True),
        encoding="utf-8",
    )
    return manifest_path


def _collect_artifacts(out_dir: Path) -> list[dict[str, Any]]:
    """Collect all files under out_dir and compute hashes."""
    artifacts: list[dict[str, Any]] = []
    for path in sorted(out_dir.rglob("*")):
        if path.is_file():
            rel = path.relative_to(out_dir).as_posix()
            artifacts.append(
                {
                    "path": rel,
                    "size": path.stat().st_size,
                    "sha256": _file_hash(path),
                }
            )
    return artifacts


def _stage(name: str, statuses: list[StageStatus], status: str, message: str = "") -> None:
    statuses.append(StageStatus(name=name, status=status, message=message))


def generate_migration_assessment(
    repo_root: Path,
    mapping_path: Path,
    dataset_path: Path | None,
    evidence_paths: list[Path],
    out_dir: Path,
    generated_at: str | None = None,
) -> MigrationAssessmentManifest:
    """Generate a complete migration assessment output package and manifest.

    Args:
        repo_root: Path to the model repository.
        mapping_path: Path to the XLSX mapping workbook.
        dataset_path: Optional path to a CSV/XLSX sample dataset.
        evidence_paths: Optional list of evidence file paths.
        out_dir: Directory where all outputs will be written.
        generated_at: Optional ISO timestamp. Defaults to the current UTC time.

    Returns:
        MigrationAssessmentManifest describing inputs, stage statuses, and artifacts.

    Raises:
        RuntimeError: if a required stage fails and the manifest cannot be written.
        ValueError: if required inputs are missing or invalid.
    """
    repo_root = repo_root.resolve()
    out_dir = out_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    config = load_repo_config(repo_root)
    repo_name = config.name if config else repo_root.name

    statuses: list[StageStatus] = []
    generated_at = generated_at or datetime.now(UTC).isoformat().replace("+00:00", "Z")

    inputs: dict[str, Any] = {
        "mapping": str(mapping_path),
        "dataset": str(dataset_path) if dataset_path else None,
        "evidence": [str(p) for p in evidence_paths],
        "enabled_domain_packs": config.enabled_domain_packs if config else [],
    }
    input_fingerprint, input_checksums = _assessment_inputs_fingerprint(
        repo_root,
        mapping_path,
        dataset_path,
        evidence_paths,
        config.enabled_domain_packs if config else None,
    )
    assessment_run_id = f"ASSESSMENT-{input_fingerprint[:16].upper()}"

    # Stage: validation
    try:
        summary, _ = _validate_repo(repo_root)
        _stage(
            "validation",
            statuses,
            "success" if summary.is_valid else "success",
            message=f"{summary.error_count} errors, {summary.warning_count} warnings",
        )
    except Exception as exc:  # pragma: no cover - defensive
        _stage("validation", statuses, "failed", message=str(exc))
        # Continue so the manifest can still be written

    # Stage: index
    try:
        _ensure_index(repo_root)
        _stage("index", statuses, "success")
    except Exception as exc:  # pragma: no cover - defensive
        _stage("index", statuses, "failed", message=str(exc))

    # Stage: workbook structural scan
    workbook_manifest = None
    try:
        workbook_manifest = scan_workbook_structure(mapping_path)
        workbook_manifest_path = out_dir / "workbook_manifest.json"
        workbook_manifest_path.write_text(
            json.dumps(workbook_manifest.to_dict(), indent=2, default=str) + "\n",
            encoding="utf-8",
        )
        detected_tables = sum(len(sheet.tables) for sheet in workbook_manifest.sheets)
        _stage(
            "structural_scan",
            statuses,
            "success",
            message=f"{len(workbook_manifest.sheets)} sheets, {detected_tables} table section(s)",
        )
    except Exception as exc:
        _stage("structural_scan", statuses, "failed", message=str(exc))

    # Stage: governed workbook suggestions
    if workbook_manifest is not None:
        try:
            workbook_suggestions = generate_workbook_suggestions(workbook_manifest)
            write_workbook_suggestions_json(
                workbook_suggestions,
                out_dir / "workbook_suggestions.json",
            )
            write_workbook_suggestions_markdown(
                workbook_suggestions,
                out_dir / "workbook_suggestions.md",
            )
            write_workbook_suggestion_review_xlsx(
                workbook_suggestions,
                out_dir / "workbook_suggestion_review.xlsx",
            )
            _stage(
                "workbook_suggestions",
                statuses,
                "success",
                message=f"{len(workbook_suggestions.suggestions)} suggestion(s)",
            )
        except Exception as exc:
            _stage("workbook_suggestions", statuses, "failed", message=str(exc))
    else:
        _stage(
            "workbook_suggestions",
            statuses,
            "skipped",
            message="Structural scan did not complete",
        )

    # Stage: mapping workbook profile
    mapping_profile: MappingWorkbookProfile | None = None
    try:
        mapping_profile = _profile_mapping_workbook(mapping_path)
        _stage("mapping_profile", statuses, "success")
    except Exception as exc:
        _stage("mapping_profile", statuses, "failed", message=str(exc))

    if mapping_profile is not None:
        profile_path = out_dir / "mapping_profile.json"
        profile_path.write_text(
            json.dumps(
                {
                    "file_path": mapping_profile.file_path,
                    "file_hash": mapping_profile.file_hash,
                    "sheet_names": mapping_profile.sheet_names,
                    "total_rows": mapping_profile.total_rows,
                    "column_names": mapping_profile.column_names,
                    "hidden_sheets": mapping_profile.hidden_sheets,
                    "formula_warnings": mapping_profile.formula_warnings,
                    "missing_owner_rows": mapping_profile.missing_owner_rows,
                    "duplicate_rows": mapping_profile.duplicate_rows,
                    "missing_mapping_rows": mapping_profile.missing_mapping_rows,
                    "obsolete_rows": mapping_profile.obsolete_rows,
                    "validation_coverage_gaps": mapping_profile.validation_coverage_gaps,
                    "unresolved_decisions": mapping_profile.unresolved_decisions,
                    "conflicting_decisions": mapping_profile.conflicting_decisions,
                    "duplicate_target_rows": mapping_profile.duplicate_target_rows,
                },
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )

        # Stable findings derived from the mapping workbook profile.
        findings = _build_findings(mapping_profile, assessment_run_id)
        findings_path = out_dir / "findings.json"
        findings_path.write_text(
            json.dumps(
                {
                    "generated_at": generated_at,
                    "finding_count": len(findings),
                    "findings": findings,
                },
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )

    # Stage: dataset readiness (optional)
    readiness_dir = out_dir / "dataset_readiness"
    if dataset_path is not None:
        try:
            readiness_report = generate_dataset_readiness_report(
                repo_root=repo_root,
                dataset_path=dataset_path,
                check_model=True,
                dry_run=False,
                promote_to_proposal=False,
                issue_draft=False,
            )
            readiness_dir.mkdir(parents=True, exist_ok=True)
            write_readiness_report(readiness_report, readiness_dir)
            _stage("dataset_readiness", statuses, "success")
        except Exception as exc:
            _stage("dataset_readiness", statuses, "failed", message=str(exc))
    else:
        _stage("dataset_readiness", statuses, "skipped", message="No dataset provided")

    # Stage: assessment package
    try:
        generate_assessment_package(
            repo_root, out_dir, generated_at=generated_at, assessment_run_id=assessment_run_id
        )
        _stage("assessment_package", statuses, "success")
    except Exception as exc:
        _stage("assessment_package", statuses, "failed", message=str(exc))

    # Stage: review pack
    try:
        review_pack_dir = out_dir / "review_pack"
        review_pack_dir.mkdir(parents=True, exist_ok=True)
        generate_review_pack(repo_root, review_pack_dir, generated_at=generated_at)
        _stage("review_pack", statuses, "success")
    except Exception as exc:
        _stage("review_pack", statuses, "failed", message=str(exc))

    # Collect artifacts and write manifest (including the manifest itself)
    artifacts = _collect_artifacts(out_dir)
    manifest_path_placeholder = {
        "path": "manifest.json",
        "size": 0,
        "sha256": "",
    }
    artifacts.append(manifest_path_placeholder)
    manifest = MigrationAssessmentManifest(
        martenweave_version=__version__,
        repo_name=repo_name,
        repo_path=str(repo_root),
        inputs=inputs,
        stage_statuses=statuses,
        generated_artifacts=artifacts,
        generated_at=generated_at,
        run_id=assessment_run_id,
        input_fingerprint=input_fingerprint,
        input_checksums=input_checksums,
    )
    manifest_path = _write_manifest(manifest, out_dir)
    manifest_path_placeholder["size"] = manifest_path.stat().st_size
    manifest_path_placeholder["sha256"] = _file_hash(manifest_path)
    _write_manifest(manifest, out_dir)

    return manifest
