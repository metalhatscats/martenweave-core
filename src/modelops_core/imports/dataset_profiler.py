"""Streaming CSV profiler with bounded memory."""

from __future__ import annotations

import csv
import hashlib
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path


@dataclass
class ProfilingStatus:
    success: bool = True
    truncated: bool = False
    sampled: bool = False
    sample_interval: int | None = None
    reason: str | None = None
    rows_processed: int = 0
    file_size_bytes: int = 0


@dataclass
class ColumnProfile:
    name: str
    position: int
    blank_count: int = 0
    non_blank_count: int = 0
    distinct_count: int = 0
    sample_values: list[str] = field(default_factory=list)
    inferred_type: str = "string"


@dataclass
class DatasetProfile:
    dataset_id: str
    file_path: str
    file_hash: str
    row_count: int = 0
    column_count: int = 0
    columns: list[ColumnProfile] = field(default_factory=list)
    status: ProfilingStatus = field(default_factory=ProfilingStatus)
    sheet_name: str | None = None
    header_row: int = 1


@dataclass
class WorkbookProfile:
    dataset_id: str
    file_path: str
    file_hash: str
    sheet_names: list[str] = field(default_factory=list)
    sheets: list[DatasetProfile] = field(default_factory=list)
    status: ProfilingStatus = field(default_factory=ProfilingStatus)


MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB
MAX_ROWS = 500_000
MAX_COLUMNS = 1_000
SAMPLE_SIZE = 5
MAX_HEADER_SCAN_ROWS = 25

_HEADER_HINTS = (
    "source",
    "target",
    "legacy",
    "sap",
    "owner",
    "steward",
    "responsible",
    "status",
    "state",
    "field",
    "table",
    "поле",
    "таблица",
    "владелец",
    "статус",
)


def _compute_file_hash(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(8192), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _infer_type(values: list[str]) -> str:
    types_seen: set[str] = set()
    for v in values:
        v_stripped = v.strip()
        if v_stripped == "":
            types_seen.add("empty")
            continue
        if v_stripped.lower() in {"true", "false", "1", "0", "yes", "no"}:
            types_seen.add("boolean")
            continue
        try:
            int(v_stripped)
            types_seen.add("integer")
            continue
        except ValueError:
            pass
        try:
            float(v_stripped)
            types_seen.add("decimal")
            continue
        except ValueError:
            pass
        types_seen.add("string")

    if len(types_seen) == 1:
        return types_seen.pop()
    if "string" in types_seen:
        return "mixed"
    if {"integer", "decimal"} <= types_seen:
        return "decimal"
    return "mixed"


def _header_candidate_score(values: tuple[object, ...]) -> int:
    """Score a likely Excel header row without treating title rows as headers."""
    labels = [str(value).strip() for value in values if value is not None and str(value).strip()]
    if len(labels) < 2:
        return 0
    normalized = [unicodedata.normalize("NFKC", label).casefold() for label in labels]
    unique = len(set(normalized)) == len(normalized)
    keyword_hits = sum(
        any(re.search(rf"\b{re.escape(hint)}", label) for hint in _HEADER_HINTS)
        for label in normalized
    )
    return (2 if unique else 0) + keyword_hits


def _find_header_row(ws: object) -> tuple[int, tuple[object, ...]]:
    """Return the most plausible header from the first bounded worksheet rows."""
    best_row = 1
    best_values: tuple[object, ...] = ()
    best_score = -1
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True), start=1
    ):
        values = tuple(row)
        score = _header_candidate_score(values)
        if score > best_score:
            best_row, best_values, best_score = row_index, values, score
    return best_row, best_values


def profile_csv(
    csv_path: Path,
    dataset_id: str,
    max_file_size: int = MAX_FILE_SIZE_BYTES,
    max_rows: int = MAX_ROWS,
    max_columns: int = MAX_COLUMNS,
    sample_interval: int | None = None,
) -> DatasetProfile:
    """Profile a CSV file with bounded memory usage.

    If *sample_interval* is provided, every Nth row is processed for
    statistics and the result is marked as sampled rather than exact.
    """
    file_size = csv_path.stat().st_size
    if file_size > max_file_size:
        return DatasetProfile(
            dataset_id=dataset_id,
            file_path=str(csv_path),
            file_hash="",
            status=ProfilingStatus(
                success=False,
                truncated=True,
                reason=f"File size {file_size} exceeds limit {max_file_size}.",
                file_size_bytes=file_size,
            ),
        )

    file_hash = _compute_file_hash(csv_path)

    with csv_path.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        try:
            headers = next(reader)
        except StopIteration:
            return DatasetProfile(
                dataset_id=dataset_id,
                file_path=str(csv_path),
                file_hash=file_hash,
                status=ProfilingStatus(
                    success=False,
                    truncated=True,
                    reason="Empty CSV file.",
                    file_size_bytes=file_size,
                ),
            )

        if len(headers) > max_columns:
            return DatasetProfile(
                dataset_id=dataset_id,
                file_path=str(csv_path),
                file_hash=file_hash,
                status=ProfilingStatus(
                    success=False,
                    truncated=True,
                    reason=f"Column count {len(headers)} exceeds limit {max_columns}.",
                    file_size_bytes=file_size,
                ),
            )

        columns: list[ColumnProfile] = [
            ColumnProfile(name=h, position=i + 1) for i, h in enumerate(headers)
        ]
        row_count = 0
        processed_count = 0
        distinct_sets: list[set[str]] = [set() for _ in headers]
        sampled = sample_interval is not None and sample_interval > 1

        for row in reader:
            if row_count >= max_rows:
                break
            row_count += 1
            if sampled and row_count % sample_interval != 0:
                continue
            processed_count += 1
            for i, value in enumerate(row):
                if i >= len(columns):
                    break
                col = columns[i]
                if value.strip() == "":
                    col.blank_count += 1
                else:
                    col.non_blank_count += 1
                    distinct_sets[i].add(value.strip())
                    if len(col.sample_values) < SAMPLE_SIZE:
                        col.sample_values.append(value.strip())

        for i, col in enumerate(columns):
            col.distinct_count = len(distinct_sets[i])
            col.inferred_type = _infer_type(list(distinct_sets[i])[: SAMPLE_SIZE * 2])

    status = ProfilingStatus(
        success=True,
        sampled=sampled,
        sample_interval=sample_interval,
        rows_processed=processed_count,
        file_size_bytes=file_size,
    )
    if sampled:
        status.reason = (
            f"Statistics estimated from every {sample_interval}th row "
            f"({processed_count} sampled out of {row_count} total)."
        )

    return DatasetProfile(
        dataset_id=dataset_id,
        file_path=str(csv_path),
        file_hash=file_hash,
        row_count=row_count,
        column_count=len(columns),
        columns=columns,
        status=status,
    )


def profile_xlsx(
    xlsx_path: Path,
    dataset_id: str,
    max_file_size: int = MAX_FILE_SIZE_BYTES,
    max_rows: int = MAX_ROWS,
    max_columns: int = MAX_COLUMNS,
    sample_interval: int | None = None,
) -> WorkbookProfile:
    """Profile an XLSX workbook, profiling each sheet independently.

    If *sample_interval* is provided, every Nth row is processed for
    statistics and the result is marked as sampled rather than exact.
    """
    from openpyxl import load_workbook

    file_size = xlsx_path.stat().st_size
    if file_size > max_file_size:
        return WorkbookProfile(
            dataset_id=dataset_id,
            file_path=str(xlsx_path),
            file_hash="",
            status=ProfilingStatus(
                success=False,
                truncated=True,
                reason=f"File size {file_size} exceeds limit {max_file_size}.",
                file_size_bytes=file_size,
            ),
        )

    file_hash = _compute_file_hash(xlsx_path)

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        return WorkbookProfile(
            dataset_id=dataset_id,
            file_path=str(xlsx_path),
            file_hash=file_hash,
            status=ProfilingStatus(
                success=False,
                truncated=True,
                reason=f"Failed to load workbook: {exc}",
                file_size_bytes=file_size,
            ),
        )

    sheet_profiles: list[DatasetProfile] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            continue
        header_row, header_values = _find_header_row(ws)
        if not header_values:
            sheet_profiles.append(
                DatasetProfile(
                    dataset_id=dataset_id,
                    file_path=str(xlsx_path),
                    file_hash=file_hash,
                    sheet_name=sheet_name,
                    status=ProfilingStatus(
                        success=False,
                        truncated=True,
                        reason="Empty sheet.",
                        file_size_bytes=file_size,
                    ),
                )
            )
            continue

        headers = [str(value) if value is not None else "" for value in header_values]

        if len(headers) > max_columns:
            sheet_profiles.append(
                DatasetProfile(
                    dataset_id=dataset_id,
                    file_path=str(xlsx_path),
                    file_hash=file_hash,
                    sheet_name=sheet_name,
                    status=ProfilingStatus(
                        success=False,
                        truncated=True,
                        reason=f"Column count {len(headers)} exceeds limit {max_columns}.",
                        file_size_bytes=file_size,
                    ),
                )
            )
            continue

        columns: list[ColumnProfile] = [
            ColumnProfile(name=h, position=i + 1) for i, h in enumerate(headers)
        ]
        row_count = 0
        processed_count = 0
        distinct_sets: list[set[str]] = [set() for _ in headers]
        sampled = sample_interval is not None and sample_interval > 1

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row_count >= max_rows:
                break
            row_count += 1
            if sampled and row_count % sample_interval != 0:
                continue
            processed_count += 1
            for i, value in enumerate(row):
                if i >= len(columns):
                    break
                col = columns[i]
                str_value = "" if value is None else str(value).strip()
                if str_value == "":
                    col.blank_count += 1
                else:
                    col.non_blank_count += 1
                    distinct_sets[i].add(str_value)
                    if len(col.sample_values) < SAMPLE_SIZE:
                        col.sample_values.append(str_value)

        for i, col in enumerate(columns):
            col.distinct_count = len(distinct_sets[i])
            col.inferred_type = _infer_type(list(distinct_sets[i])[: SAMPLE_SIZE * 2])

        status = ProfilingStatus(
            success=True,
            sampled=sampled,
            sample_interval=sample_interval,
            rows_processed=processed_count,
            file_size_bytes=file_size,
        )
        if sampled:
            status.reason = (
                f"Statistics estimated from every {sample_interval}th row "
                f"({processed_count} sampled out of {row_count} total)."
            )

        sheet_profiles.append(
            DatasetProfile(
                dataset_id=dataset_id,
                file_path=str(xlsx_path),
                file_hash=file_hash,
                sheet_name=sheet_name,
                header_row=header_row,
                row_count=row_count,
                column_count=len(columns),
                columns=columns,
                status=status,
            )
        )

    wb.close()

    return WorkbookProfile(
        dataset_id=dataset_id,
        file_path=str(xlsx_path),
        file_hash=file_hash,
        sheet_names=list(wb.sheetnames),
        sheets=sheet_profiles,
        status=ProfilingStatus(
            success=all(s.status.success for s in sheet_profiles) if sheet_profiles else True,
            sampled=any(s.status.sampled for s in sheet_profiles),
            sample_interval=sample_interval,
            rows_processed=sum(s.status.rows_processed for s in sheet_profiles),
            file_size_bytes=file_size,
        ),
    )


def dataset_profile_to_dict(profile: DatasetProfile | WorkbookProfile) -> dict:
    """Convert a profile dataclass to a deterministic dict for JSON serialization."""
    return asdict(profile)
