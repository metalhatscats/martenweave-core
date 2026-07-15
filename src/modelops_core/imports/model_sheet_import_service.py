"""Import spreadsheet edits as PatchProposals."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from modelops_core.patching.patch_model import PatchOperation
from modelops_core.repository import parse_file, scan_repository
from modelops_core.schemas import ObjectType, SourceState

# Columns that are metadata / not frontmatter fields
_META_COLUMNS = {"source_file", "reviewer_notes"}

# Reference-like fields that are validated for broken references
_REFERENCE_LIKE_FIELDS = {
    "domain",
    "entity",
    "parent_entity",
    "entity_context",
    "attribute",
    "business_attribute",
    "field_endpoint",
    "value_list",
    "mapping",
    "validation_rule",
    "system",
    "migration_object",
    "source_value_list",
    "target_value_list",
    "owner",
    "business_owner",
    "technical_owner",
    "data_steward",
    "approver",
    "accountable_team",
}


class SpreadsheetImportError(ValueError):
    """Raised when a review workbook cannot safely identify its rows."""


def _read_csv_dir(csv_dir: Path) -> dict[str, list[dict[str, str]]]:
    """Read all CSV files in a directory, keyed by object type."""
    rows_by_type: dict[str, list[dict[str, str]]] = {}
    for csv_file in sorted(csv_dir.glob("*.csv")):
        obj_type = csv_file.stem.replace("_", " ")
        with csv_file.open("r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows_by_type[obj_type] = list(reader)
    return rows_by_type


def _read_xlsx(xlsx_path: Path, max_rows: int | None = None) -> dict[str, list[dict[str, str]]]:
    """Read all sheets from an XLSX workbook, keyed by object type."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for XLSX import. Install it with: pip install openpyxl"
        ) from exc

    rows_by_type: dict[str, list[dict[str, str]]] = {}
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        # The business-review export includes instructions, not model rows.
        if sheet_name == "Read Me":
            continue
        ws = wb[sheet_name]
        obj_type = sheet_name.replace("_", " ")
        headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, str]] = []
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if max_rows is not None and row_count >= max_rows:
                break
            row_count += 1
            row_dict = {
                h: str(v) if v is not None else "" for h, v in zip(headers, row, strict=False)
            }
            if any(row_dict.values()):
                rows.append(row_dict)
        rows_by_type[obj_type] = rows
    wb.close()
    return rows_by_type


def _require_row_ids(rows_by_type: dict[str, list[dict[str, str]]]) -> None:
    """Reject review input that cannot be traced to a stable canonical ID."""
    problems: list[str] = []
    for obj_type, rows in rows_by_type.items():
        for row_number, row in enumerate(rows, start=2):
            if not row.get("id", "").strip():
                problems.append(f"sheet '{obj_type}' row {row_number}")
    if problems:
        locations = ", ".join(problems[:5])
        remaining = len(problems) - 5
        suffix = f" (and {remaining} more)" if remaining > 0 else ""
        raise SpreadsheetImportError(
            "Every non-empty review row must include a stable 'id'; missing ID at "
            f"{locations}{suffix}."
        )


def _detect_formulas(xlsx_path: Path) -> list[str]:
    """Detect formula cells in an XLSX workbook.

    Returns a list of human-readable warning strings with sheet and cell refs.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    warnings: list[str] = []
    wb = load_workbook(xlsx_path, data_only=False, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if (
                    cell.value is not None
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    warnings.append(
                        f"Formula detected in sheet '{sheet_name}' cell {cell.coordinate}: "
                        f"formulas are not evaluated during import."
                    )
    wb.close()
    return warnings


def _validate_import(
    rows_by_type: dict[str, list[dict[str, str]]],
    existing: dict[str, dict[str, Any]],
) -> list[str]:
    """Validate spreadsheet rows and return actionable warnings.

    Checks:
    - Missing 'id' column in sheets
    - Unknown sheet / object type names
    - Broken references to objects not in existing or workbook
    """
    warnings: list[str] = []
    known_types = {ot.value for ot in ObjectType}
    all_ids_in_workbook: set[str] = set()

    for obj_type, rows in rows_by_type.items():
        if obj_type not in known_types:
            warnings.append(
                f"Sheet '{obj_type}' does not match a known object type. "
                f"Expected one of: {', '.join(sorted(known_types)[:5])}..."
            )

        if rows and "id" not in rows[0]:
            warnings.append(
                f"Sheet '{obj_type}' is missing required 'id' column. All rows will be skipped."
            )
            continue

        for row in rows:
            obj_id = row.get("id")
            if obj_id:
                all_ids_in_workbook.add(obj_id)

    # Check references
    for obj_type, rows in rows_by_type.items():
        for idx, row in enumerate(rows, start=2):
            obj_id = row.get("id")
            if not obj_id:
                continue
            for col, val in row.items():
                if col in _META_COLUMNS or col == "id" or not val:
                    continue
                if col in _REFERENCE_LIKE_FIELDS or col.endswith("_id"):
                    refs: list[str] = []
                    if "; " in val:
                        refs = [v.strip() for v in val.split(";") if v.strip()]
                    else:
                        refs = [val.strip()]
                    for ref_id in refs:
                        if ref_id not in existing and ref_id not in all_ids_in_workbook:
                            warnings.append(
                                f"Broken reference in sheet '{obj_type}' row {idx} "
                                f"({obj_id}): '{col}' points to '{ref_id}' which does not exist."
                            )

    return warnings


def _load_existing_objects(repo_model_path: Path) -> dict[str, dict[str, Any]]:
    """Load all canonical objects keyed by ID."""
    existing: dict[str, dict[str, Any]] = {}
    for file_path in scan_repository(repo_model_path):
        parsed = parse_file(file_path)
        if parsed.parser_error or not parsed.frontmatter:
            continue
        obj_id = str(parsed.frontmatter.get("id", ""))
        if obj_id:
            existing[obj_id] = dict(parsed.frontmatter)
    return existing


def _unflatten_value(value: str) -> str | list[str]:
    """Convert a flattened string back to original type if it looks like a list."""
    if "; " in value:
        return [v.strip() for v in value.split(";") if v.strip()]
    return value


def _diff_row(row: dict[str, str], existing: dict[str, Any]) -> list[PatchOperation]:
    """Compare a spreadsheet row against an existing object and return update ops."""
    ops: list[PatchOperation] = []
    obj_id = row["id"]
    obj_type = row.get("type", str(existing.get("type", "")))

    for col, new_val_raw in row.items():
        if col in _META_COLUMNS or col == "id":
            continue
        new_val = _unflatten_value(new_val_raw)
        old_val = existing.get(col)

        # Normalize for comparison
        if isinstance(old_val, list):
            old_val_str = "; ".join(str(v) for v in old_val)
        elif old_val is None:
            old_val_str = ""
        else:
            old_val_str = str(old_val)

        new_val_str = (
            "; ".join(str(v) for v in new_val) if isinstance(new_val, list) else str(new_val)
        )

        if old_val_str != new_val_str:
            ops.append(
                PatchOperation(
                    op="update_object",
                    object_id=obj_id,
                    object_type=obj_type,
                    target_path=col,
                    before=old_val_str if old_val_str else None,
                    after=new_val_str if new_val_str else None,
                    reason=f"Changed '{col}' from spreadsheet import",
                )
            )

    return ops


def _build_proposal(
    rows_by_type: dict[str, list[dict[str, str]]],
    existing: dict[str, dict[str, Any]],
    source: str,
    extra_warnings: list[str] | None = None,
    require_stable_ids: bool = False,
) -> dict[str, Any]:
    """Build a PatchProposal from spreadsheet rows compared to existing objects."""
    if require_stable_ids:
        _require_row_ids(rows_by_type)
    operations: list[PatchOperation] = []
    warnings: list[str] = list(extra_warnings or [])
    affected_objects: list[str] = []
    seen_ids: set[str] = set()

    warnings.extend(_validate_import(rows_by_type, existing))

    for obj_type, rows in rows_by_type.items():
        for row in rows:
            obj_id = row.get("id", "")

            if not obj_id:
                warnings.append("Row skipped: missing id")
                continue

            if obj_id in seen_ids:
                warnings.append(f"Duplicate ID in spreadsheet: {obj_id}")
                continue
            seen_ids.add(obj_id)

            if obj_id in existing:
                ops = _diff_row(row, existing[obj_id])
                operations.extend(ops)
                if ops:
                    affected_objects.append(obj_id)
            else:
                # New object — build frontmatter from row
                frontmatter: dict[str, Any] = {"id": obj_id, "type": obj_type}
                for col, val in row.items():
                    if col in _META_COLUMNS or col == "id":
                        continue
                    if val:
                        frontmatter[col] = _unflatten_value(val)

                operations.append(
                    PatchOperation(
                        op="create_object",
                        object_id=obj_id,
                        object_type=obj_type,
                        after=frontmatter,
                        reason="New object from spreadsheet import",
                    )
                )
                affected_objects.append(obj_id)

    proposal_stem = Path(source).stem.upper()
    if require_stable_ids:
        proposal_stem = proposal_stem.replace("_", "-")
    proposal_id = f"PP-IMPORT-{proposal_stem}"
    return {
        "id": proposal_id,
        "type": "PatchProposal",
        "status": "pending_review",
        "name": proposal_id,
        "title": f"Spreadsheet review: {Path(source).name}",
        "validation_status": "pending",
        "operations": [op.model_dump() for op in operations],
        "affected_objects": affected_objects,
        "source_evidence": source,
        "source_state": SourceState.PROPOSAL.value,
        "created_by": "system",
        "warnings": warnings,
        "assumptions": ["Spreadsheet columns map directly to canonical frontmatter fields."],
        "human_checks": ["Review each operation before applying."],
    }


def import_model_sheet_csv(
    csv_dir: Path,
    repo_model_path: Path,
) -> dict[str, Any]:
    """Import a directory of CSV files and produce a PatchProposal.

    Args:
        csv_dir: Directory containing exported CSV files.
        repo_model_path: Path to the model directory.

    Returns:
        A PatchProposal dict capturing detected changes.
    """
    rows_by_type = _read_csv_dir(csv_dir)
    existing = _load_existing_objects(repo_model_path)
    return _build_proposal(rows_by_type, existing, str(csv_dir))


def import_model_sheet_xlsx(
    xlsx_path: Path,
    repo_model_path: Path,
    max_rows: int | None = None,
    require_stable_ids: bool = False,
) -> dict[str, Any]:
    """Import an XLSX workbook and produce a PatchProposal.

    Args:
        xlsx_path: Path to the exported XLSX workbook.
        repo_model_path: Path to the model directory.
        max_rows: Maximum rows to read per sheet. If exceeded, rows are truncated.
        require_stable_ids: Reject non-empty rows without an ID. Intended for
            the controlled business-review workflow; the general importer
            preserves its legacy warning-oriented behavior.

    Returns:
        A PatchProposal dict capturing detected changes.
    """
    rows_by_type = _read_xlsx(xlsx_path, max_rows=max_rows)
    existing = _load_existing_objects(repo_model_path)
    formula_warnings = _detect_formulas(xlsx_path)
    extra_warnings: list[str] = list(formula_warnings)
    if max_rows is not None:
        for obj_type, rows in rows_by_type.items():
            if len(rows) >= max_rows:
                extra_warnings.append(
                    f"Sheet '{obj_type}' truncated at {max_rows} rows (max_import_rows limit)."
                )
    return _build_proposal(
        rows_by_type,
        existing,
        str(xlsx_path),
        extra_warnings=extra_warnings,
        require_stable_ids=require_stable_ids,
    )
