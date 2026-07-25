"""Normalized inspection for local machine-readable schema documents."""

from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from html.parser import HTMLParser
from pathlib import Path, PurePosixPath
from typing import Any
from urllib.parse import unquote, urlparse
from xml.etree import ElementTree as ET

import yaml

_HTTP_METHODS = {"get", "post", "put", "patch", "delete", "options", "head"}
_PARSER_VERSION = "1.0"
_EDMX_NAMESPACES = {
    "edmx": "http://docs.oasis-open.org/odata/ns/edmx",
    "edm": "http://docs.oasis-open.org/odata/ns/edm",
}
_WSDL_NAMESPACES = {
    "wsdl": "http://schemas.xmlsoap.org/wsdl/",
    "soap": "http://schemas.xmlsoap.org/wsdl/soap/",
    "soap12": "http://schemas.xmlsoap.org/wsdl/soap12/",
}
_XSD_NAMESPACE = "http://www.w3.org/2001/XMLSchema"
_XSD_NAMESPACES = {
    "xs": _XSD_NAMESPACE,
    "xsd": _XSD_NAMESPACE,
}
_HEADER_TOKEN = re.compile(r"[^a-z0-9]+")
_IDOC_SEGMENT_NAME = re.compile(r"^(?:EDI_DC\d+|E[1-9][A-Z0-9_]+|Z[A-Z0-9_]+|Y[A-Z0-9_]+)$")
_IDOC_BASIC_TYPE_NAME = re.compile(r"^[A-Z][A-Z0-9_]{3,}\d{2,}$")
_WE60_SEGMENT_ALIASES = {
    "segment",
    "segment_type",
    "segment_name",
    "segmenttype",
    "segment_name_type",
    "segmtype",
}
_WE60_FIELD_ALIASES = {
    "field",
    "field_name",
    "fieldname",
    "field_name_type",
    "attribute",
    "attribute_name",
}
_WE60_DESCRIPTION_ALIASES = {
    "description",
    "field_description",
    "segment_description",
    "meaning",
    "documentation",
    "short_text",
}
_WE60_DATA_TYPE_ALIASES = {"data_type", "datatype", "type", "domain", "built_in_type"}
_WE60_LENGTH_ALIASES = {"length", "output_length", "internal_length", "intlen"}
_WE60_REQUIRED_ALIASES = {"required", "mandatory", "obligatory"}
_WE60_MIN_ALIASES = {"min", "minimum", "min_occurs", "minoccurs"}
_WE60_MAX_ALIASES = {"max", "maximum", "max_occurs", "maxoccurs"}
_WE60_PARENT_ALIASES = {"parent_segment", "parent", "superior_segment"}
_INTEGRATION_RESOURCE_DIRS = {
    "src/main/resources/wsdl": "wsdl_resource",
    "src/main/resources/xsd": "xsd_resource",
    "src/main/resources/edmx": "edmx_resource",
    "src/main/resources/mapping": "mapping_resource",
}
_ZIP_ENTRY_LIMIT = 200
_ZIP_TOTAL_UNCOMPRESSED_LIMIT = 10_000_000
_CDS_METADATA_ALIASES: dict[str, set[str]] = {
    "artifact": {
        "artifact",
        "artifact_name",
        "cds_artifact",
        "cds_artifact_name",
        "view_name",
        "entity_name",
        "artifactname",
    },
    "artifact_kind": {"artifact_kind", "kind", "artifact_type", "cds_kind"},
    "element": {
        "element",
        "element_name",
        "field_name",
        "column_name",
        "attribute_name",
        "elementname",
    },
    "element_kind": {"element_kind", "field_kind", "member_kind"},
    "data_type": {"data_type", "datatype", "type_name", "type", "builtin_type", "sql_type"},
    "description": {"description", "label", "end_user_text", "short_text"},
    "length": {"length", "max_length", "type_length"},
    "precision": {"precision", "total_digits"},
    "scale": {"scale", "fraction_digits", "decimal_places"},
    "key": {"is_key", "key", "key_flag", "iskey"},
    "association_target": {
        "association_target",
        "target_artifact",
        "target_artifact_name",
        "target_entity",
        "association",
        "association_target_name",
    },
    "min_occurs": {"min_occurs", "min", "cardinality_min", "minimum"},
    "max_occurs": {"max_occurs", "max", "cardinality_max", "maximum"},
    "namespace": {"namespace", "schema_name", "package_name", "package"},
}
_FIELD_CATALOG_ALIASES: dict[str, set[str]] = {
    "entity": {
        "entity",
        "entity_name",
        "business_entity",
        "object",
        "object_name",
        "table",
        "table_name",
        "message",
        "message_name",
        "segment",
        "segment_name",
        "business_object",
    },
    "field": {
        "field",
        "field_name",
        "field_path",
        "path",
        "technical_name",
        "technical_field",
        "column",
        "column_name",
        "name",
        "attribute",
        "attribute_name",
        "element",
        "element_name",
    },
    "description": {
        "description",
        "definition",
        "label",
        "business_name",
        "title",
        "long_text",
    },
    "data_type": {"data_type", "datatype", "type", "field_type", "domain", "format"},
    "required": {"required", "mandatory", "is_required", "required_flag"},
    "nullable": {"nullable", "is_nullable", "optional"},
    "key": {"key", "is_key", "primary_key", "key_flag"},
    "length": {"length", "max_length", "size"},
    "precision": {"precision", "total_digits"},
    "scale": {"scale", "fraction_digits", "decimal_places"},
    "enumerations": {"enum", "allowed_values", "values", "value_list", "permitted_values"},
    "cardinality": {"cardinality", "occurs", "multiplicity"},
}
_MAPPING_TEMPLATE_ALIASES: dict[str, set[str]] = {
    "source_field": {
        "source_field",
        "source field",
        "legacy_field",
        "legacy field",
        "old_field",
        "old field",
        "source_attribute",
    },
    "source_system": {"source_system", "source system", "legacy_system", "legacy system"},
    "target_table": {
        "target_table",
        "target table",
        "sap_table",
        "sap table",
        "target_structure",
        "target structure",
    },
    "target_field": {
        "target_field",
        "target field",
        "sap_field",
        "sap field",
        "new_field",
        "new field",
    },
    "target_system": {"target_system", "target system", "sap_system", "sap system"},
    "rule": {"rule", "mapping_type", "mapping type", "transform", "transformation"},
    "owner": {"owner", "steward", "responsible"},
    "status": {"status", "state", "mapping_status"},
    "comment": {"comment", "note", "notes", "remarks"},
    "description": {"description", "label", "business name", "target description"},
    "data_type": {"data_type", "datatype", "type", "target_type"},
    "required": {"required", "mandatory", "is_required", "target_required"},
}
_MIGRATION_COCKPIT_ALIASES: dict[str, set[str]] = {
    "migration_object": {
        "migration_object",
        "migration object",
        "object",
        "migrationobject",
    },
    "structure": {"structure", "target_structure", "target structure", "segment"},
    "field": {
        "field",
        "field_name",
        "field name",
        "target_field",
        "target field",
        "technical_field",
        "technical field",
    },
    "description": {"description", "field description", "label", "business name"},
    "data_type": {"data_type", "data type", "datatype", "type", "domain"},
    "required": {"required", "mandatory", "is_required", "required flag"},
    "length": {"length", "max_length", "max length"},
    "enumerations": {"allowed_values", "allowed values", "value_list", "value list", "enum"},
}


@dataclass
class NormalizedSchemaField:
    entity_name: str
    field_path: str
    parent_path: str | None
    data_type: str | None
    required: bool | None
    cardinality: str | None
    length: int | None = None
    precision: int | None = None
    scale: int | None = None
    is_key: bool = False
    enumerations: list[str] = field(default_factory=list)
    associations: list[str] = field(default_factory=list)
    annotations: dict[str, Any] = field(default_factory=dict)
    description: str | None = None
    source_evidence: str = ""
    warnings: list[str] = field(default_factory=list)


@dataclass
class NormalizedSchemaEntity:
    name: str
    kind: str
    description: str | None
    source_evidence: str


@dataclass
class NormalizedSchemaOperation:
    operation_id: str
    method: str
    path: str
    protocol: str
    request_body_schema: str | None
    response_schemas: list[dict[str, str]]
    parameters: list[str] = field(default_factory=list)
    description: str | None = None
    source_evidence: str = ""


@dataclass
class NormalizedSchemaDocument:
    source_path: str
    source_format: str
    source_identity: str
    source_version: str | None
    namespace: str | None
    parser_version: str
    checksum: str
    entities: list[NormalizedSchemaEntity] = field(default_factory=list)
    fields: list[NormalizedSchemaField] = field(default_factory=list)
    operations: list[NormalizedSchemaOperation] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class SchemaInspectionError(ValueError):
    """Raised when a schema file cannot be inspected safely."""


@dataclass
class _HtmlLink:
    href: str
    text: str


@dataclass
class _HtmlTable:
    heading: str | None
    headers: list[str]
    rows: list[list[str]]


@dataclass
class _HtmlDocument:
    title: str | None
    headings: list[str]
    tables: list[_HtmlTable]
    links: list[_HtmlLink]


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.title: str | None = None
        self.headings: list[str] = []
        self.tables: list[_HtmlTable] = []
        self.links: list[_HtmlLink] = []
        self._capture_title = False
        self._capture_heading = False
        self._heading_parts: list[str] = []
        self._title_parts: list[str] = []
        self._last_heading: str | None = None
        self._current_table: dict[str, Any] | None = None
        self._current_row: list[tuple[str, str]] | None = None
        self._current_cell_tag: str | None = None
        self._current_cell_parts: list[str] = []
        self._current_link_href: str | None = None
        self._current_link_parts: list[str] = []
        self._skip_tag_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"script", "style"}:
            self._skip_tag_depth += 1
            return
        if self._skip_tag_depth:
            return

        attributes = {key.lower(): value for key, value in attrs}
        lowered = tag.lower()
        if lowered == "title":
            self._capture_title = True
            self._title_parts = []
        elif lowered in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            self._capture_heading = True
            self._heading_parts = []
        elif lowered == "table":
            self._current_table = {"heading": self._last_heading, "headers": [], "rows": []}
        elif lowered == "tr" and self._current_table is not None:
            self._current_row = []
        elif lowered in {"th", "td"} and self._current_row is not None:
            self._current_cell_tag = lowered
            self._current_cell_parts = []
        elif lowered == "br" and self._current_cell_tag is not None:
            self._current_cell_parts.append("\n")
        elif lowered == "a":
            href = attributes.get("href")
            if href:
                self._current_link_href = href
                self._current_link_parts = []

    def handle_endtag(self, tag: str) -> None:
        lowered = tag.lower()
        if lowered in {"script", "style"}:
            if self._skip_tag_depth:
                self._skip_tag_depth -= 1
            return
        if self._skip_tag_depth:
            return

        if lowered == "title" and self._capture_title:
            self.title = _normalize_html_text("".join(self._title_parts)) or None
            self._capture_title = False
        elif lowered in {"h1", "h2", "h3", "h4", "h5", "h6"} and self._capture_heading:
            heading = _normalize_html_text("".join(self._heading_parts))
            if heading:
                self.headings.append(heading)
                self._last_heading = heading
            self._capture_heading = False
        elif lowered in {"th", "td"} and self._current_row is not None and self._current_cell_tag:
            text = _normalize_html_text("".join(self._current_cell_parts))
            self._current_row.append((text, self._current_cell_tag))
            self._current_cell_tag = None
            self._current_cell_parts = []
        elif lowered == "tr" and self._current_table is not None and self._current_row is not None:
            if self._current_row:
                texts = [text for text, _ in self._current_row]
                tags = {cell_tag for _, cell_tag in self._current_row}
                if not self._current_table["headers"] and tags == {"th"}:
                    self._current_table["headers"] = texts
                else:
                    self._current_table["rows"].append(texts)
            self._current_row = None
        elif lowered == "table" and self._current_table is not None:
            self.tables.append(
                _HtmlTable(
                    heading=self._current_table["heading"],
                    headers=self._current_table["headers"],
                    rows=self._current_table["rows"],
                )
            )
            self._current_table = None
        elif lowered == "a" and self._current_link_href is not None:
            self.links.append(
                _HtmlLink(
                    href=self._current_link_href,
                    text=_normalize_html_text("".join(self._current_link_parts)),
                )
            )
            self._current_link_href = None
            self._current_link_parts = []

    def handle_data(self, data: str) -> None:
        if self._skip_tag_depth:
            return
        if self._capture_title:
            self._title_parts.append(data)
        if self._capture_heading:
            self._heading_parts.append(data)
        if self._current_cell_tag is not None:
            self._current_cell_parts.append(data)
        if self._current_link_href is not None:
            self._current_link_parts.append(data)


def _file_checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8192), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_document(path: Path) -> Any:
    raw = path.read_text(encoding="utf-8")
    if path.suffix.lower() == ".json":
        return json.loads(raw)
    if path.suffix.lower() in {".xml", ".edmx", ".xsd", ".wsdl", ".iflw"}:
        return ET.fromstring(raw)
    return yaml.safe_load(raw)


def _normalize_annotations(node: dict[str, Any]) -> dict[str, Any]:
    annotations: dict[str, Any] = {}
    for key in ("format", "nullable", "readOnly", "writeOnly", "deprecated"):
        if key in node:
            annotations[key] = node[key]
    for key, value in node.items():
        if key.startswith("x-"):
            annotations[key] = value
    return annotations


def _normalize_header_name(value: str) -> str:
    normalized = _HEADER_TOKEN.sub("_", value.strip().lower()).strip("_")
    return normalized


def _normalize_html_text(value: str) -> str:
    text = html.unescape(value).replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def _slugify_path_segment(value: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip()).strip("-").lower()
    return normalized or "integration-flow"


def _xml_local_name(tag: str) -> str:
    return tag.split("}", 1)[-1]


def _xml_namespace(tag: str) -> str:
    if tag.startswith("{") and "}" in tag:
        return tag[1:].split("}", 1)[0]
    return ""


def _is_probable_idoc_segment(element: ET.Element) -> bool:
    local_name = _xml_local_name(element.tag)
    segment_flag = str(element.attrib.get("SEGMENT", "")).strip().lower()
    if segment_flag in {"1", "true", "x", "yes"}:
        return True
    return _IDOC_SEGMENT_NAME.fullmatch(local_name) is not None


def _find_idoc_container(document: ET.Element) -> ET.Element | None:
    root_name = _xml_local_name(document.tag)
    if root_name == "IDOC" and any(_is_probable_idoc_segment(child) for child in document):
        return document
    for child in document:
        if _xml_local_name(child.tag) != "IDOC":
            continue
        if any(_is_probable_idoc_segment(grandchild) for grandchild in child):
            return child
    return None


def _header_index(headers: list[str], aliases: set[str]) -> int | None:
    normalized_headers = [_normalize_header_name(header) for header in headers]
    for idx, header in enumerate(normalized_headers):
        if header in aliases:
            return idx
    return None


def _extract_segment_name(text: str | None) -> str | None:
    if not text:
        return None
    for token in re.findall(r"[A-Z0-9_]+", text.upper()):
        if _IDOC_SEGMENT_NAME.fullmatch(token):
            return token
    return None


def _extract_basic_type_name(text: str | None) -> str | None:
    if not text:
        return None
    explicit_match = re.search(
        r"\b(?:basic\s+type|idoc\s+type)\b[^A-Z0-9_]*([A-Z0-9_]+)",
        text,
        re.I,
    )
    if explicit_match:
        candidate = explicit_match.group(1).upper()
        if _IDOC_BASIC_TYPE_NAME.fullmatch(candidate):
            return candidate
    for token in re.findall(r"[A-Z0-9_]+", text.upper()):
        if token == "WE60":
            continue
        if _IDOC_BASIC_TYPE_NAME.fullmatch(token):
            return token
    return None


def _looks_like_we60_html(document: _HtmlDocument) -> bool:
    text = " ".join(filter(None, [document.title, *document.headings]))
    normalized = text.lower()
    if "we60" in normalized and "idoc" in normalized:
        return True
    for table in document.tables:
        header_set = {_normalize_header_name(header) for header in table.headers}
        if header_set & _WE60_SEGMENT_ALIASES and header_set & _WE60_DESCRIPTION_ALIASES:
            return True
        if header_set & _WE60_FIELD_ALIASES and header_set & _WE60_DATA_TYPE_ALIASES:
            return True
    return False


def _parse_html_document(path: Path) -> _HtmlDocument:
    parser = _HtmlTableParser()
    parser.feed(path.read_text(encoding="utf-8"))
    parser.close()
    return _HtmlDocument(
        title=parser.title,
        headings=parser.headings,
        tables=parser.tables,
        links=parser.links,
    )


def _resolve_local_html_link(base_path: Path, href: str) -> Path | None:
    parsed = urlparse(href)
    if parsed.scheme or parsed.netloc:
        return None
    relative = unquote(parsed.path)
    if not relative:
        return None
    candidate = (base_path.parent / relative).resolve()
    root = base_path.parent.resolve()
    try:
        candidate.relative_to(root)
    except ValueError:
        return None
    if candidate.suffix.lower() not in {".html", ".htm"} or not candidate.is_file():
        return None
    return candidate


def _parse_occurs_value(value: str | None) -> str | None:
    if value is None:
        return None
    text = value.strip().upper()
    if not text:
        return None
    if text in {"*", "N", "UNBOUNDED"}:
        return "*"
    if text.isdigit():
        return text
    return None


def _cardinality_from_occurs(
    minimum: str | None,
    maximum: str | None,
) -> tuple[bool | None, str | None]:
    min_occurs = _parse_occurs_value(minimum)
    max_occurs = _parse_occurs_value(maximum)
    if min_occurs is None and max_occurs is None:
        return None, None
    required = None if min_occurs is None else min_occurs != "0"
    if min_occurs is not None and max_occurs is not None:
        return required, f"{min_occurs}..{max_occurs}"
    return required, None


def _parse_manifest_text(raw_text: str) -> dict[str, str]:
    manifest: dict[str, str] = {}
    current_key: str | None = None
    for line in raw_text.splitlines():
        if not line.strip():
            current_key = None
            continue
        if line.startswith(" ") and current_key is not None:
            manifest[current_key] += line[1:]
            continue
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        current_key = key.strip()
        manifest[current_key] = value.strip()
    return manifest


def _iflow_identity(document: ET.Element, fallback: str) -> str:
    for candidate in (
        document.attrib.get("name"),
        document.attrib.get("id"),
    ):
        if candidate:
            return candidate
    for element in document.iter():
        if _xml_local_name(element.tag) not in {"process", "collaboration", "participant"}:
            continue
        for candidate in (element.attrib.get("name"), element.attrib.get("id")):
            if candidate:
                return candidate
    return fallback


def _iflow_protocol(document: ET.Element) -> str:
    for element in document.iter():
        for key in ("componentType", "adapterType", "type", "protocol"):
            value = element.attrib.get(key)
            if value:
                return value
    return "integration_flow"


def _iflow_endpoint_path(document: ET.Element, flow_name: str) -> str:
    for element in document.iter():
        for key in ("endpointAddress", "address", "urlPath", "path"):
            value = element.attrib.get(key)
            if value:
                return value
    return f"/{_slugify_path_segment(flow_name)}"


def _inspect_iflow_document(
    *,
    source_identity: str,
    flow_name: str,
    source_evidence: str,
    document: ET.Element,
) -> tuple[list[NormalizedSchemaEntity], list[NormalizedSchemaOperation], list[str]]:
    protocol = _iflow_protocol(document)
    path = _iflow_endpoint_path(document, flow_name)
    participant_names = [
        element.attrib.get("name", "")
        for element in document.iter()
        if _xml_local_name(element.tag) in {"participant", "messageStartEvent", "serviceTask"}
        and element.attrib.get("name")
    ]
    description = f"Imported integration flow artifact for {flow_name}."
    if participant_names:
        description = (
            f"Imported integration flow artifact for {flow_name}; participants: "
            f"{', '.join(participant_names[:4])}."
        )
    entities = [
        NormalizedSchemaEntity(
            name=flow_name,
            kind="integration_flow",
            description=description,
            source_evidence=source_evidence,
        )
    ]
    operations = [
        NormalizedSchemaOperation(
            operation_id=flow_name,
            method="flow",
            path=path,
            protocol=protocol,
            request_body_schema=None,
            response_schemas=[],
            parameters=[],
            description=f"Integration flow artifact in package {source_identity}.",
            source_evidence=source_evidence,
        )
    ]
    warnings = [
        (
            f"Integration flow operation metadata for {flow_name} is inferred from design-time "
            "artifact structure and may require manual review."
        )
    ]
    return entities, operations, warnings


def _infer_schema_kind(document: Any, *, suffix: str | None = None) -> str:
    if isinstance(document, ET.Element):
        if suffix == ".iflw":
            return "integration_flow_artifact"
        local_name = _xml_local_name(document.tag)
        namespace = _xml_namespace(document.tag)
        if local_name.lower() == "edmx":
            return "edmx"
        if local_name.lower() == "definitions" and namespace == _WSDL_NAMESPACES["wsdl"]:
            return "wsdl"
        if local_name.lower() == "schema" and namespace == _XSD_NAMESPACE:
            return "xsd"
        if suffix == ".xml":
            if _find_idoc_container(document) is not None:
                return "idoc_payload"
            return "xml_payload"
        raise SchemaInspectionError(
            "Unsupported XML schema document: expected OData EDMX, WSDL, XML payload, "
            "or XML Schema."
        )
    if suffix == ".json" and isinstance(document, (dict, list, str, int, float, bool, type(None))):
        rowset = _extract_json_rowset(document)
        if rowset is not None and _looks_like_cds_metadata_rows(rowset):
            return "cds_metadata_export"
        if isinstance(document, dict):
            if "openapi" in document or "swagger" in document:
                return "openapi"
            if "$schema" in document or "properties" in document or "$defs" in document:
                return "json_schema"
            return "json_payload"
        return "json_payload"
    if not isinstance(document, dict):
        raise SchemaInspectionError(
            "Schema document must be a JSON, YAML, OData EDMX, WSDL, XML payload, or "
            "XML Schema object."
        )
    if "openapi" in document or "swagger" in document:
        return "openapi"
    if "$schema" in document or "properties" in document or "$defs" in document:
        return "json_schema"
    raise SchemaInspectionError(
        "Unsupported schema document: expected JSON Schema, OpenAPI, OData EDMX, "
        "WSDL, XML payload, or XML Schema."
    )


def _payload_value_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int) and not isinstance(value, bool):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return type(value).__name__


def _parse_catalog_bool(value: str | None) -> bool | None:
    if value is None:
        return None
    normalized = value.strip().lower()
    if normalized in {"true", "yes", "y", "1", "required", "mandatory", "x"}:
        return True
    if normalized in {"false", "no", "n", "0", "optional", "nullable"}:
        return False
    return None


def _parse_catalog_int(value: str | None) -> int | None:
    if value is None:
        return None
    stripped = value.strip()
    return int(stripped) if stripped.isdigit() else None


def _split_catalog_values(value: str | None) -> list[str]:
    if value is None:
        return []
    text = value.strip()
    if not text:
        return []
    for separator in (";", "|", ","):
        if separator in text:
            return [part.strip() for part in text.split(separator) if part.strip()]
    return [text]


def _catalog_value(row: dict[str, str], role: str) -> str | None:
    for alias in _FIELD_CATALOG_ALIASES[role]:
        value = row.get(alias)
        if value:
            stripped = value.strip()
            if stripped:
                return stripped
    return None


def _mapping_value(row: dict[str, str], role: str) -> str | None:
    for alias in _MAPPING_TEMPLATE_ALIASES[role]:
        value = row.get(alias)
        if value:
            stripped = value.strip()
            if stripped:
                return stripped
    return None


def _migration_cockpit_value(row: dict[str, str], role: str) -> str | None:
    for alias in _MIGRATION_COCKPIT_ALIASES[role]:
        value = row.get(alias)
        if value:
            stripped = value.strip()
            if stripped:
                return stripped
    return None


def _cds_metadata_value(row: dict[str, str], role: str) -> str | None:
    for alias in _CDS_METADATA_ALIASES[role]:
        value = row.get(alias)
        if value:
            stripped = value.strip()
            if stripped:
                return stripped
    return None


def _normalize_tabular_row(row: dict[str, Any]) -> dict[str, str]:
    return {
        _normalize_header_name(str(key)): str(value).strip()
        for key, value in row.items()
        if key is not None
    }


def _extract_json_rowset(document: Any) -> list[dict[str, str]] | None:
    rows: Any
    if isinstance(document, list):
        rows = document
    elif isinstance(document, dict):
        rows = document.get("rows")
    else:
        return None
    if not isinstance(rows, list) or not rows:
        return None

    normalized_rows: list[dict[str, str]] = []
    for row in rows:
        if not isinstance(row, dict):
            return None
        normalized_rows.append(_normalize_tabular_row(row))
    return normalized_rows


def _looks_like_cds_metadata_rows(rows: list[dict[str, str]]) -> bool:
    if not rows:
        return False
    headers = set(rows[0])
    return (
        bool(headers & _CDS_METADATA_ALIASES["artifact"])
        and bool(headers & _CDS_METADATA_ALIASES["element"])
        and (
            bool(headers & _CDS_METADATA_ALIASES["data_type"])
            or bool(headers & _CDS_METADATA_ALIASES["element_kind"])
            or bool(headers & _CDS_METADATA_ALIASES["association_target"])
        )
    )


def _detect_cds_metadata_rows(
    rows_by_sheet: dict[str, list[dict[str, str]]],
) -> dict[str, list[dict[str, str]]]:
    detected: dict[str, list[dict[str, str]]] = {}
    for sheet_name, rows in rows_by_sheet.items():
        normalized_rows = [_normalize_tabular_row(row) for row in rows]
        if _looks_like_cds_metadata_rows(normalized_rows):
            detected[sheet_name] = normalized_rows
    return detected


def _walk_json_payload(
    *,
    entity_name: str,
    value: Any,
    source_pointer: str,
    field_path: str | None,
    required: bool | None,
    fields: list[NormalizedSchemaField],
    warnings: list[str],
) -> None:
    data_type = _payload_value_type(value)
    if field_path is not None:
        fields.append(
            NormalizedSchemaField(
                entity_name=entity_name,
                field_path=field_path,
                parent_path=field_path.rsplit(".", 1)[0] if "." in field_path else None,
                data_type=data_type,
                required=required,
                cardinality=_cardinality(required, data_type),
                length=len(value) if isinstance(value, str) else None,
                precision=None,
                scale=None,
                is_key=field_path.split(".")[-1].replace("[]", "").lower() == "id",
                annotations={"observed_type": data_type},
                description=None,
                source_evidence=source_pointer,
                warnings=[],
            )
        )

    if isinstance(value, dict):
        for key, child in value.items():
            child_path = key if field_path is None else f"{field_path}.{key}"
            _walk_json_payload(
                entity_name=entity_name,
                value=child,
                source_pointer=f"{source_pointer}/{key}",
                field_path=child_path,
                required=True,
                fields=fields,
                warnings=warnings,
            )
        return

    if not isinstance(value, list):
        return

    if not value:
        warnings.append(f"Empty array at {source_pointer} requires manual review.")
        return

    observed_types = {_payload_value_type(item) for item in value}
    if len(observed_types) > 1:
        warnings.append(
            f"Heterogeneous array at {source_pointer} uses representative item inference."
        )

    representative = next((item for item in value if item is not None), value[0])
    child_path = "items[]" if field_path is None else f"{field_path}[]"
    _walk_json_payload(
        entity_name=entity_name,
        value=representative,
        source_pointer=f"{source_pointer}/*",
        field_path=child_path,
        required=None,
        fields=fields,
        warnings=warnings,
    )


def _read_csv_catalog_rows(path: Path) -> dict[str, list[dict[str, str]]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return {
            path.stem: [
                {key or "": str(value or "") for key, value in row.items()} for row in reader
            ]
        }


def _read_xlsx_catalog_rows(path: Path) -> dict[str, list[dict[str, str]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    rows_by_sheet: dict[str, list[dict[str, str]]] = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            rows_by_sheet[sheet_name] = []
            continue
        headers = [str(value) if value is not None else "" for value in header_row]
        rows: list[dict[str, str]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_dict = {
                header: str(value) if value is not None else ""
                for header, value in zip(headers, row, strict=False)
            }
            if any(value.strip() for value in row_dict.values()):
                rows.append(row_dict)
        rows_by_sheet[sheet_name] = rows
    workbook.close()
    return rows_by_sheet


def _read_xlsx_mapping_sections(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    from modelops_core.pilot.structural_scan import scan_workbook_structure

    manifest = scan_workbook_structure(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    sections: list[dict[str, Any]] = []
    try:
        for sheet_result in manifest.sheets:
            if not sheet_result.included or sheet_result.purpose != "mapping":
                continue
            worksheet = workbook[sheet_result.name]
            for table in sheet_result.tables:
                header_row_values = next(
                    worksheet.iter_rows(
                        min_row=table.header_row,
                        max_row=table.header_row,
                        values_only=True,
                    ),
                    None,
                )
                if header_row_values is None:
                    continue
                headers = [
                    str(value).strip() if value is not None else ""
                    for value in header_row_values[: table.column_count or None]
                ]
                normalized_headers = {
                    _normalize_header_name(header) for header in headers if header
                }
                if not (
                    normalized_headers & _MAPPING_TEMPLATE_ALIASES["source_field"]
                    and normalized_headers & _MAPPING_TEMPLATE_ALIASES["target_field"]
                ):
                    continue

                rows: list[dict[str, str]] = []
                for row_number, row_values in enumerate(
                    worksheet.iter_rows(
                        min_row=table.header_row + 1,
                        max_row=table.end_row,
                        values_only=True,
                    ),
                    start=table.header_row + 1,
                ):
                    row_dict = {
                        header: str(value).strip() if value is not None else ""
                        for header, value in zip(headers, row_values, strict=False)
                        if header
                    }
                    if any(value for value in row_dict.values()):
                        rows.append({"_row_number": str(row_number), **row_dict})

                sections.append(
                    {
                        "sheet_name": sheet_result.name,
                        "table_id": table.table_id,
                        "header_row": table.header_row,
                        "title_rows": table.title_rows,
                        "rows": rows,
                    }
                )
    finally:
        workbook.close()
    return sections


def _detect_migration_cockpit_rows(
    rows_by_sheet: dict[str, list[dict[str, str]]],
) -> dict[str, list[dict[str, str]]]:
    detected: dict[str, list[dict[str, str]]] = {}
    for sheet_name, rows in rows_by_sheet.items():
        if not rows:
            continue
        normalized_headers = {_normalize_header_name(header) for header in rows[0]}
        if not (
            normalized_headers & _MIGRATION_COCKPIT_ALIASES["migration_object"]
            and normalized_headers & _MIGRATION_COCKPIT_ALIASES["field"]
        ):
            continue
        detected[sheet_name] = rows
    return detected


def _inspect_cds_metadata_rows(
    *,
    path: Path,
    rows_by_sheet: dict[str, list[dict[str, str]]],
    source_file_format: str,
) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []
    seen_entities: set[str] = set()
    namespaces: set[str] = set()
    artifacts: set[str] = set()

    for sheet_name, rows in rows_by_sheet.items():
        if not rows:
            warnings.append(f"CDS metadata section '{sheet_name}' is empty.")
            continue

        for row_number, row in enumerate(rows, start=2):
            normalized_row = _normalize_tabular_row(row)
            artifact_name = _cds_metadata_value(normalized_row, "artifact")
            element_name = _cds_metadata_value(normalized_row, "element")
            if artifact_name is None or element_name is None:
                warnings.append(
                    f"CDS metadata section '{sheet_name}' row {row_number} is missing artifact "
                    "or element name."
                )
                continue

            artifacts.add(artifact_name)
            namespace = _cds_metadata_value(normalized_row, "namespace")
            if namespace is not None:
                namespaces.add(namespace)

            artifact_kind = _cds_metadata_value(normalized_row, "artifact_kind") or "cds_artifact"
            if artifact_name not in seen_entities:
                seen_entities.add(artifact_name)
                entities.append(
                    NormalizedSchemaEntity(
                        name=artifact_name,
                        kind=artifact_kind.lower(),
                        description=(
                            f"Imported from {source_file_format.upper()} CDS metadata export."
                        ),
                        source_evidence=f"{sheet_name}!row={row_number}",
                    )
                )

            required, cardinality = _cardinality_from_occurs(
                _cds_metadata_value(normalized_row, "min_occurs"),
                _cds_metadata_value(normalized_row, "max_occurs"),
            )
            is_key = _parse_catalog_bool(_cds_metadata_value(normalized_row, "key")) is True
            association_target = _cds_metadata_value(normalized_row, "association_target")
            element_kind = _cds_metadata_value(normalized_row, "element_kind")
            associations = [association_target] if association_target else []
            fields.append(
                NormalizedSchemaField(
                    entity_name=artifact_name,
                    field_path=element_name,
                    parent_path=None,
                    data_type=_cds_metadata_value(normalized_row, "data_type"),
                    required=required,
                    cardinality=cardinality,
                    length=_parse_catalog_int(_cds_metadata_value(normalized_row, "length")),
                    precision=_parse_catalog_int(_cds_metadata_value(normalized_row, "precision")),
                    scale=_parse_catalog_int(_cds_metadata_value(normalized_row, "scale")),
                    is_key=is_key,
                    enumerations=[],
                    associations=associations,
                    annotations={
                        "artifact_kind": artifact_kind,
                        "element_kind": element_kind,
                        "association_target": association_target,
                        "source_file_format": source_file_format,
                        "source_sheet": sheet_name,
                    },
                    description=_cds_metadata_value(normalized_row, "description"),
                    source_evidence=f"{sheet_name}!row={row_number}",
                    warnings=[],
                )
            )

    source_identity = next(iter(artifacts)) if len(artifacts) == 1 else path.stem
    if len(artifacts) > 1:
        warnings.append(
            f"CDS metadata export contains multiple artifacts ({len(artifacts)}); using "
            f"'{path.stem}' as source identity."
        )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="cds_metadata_export",
        source_identity=source_identity,
        source_version=None,
        namespace=next(iter(namespaces)) if len(namespaces) == 1 else None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        warnings=warnings,
    )


def _inspect_field_catalog_rows(
    *,
    path: Path,
    rows_by_sheet: dict[str, list[dict[str, str]]],
    source_file_format: str,
) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []
    seen_entities: set[str] = set()

    for sheet_name, rows in rows_by_sheet.items():
        default_entity = path.stem if source_file_format == "csv" else sheet_name
        if not rows:
            warnings.append(f"Field catalogue section '{sheet_name}' is empty.")
            continue

        normalized_headers = {_normalize_header_name(header) for header in rows[0]}
        if not normalized_headers & _FIELD_CATALOG_ALIASES["field"]:
            warnings.append(
                f"Field catalogue section '{sheet_name}' has no recognized field-name column."
            )
            continue

        for row_number, row in enumerate(rows, start=2):
            normalized_row = {
                _normalize_header_name(key): str(value).strip()
                for key, value in row.items()
                if key is not None
            }
            field_name = _catalog_value(normalized_row, "field")
            if field_name is None:
                warnings.append(
                    f"Field catalogue section '{sheet_name}' row {row_number} has no field name."
                )
                continue

            entity_name = _catalog_value(normalized_row, "entity") or default_entity
            if entity_name not in seen_entities:
                seen_entities.add(entity_name)
                entities.append(
                    NormalizedSchemaEntity(
                        name=entity_name,
                        kind="entity",
                        description=f"Imported from {source_file_format} field catalogue.",
                        source_evidence=f"{sheet_name}!row={row_number}",
                    )
                )

            data_type = _catalog_value(normalized_row, "data_type")
            description = _catalog_value(normalized_row, "description")
            required_value = _catalog_value(normalized_row, "required")
            nullable_value = _catalog_value(normalized_row, "nullable")
            required = _parse_catalog_bool(required_value)
            if required is None:
                nullable = _parse_catalog_bool(nullable_value)
                required = None if nullable is None else not nullable
            enumerations = _split_catalog_values(_catalog_value(normalized_row, "enumerations"))
            cardinality = _catalog_value(normalized_row, "cardinality")
            if cardinality is None and required is not None:
                cardinality = _cardinality(required, data_type or "string")
            fields.append(
                NormalizedSchemaField(
                    entity_name=entity_name,
                    field_path=field_name,
                    parent_path=field_name.rsplit(".", 1)[0] if "." in field_name else None,
                    data_type=data_type,
                    required=required,
                    cardinality=cardinality,
                    length=_parse_catalog_int(_catalog_value(normalized_row, "length")),
                    precision=_parse_catalog_int(_catalog_value(normalized_row, "precision")),
                    scale=_parse_catalog_int(_catalog_value(normalized_row, "scale")),
                    is_key=_parse_catalog_bool(_catalog_value(normalized_row, "key")) is True
                    or field_name.split(".")[-1].replace("[]", "").lower() == "id",
                    enumerations=enumerations,
                    annotations={
                        "source_file_format": source_file_format,
                        "source_sheet": sheet_name,
                    },
                    description=description,
                    source_evidence=f"{sheet_name}!row={row_number}",
                )
            )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="field_catalog",
        source_identity=path.stem,
        source_version=None,
        namespace=None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        warnings=warnings,
    )


def _inspect_mapping_sections(
    *,
    path: Path,
    sections: list[dict[str, Any]],
) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []
    seen_entities: set[str] = set()

    for section in sections:
        sheet_name = str(section["sheet_name"])
        rows = list(section["rows"])
        if not rows:
            warnings.append(f"Mapping section '{section['table_id']}' is empty.")
            continue

        for row in rows:
            row_number = row.get("_row_number", "?")
            normalized_row = {
                _normalize_header_name(key): str(value).strip()
                for key, value in row.items()
                if key != "_row_number" and key is not None
            }
            target_field = _mapping_value(normalized_row, "target_field")
            if target_field is None:
                warnings.append(
                    f"Mapping section '{section['table_id']}' row {row_number} has no target field."
                )
                continue

            entity_name = (
                _mapping_value(normalized_row, "target_table")
                or _mapping_value(normalized_row, "target_system")
                or sheet_name
            )
            if entity_name not in seen_entities:
                seen_entities.add(entity_name)
                entities.append(
                    NormalizedSchemaEntity(
                        name=entity_name,
                        kind="entity",
                        description="Imported from SAP-style mapping workbook evidence.",
                        source_evidence=f"{sheet_name}!row={row_number}",
                    )
                )

            source_field = _mapping_value(normalized_row, "source_field")
            required = _parse_catalog_bool(_mapping_value(normalized_row, "required"))
            data_type = _mapping_value(normalized_row, "data_type")
            description = _mapping_value(normalized_row, "description") or _mapping_value(
                normalized_row, "comment"
            )
            fields.append(
                NormalizedSchemaField(
                    entity_name=entity_name,
                    field_path=target_field,
                    parent_path=target_field.rsplit(".", 1)[0] if "." in target_field else None,
                    data_type=data_type,
                    required=required,
                    cardinality=_cardinality(required, data_type or "string")
                    if required is not None
                    else None,
                    is_key=target_field.split(".")[-1].lower() == "id",
                    annotations={
                        "source_field": source_field,
                        "source_system": _mapping_value(normalized_row, "source_system"),
                        "target_system": _mapping_value(normalized_row, "target_system"),
                        "mapping_rule": _mapping_value(normalized_row, "rule"),
                        "mapping_status": _mapping_value(normalized_row, "status"),
                        "owner": _mapping_value(normalized_row, "owner"),
                        "mapping_table_id": section["table_id"],
                    },
                    description=description,
                    source_evidence=f"{sheet_name}!row={row_number}",
                )
            )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="sap_mapping_workbook",
        source_identity=path.stem,
        source_version=None,
        namespace=None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        warnings=warnings,
    )


def _inspect_migration_cockpit_rows(
    *,
    path: Path,
    rows_by_sheet: dict[str, list[dict[str, str]]],
) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []
    seen_entities: set[str] = set()

    for sheet_name, rows in rows_by_sheet.items():
        for row_number, row in enumerate(rows, start=2):
            normalized_row = {
                _normalize_header_name(key): str(value).strip()
                for key, value in row.items()
                if key is not None
            }
            field_name = _migration_cockpit_value(normalized_row, "field")
            if field_name is None:
                warnings.append(
                    f"Migration Cockpit sheet '{sheet_name}' row {row_number} has no field name."
                )
                continue

            migration_object = _migration_cockpit_value(normalized_row, "migration_object")
            structure = _migration_cockpit_value(normalized_row, "structure")
            entity_name = structure or migration_object or sheet_name
            if entity_name not in seen_entities:
                seen_entities.add(entity_name)
                description = (
                    f"Imported from Migration Cockpit workbook evidence for "
                    f"{migration_object or sheet_name}."
                )
                entities.append(
                    NormalizedSchemaEntity(
                        name=entity_name,
                        kind="entity",
                        description=description,
                        source_evidence=f"{sheet_name}!row={row_number}",
                    )
                )

            description = _migration_cockpit_value(normalized_row, "description")
            data_type = _migration_cockpit_value(normalized_row, "data_type")
            required = _parse_catalog_bool(_migration_cockpit_value(normalized_row, "required"))
            enumerations = _split_catalog_values(
                _migration_cockpit_value(normalized_row, "enumerations")
            )
            fields.append(
                NormalizedSchemaField(
                    entity_name=entity_name,
                    field_path=field_name,
                    parent_path=field_name.rsplit(".", 1)[0] if "." in field_name else None,
                    data_type=data_type,
                    required=required,
                    cardinality=_cardinality(required, data_type or "string")
                    if required is not None
                    else None,
                    length=_parse_catalog_int(_migration_cockpit_value(normalized_row, "length")),
                    is_key=field_name.split(".")[-1].replace("[]", "").lower() == "id",
                    enumerations=enumerations,
                    annotations={
                        "migration_object": migration_object,
                        "structure": structure,
                        "source_sheet": sheet_name,
                        "template_type": "sap_migration_cockpit",
                    },
                    description=description,
                    source_evidence=f"{sheet_name}!row={row_number}",
                )
            )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="sap_migration_cockpit_template",
        source_identity=path.stem,
        source_version=None,
        namespace=None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        warnings=warnings,
    )


def _infer_scalar_text_value(value: str) -> Any:
    lower = value.lower()
    if lower in {"true", "false"}:
        return lower == "true"
    if lower in {"null", "none"}:
        return None
    try:
        if value.isdigit() or (value.startswith("-") and value[1:].isdigit()):
            return int(value)
    except (AttributeError, ValueError):
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _append_xml_payload_field(
    *,
    entity_name: str,
    field_path: str,
    data_type: str,
    required: bool | None,
    cardinality: str | None,
    source_evidence: str,
    annotations: dict[str, Any] | None,
    fields: list[NormalizedSchemaField],
) -> None:
    fields.append(
        NormalizedSchemaField(
            entity_name=entity_name,
            field_path=field_path,
            parent_path=field_path.rsplit(".", 1)[0] if "." in field_path else None,
            data_type=data_type,
            required=required,
            cardinality=cardinality,
            length=None,
            precision=None,
            scale=None,
            is_key=field_path.split(".")[-1].replace("[]", "").lstrip("@").lower() == "id",
            annotations=annotations or {},
            description=None,
            source_evidence=source_evidence,
            warnings=[],
        )
    )


def _walk_xml_payload_element(
    *,
    entity_name: str,
    element: ET.Element,
    parent_path: str | None,
    source_pointer: str,
    required: bool | None,
    fields: list[NormalizedSchemaField],
    warnings: list[str],
) -> None:
    element_name = _xml_local_name(element.tag)
    field_path = element_name if parent_path is None else f"{parent_path}.{element_name}"
    child_elements = list(element)
    text_value = (element.text or "").strip()
    has_mixed_text = bool(text_value and child_elements)

    if has_mixed_text:
        warnings.append(f"Mixed XML content at {source_pointer} requires manual review.")

    if child_elements:
        _append_xml_payload_field(
            entity_name=entity_name,
            field_path=field_path,
            data_type="object",
            required=required,
            cardinality=_cardinality(required, "object"),
            source_evidence=source_pointer,
            annotations={"xml_tag": element_name},
            fields=fields,
        )
    elif text_value:
        scalar = _infer_scalar_text_value(text_value)
        data_type = _payload_value_type(scalar)
        _append_xml_payload_field(
            entity_name=entity_name,
            field_path=field_path,
            data_type=data_type,
            required=required,
            cardinality=_cardinality(required, data_type),
            source_evidence=source_pointer,
            annotations={"xml_tag": element_name},
            fields=fields,
        )
    else:
        _append_xml_payload_field(
            entity_name=entity_name,
            field_path=field_path,
            data_type="null",
            required=required,
            cardinality=_cardinality(required, "null"),
            source_evidence=source_pointer,
            annotations={"xml_tag": element_name},
            fields=fields,
        )

    for attr_name, attr_value in element.attrib.items():
        attr_path = f"{field_path}.@{attr_name}"
        scalar = _infer_scalar_text_value(attr_value)
        data_type = _payload_value_type(scalar)
        _append_xml_payload_field(
            entity_name=entity_name,
            field_path=attr_path,
            data_type=data_type,
            required=True,
            cardinality="1..1",
            source_evidence=f"{source_pointer}/@{attr_name}",
            annotations={"xml_attribute": attr_name},
            fields=fields,
        )

    if not child_elements:
        return

    grouped_children: dict[str, list[ET.Element]] = defaultdict(list)
    for child in child_elements:
        grouped_children[_xml_local_name(child.tag)].append(child)

    for child_name, grouped in grouped_children.items():
        child_pointer = f"{source_pointer}/{child_name}"
        if len(grouped) == 1:
            _walk_xml_payload_element(
                entity_name=entity_name,
                element=grouped[0],
                parent_path=field_path,
                source_pointer=child_pointer,
                required=True,
                fields=fields,
                warnings=warnings,
            )
            continue

        array_path = f"{field_path}.{child_name}"
        _append_xml_payload_field(
            entity_name=entity_name,
            field_path=array_path,
            data_type="array",
            required=True,
            cardinality="1..*",
            source_evidence=child_pointer,
            annotations={"xml_tag": child_name},
            fields=fields,
        )

        observed_child_types = Counter(
            _payload_value_type((child.text or "").strip())
            if not list(child) and not child.attrib
            else "object"
            for child in grouped
        )
        if len(observed_child_types) > 1:
            warnings.append(
                f"Heterogeneous repeated XML elements at {child_pointer} use representative "
                "item inference."
            )

        representative = grouped[0]
        representative_field_path = f"{array_path}[]"
        rep_children = list(representative)
        rep_text = (representative.text or "").strip()
        if rep_children:
            _append_xml_payload_field(
                entity_name=entity_name,
                field_path=representative_field_path,
                data_type="object",
                required=None,
                cardinality=None,
                source_evidence=f"{child_pointer}/*",
                annotations={"xml_tag": child_name},
                fields=fields,
            )
            for attr_name, attr_value in representative.attrib.items():
                scalar = _infer_scalar_text_value(attr_value)
                _append_xml_payload_field(
                    entity_name=entity_name,
                    field_path=f"{representative_field_path}.@{attr_name}",
                    data_type=_payload_value_type(scalar),
                    required=True,
                    cardinality="1..1",
                    source_evidence=f"{child_pointer}/*/@{attr_name}",
                    annotations={"xml_attribute": attr_name},
                    fields=fields,
                )
            nested_grouped: dict[str, list[ET.Element]] = defaultdict(list)
            for nested in rep_children:
                nested_grouped[_xml_local_name(nested.tag)].append(nested)
            for nested_name, nested_children in nested_grouped.items():
                if len(nested_children) == 1:
                    _walk_xml_payload_element(
                        entity_name=entity_name,
                        element=nested_children[0],
                        parent_path=representative_field_path,
                        source_pointer=f"{child_pointer}/*/{nested_name}",
                        required=True,
                        fields=fields,
                        warnings=warnings,
                    )
                else:
                    warnings.append(
                        f"Nested repeated XML elements at {child_pointer}/{nested_name} use "
                        "representative item inference."
                    )
                    _append_xml_payload_field(
                        entity_name=entity_name,
                        field_path=f"{representative_field_path}.{nested_name}",
                        data_type="array",
                        required=True,
                        cardinality="1..*",
                        source_evidence=f"{child_pointer}/*/{nested_name}",
                        annotations={"xml_tag": nested_name},
                        fields=fields,
                    )
                    _walk_xml_payload_element(
                        entity_name=entity_name,
                        element=nested_children[0],
                        parent_path=representative_field_path,
                        source_pointer=f"{child_pointer}/*/{nested_name}",
                        required=None,
                        fields=fields,
                        warnings=warnings,
                    )
            continue

        scalar = _infer_scalar_text_value(rep_text)
        data_type = _payload_value_type(scalar)
        _append_xml_payload_field(
            entity_name=entity_name,
            field_path=representative_field_path,
            data_type=data_type,
            required=None,
            cardinality=None,
            source_evidence=f"{child_pointer}/*",
            annotations={"xml_tag": child_name},
            fields=fields,
        )


def _schema_type(node: dict[str, Any]) -> str | None:
    if "$ref" in node:
        return "ref"
    explicit = node.get("type")
    if explicit:
        return str(explicit)
    if "properties" in node:
        return "object"
    if "items" in node:
        return "array"
    return None


def _cardinality(required: bool | None, data_type: str | None) -> str | None:
    if required is None:
        return None
    if data_type == "array":
        return "1..*" if required else "0..*"
    return "1..1" if required else "0..1"


def _walk_json_schema(
    *,
    entity_name: str,
    schema: dict[str, Any],
    source_pointer: str,
    field_path: str | None,
    required: bool | None,
    fields: list[NormalizedSchemaField],
    warnings: list[str],
) -> None:
    data_type = _schema_type(schema)
    associations: list[str] = []
    field_warnings: list[str] = []
    if "$ref" in schema:
        ref = str(schema["$ref"])
        associations.append(ref)
        if ref.startswith("http://") or ref.startswith("https://"):
            warning = f"Remote reference not resolved: {ref}"
            warnings.append(warning)
            field_warnings.append(warning)
    if field_path is not None:
        fields.append(
            NormalizedSchemaField(
                entity_name=entity_name,
                field_path=field_path,
                parent_path=field_path.rsplit(".", 1)[0] if "." in field_path else None,
                data_type=data_type,
                required=required,
                cardinality=_cardinality(required, data_type),
                length=schema.get("maxLength"),
                precision=schema.get("maximum"),
                scale=None,
                is_key=field_path.split(".")[-1].lower() == "id",
                enumerations=[str(value) for value in schema.get("enum", [])],
                associations=associations,
                annotations=_normalize_annotations(schema),
                description=schema.get("description"),
                source_evidence=source_pointer,
                warnings=field_warnings,
            )
        )

    properties = schema.get("properties")
    if isinstance(properties, dict):
        required_fields = set(schema.get("required", []))
        for name, child in properties.items():
            child_path = name if field_path is None else f"{field_path}.{name}"
            _walk_json_schema(
                entity_name=entity_name,
                schema=child or {},
                source_pointer=f"{source_pointer}/properties/{name}",
                field_path=child_path,
                required=name in required_fields,
                fields=fields,
                warnings=warnings,
            )

    items = schema.get("items")
    if isinstance(items, dict) and field_path is not None:
        item_path = f"{field_path}[]"
        _walk_json_schema(
            entity_name=entity_name,
            schema=items,
            source_pointer=f"{source_pointer}/items",
            field_path=item_path,
            required=None,
            fields=fields,
            warnings=warnings,
        )

    for composition_key in ("allOf", "anyOf", "oneOf"):
        if composition_key in schema:
            warnings.append(
                f"{composition_key} composition in {source_pointer} requires manual review."
            )


def inspect_json_schema(path: Path, document: dict[str, Any]) -> NormalizedSchemaDocument:
    root_name = str(document.get("title") or path.stem)
    checksum = _file_checksum(path)
    warnings: list[str] = []
    fields: list[NormalizedSchemaField] = []
    _walk_json_schema(
        entity_name=root_name,
        schema=document,
        source_pointer="#",
        field_path=None,
        required=None,
        fields=fields,
        warnings=warnings,
    )
    namespace = document.get("$id") or document.get("$schema")
    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="json_schema",
        source_identity=root_name,
        source_version=document.get("$schema"),
        namespace=str(namespace) if namespace else None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=[
            NormalizedSchemaEntity(
                name=root_name,
                kind="entity",
                description=document.get("description"),
                source_evidence="#",
            )
        ],
        fields=fields,
        warnings=warnings,
    )


def inspect_json_payload(path: Path, document: Any) -> NormalizedSchemaDocument:
    source_identity = path.stem
    checksum = _file_checksum(path)
    warnings: list[str] = []
    fields: list[NormalizedSchemaField] = []

    if isinstance(document, dict):
        for key, value in document.items():
            _walk_json_payload(
                entity_name=source_identity,
                value=value,
                source_pointer=f"#/{key}",
                field_path=key,
                required=True,
                fields=fields,
                warnings=warnings,
            )
    elif isinstance(document, list):
        _walk_json_payload(
            entity_name=source_identity,
            value=document,
            source_pointer="#",
            field_path="items",
            required=True,
            fields=fields,
            warnings=warnings,
        )
    else:
        _walk_json_payload(
            entity_name=source_identity,
            value=document,
            source_pointer="#",
            field_path="value",
            required=True,
            fields=fields,
            warnings=warnings,
        )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="json_payload",
        source_identity=source_identity,
        source_version=None,
        namespace=None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=[
            NormalizedSchemaEntity(
                name=source_identity,
                kind="message_instance",
                description="Observed JSON payload structure.",
                source_evidence="#",
            )
        ],
        fields=fields,
        warnings=warnings,
    )


def inspect_cds_metadata_json(path: Path, document: Any) -> NormalizedSchemaDocument:
    rowset = _extract_json_rowset(document)
    if rowset is None or not _looks_like_cds_metadata_rows(rowset):
        raise SchemaInspectionError("JSON document is not a supported CDS metadata export.")
    return _inspect_cds_metadata_rows(
        path=path,
        rows_by_sheet={path.stem: rowset},
        source_file_format="json",
    )


def inspect_field_catalog_csv(path: Path) -> NormalizedSchemaDocument:
    rows_by_sheet = _read_csv_catalog_rows(path)
    cds_rows = _detect_cds_metadata_rows(rows_by_sheet)
    if cds_rows:
        return _inspect_cds_metadata_rows(
            path=path,
            rows_by_sheet=cds_rows,
            source_file_format="csv",
        )
    return _inspect_field_catalog_rows(
        path=path,
        rows_by_sheet=rows_by_sheet,
        source_file_format="csv",
    )


def inspect_field_catalog_xlsx(path: Path) -> NormalizedSchemaDocument:
    return _inspect_field_catalog_rows(
        path=path,
        rows_by_sheet=_read_xlsx_catalog_rows(path),
        source_file_format="xlsx",
    )


def inspect_xlsx_schema_evidence(path: Path) -> NormalizedSchemaDocument:
    mapping_sections = _read_xlsx_mapping_sections(path)
    if mapping_sections:
        return _inspect_mapping_sections(path=path, sections=mapping_sections)
    rows_by_sheet = _read_xlsx_catalog_rows(path)
    migration_cockpit_rows = _detect_migration_cockpit_rows(rows_by_sheet)
    if migration_cockpit_rows:
        return _inspect_migration_cockpit_rows(
            path=path,
            rows_by_sheet=migration_cockpit_rows,
        )
    cds_rows = _detect_cds_metadata_rows(rows_by_sheet)
    if cds_rows:
        return _inspect_cds_metadata_rows(
            path=path,
            rows_by_sheet=cds_rows,
            source_file_format="xlsx",
        )
    return _inspect_field_catalog_rows(
        path=path,
        rows_by_sheet=rows_by_sheet,
        source_file_format="xlsx",
    )


def inspect_xml_payload(path: Path, document: ET.Element) -> NormalizedSchemaDocument:
    source_identity = _xml_local_name(document.tag)
    checksum = _file_checksum(path)
    warnings: list[str] = []
    fields: list[NormalizedSchemaField] = []
    namespace = _xml_namespace(document.tag) or None

    for attr_name, attr_value in document.attrib.items():
        scalar = _infer_scalar_text_value(attr_value)
        _append_xml_payload_field(
            entity_name=source_identity,
            field_path=f"@{attr_name}",
            data_type=_payload_value_type(scalar),
            required=True,
            cardinality="1..1",
            source_evidence=f"#/@{attr_name}",
            annotations={"xml_attribute": attr_name},
            fields=fields,
        )

    child_elements = list(document)
    root_text = (document.text or "").strip()
    if child_elements:
        grouped_children: dict[str, list[ET.Element]] = defaultdict(list)
        for child in child_elements:
            grouped_children[_xml_local_name(child.tag)].append(child)
        for child_name, grouped in grouped_children.items():
            if len(grouped) == 1:
                _walk_xml_payload_element(
                    entity_name=source_identity,
                    element=grouped[0],
                    parent_path=None,
                    source_pointer=f"#/{child_name}",
                    required=True,
                    fields=fields,
                    warnings=warnings,
                )
                continue

            _append_xml_payload_field(
                entity_name=source_identity,
                field_path=child_name,
                data_type="array",
                required=True,
                cardinality="1..*",
                source_evidence=f"#/{child_name}",
                annotations={"xml_tag": child_name},
                fields=fields,
            )
            representative = grouped[0]
            rep_children = list(representative)
            rep_text = (representative.text or "").strip()
            rep_path = f"{child_name}[]"
            if rep_children:
                _append_xml_payload_field(
                    entity_name=source_identity,
                    field_path=rep_path,
                    data_type="object",
                    required=None,
                    cardinality=None,
                    source_evidence=f"#/{child_name}/*",
                    annotations={"xml_tag": child_name},
                    fields=fields,
                )
                for nested in rep_children:
                    _walk_xml_payload_element(
                        entity_name=source_identity,
                        element=nested,
                        parent_path=rep_path,
                        source_pointer=f"#/{child_name}/*/{_xml_local_name(nested.tag)}",
                        required=True,
                        fields=fields,
                        warnings=warnings,
                    )
            else:
                scalar = _infer_scalar_text_value(rep_text)
                _append_xml_payload_field(
                    entity_name=source_identity,
                    field_path=rep_path,
                    data_type=_payload_value_type(scalar),
                    required=None,
                    cardinality=None,
                    source_evidence=f"#/{child_name}/*",
                    annotations={"xml_tag": child_name},
                    fields=fields,
                )
    elif root_text:
        scalar = _infer_scalar_text_value(root_text)
        _append_xml_payload_field(
            entity_name=source_identity,
            field_path="value",
            data_type=_payload_value_type(scalar),
            required=True,
            cardinality="1..1",
            source_evidence="#",
            annotations={"xml_tag": source_identity},
            fields=fields,
        )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="xml_payload",
        source_identity=source_identity,
        source_version=None,
        namespace=namespace,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=[
            NormalizedSchemaEntity(
                name=source_identity,
                kind="message_instance",
                description="Observed XML payload structure.",
                source_evidence="#",
            )
        ],
        fields=fields,
        warnings=warnings,
    )


def _idoc_identity(path: Path, document: ET.Element, idoc_root: ET.Element) -> str:
    root_name = _xml_local_name(document.tag)
    if root_name != "IDOC":
        return root_name

    for child in idoc_root:
        if _xml_local_name(child.tag) != "EDI_DC40":
            continue
        for candidate in ("IDOCTYP", "MESTYP", "CIMTYP"):
            for control_field in child:
                if _xml_local_name(control_field.tag) != candidate:
                    continue
                value = (control_field.text or "").strip()
                if value:
                    return value
    return path.stem


def _append_idoc_field(
    *,
    basic_type: str,
    segment_name: str,
    segment_path: str,
    parent_segment: str | None,
    field_path: str,
    data_type: str,
    required: bool | None,
    cardinality: str | None,
    source_evidence: str,
    annotations: dict[str, Any] | None,
    fields: list[NormalizedSchemaField],
    seen_fields: set[tuple[str, str]],
) -> None:
    key = (segment_name, field_path)
    if key in seen_fields:
        return
    seen_fields.add(key)
    merged_annotations = {
        "idoc_basic_type": basic_type,
        "idoc_segment": segment_name,
        "idoc_segment_path": segment_path,
    }
    if parent_segment is not None:
        merged_annotations["idoc_parent_segment"] = parent_segment
    if annotations:
        merged_annotations.update(annotations)
    _append_xml_payload_field(
        entity_name=segment_name,
        field_path=field_path,
        data_type=data_type,
        required=required,
        cardinality=cardinality,
        source_evidence=source_evidence,
        annotations=merged_annotations,
        fields=fields,
    )


def _append_idoc_entity(
    *,
    basic_type: str,
    segment_name: str,
    parent_segment: str | None,
    source_evidence: str,
    entities: list[NormalizedSchemaEntity],
    seen_entities: set[str],
) -> None:
    if segment_name in seen_entities:
        return
    seen_entities.add(segment_name)
    if segment_name == "EDI_DC40":
        description = f"Observed IDoc control record for {basic_type}."
        kind = "idoc_control_record"
    elif parent_segment is None:
        description = f"Observed top-level IDoc segment for {basic_type}."
        kind = "idoc_segment"
    else:
        description = f"Observed nested IDoc segment under {parent_segment} for {basic_type}."
        kind = "idoc_segment"
    entities.append(
        NormalizedSchemaEntity(
            name=segment_name,
            kind=kind,
            description=description,
            source_evidence=source_evidence,
        )
    )


def _walk_idoc_field_element(
    *,
    basic_type: str,
    segment_name: str,
    segment_path: str,
    parent_segment: str | None,
    element: ET.Element,
    field_path: str,
    source_pointer: str,
    fields: list[NormalizedSchemaField],
    seen_fields: set[tuple[str, str]],
    warnings: list[str],
) -> None:
    child_elements = list(element)
    nested_segments = [child for child in child_elements if _is_probable_idoc_segment(child)]
    text_value = (element.text or "").strip()

    if nested_segments:
        warnings.append(
            f"Mixed IDoc field/segment content at {source_pointer} requires manual review."
        )

    if child_elements and not nested_segments:
        _append_idoc_field(
            basic_type=basic_type,
            segment_name=segment_name,
            segment_path=segment_path,
            parent_segment=parent_segment,
            field_path=field_path,
            data_type="object",
            required=True,
            cardinality="1..1",
            source_evidence=source_pointer,
            annotations={"xml_tag": _xml_local_name(element.tag)},
            fields=fields,
            seen_fields=seen_fields,
        )
        for child in child_elements:
            child_name = _xml_local_name(child.tag)
            _walk_idoc_field_element(
                basic_type=basic_type,
                segment_name=segment_name,
                segment_path=segment_path,
                parent_segment=parent_segment,
                element=child,
                field_path=f"{field_path}.{child_name}",
                source_pointer=f"{source_pointer}/{child_name}",
                fields=fields,
                seen_fields=seen_fields,
                warnings=warnings,
            )
    elif text_value:
        scalar = _infer_scalar_text_value(text_value)
        _append_idoc_field(
            basic_type=basic_type,
            segment_name=segment_name,
            segment_path=segment_path,
            parent_segment=parent_segment,
            field_path=field_path,
            data_type=_payload_value_type(scalar),
            required=True,
            cardinality="1..1",
            source_evidence=source_pointer,
            annotations={"xml_tag": _xml_local_name(element.tag)},
            fields=fields,
            seen_fields=seen_fields,
        )
    else:
        _append_idoc_field(
            basic_type=basic_type,
            segment_name=segment_name,
            segment_path=segment_path,
            parent_segment=parent_segment,
            field_path=field_path,
            data_type="null",
            required=True,
            cardinality="1..1",
            source_evidence=source_pointer,
            annotations={"xml_tag": _xml_local_name(element.tag)},
            fields=fields,
            seen_fields=seen_fields,
        )

    for attr_name, attr_value in element.attrib.items():
        scalar = _infer_scalar_text_value(attr_value)
        _append_idoc_field(
            basic_type=basic_type,
            segment_name=segment_name,
            segment_path=segment_path,
            parent_segment=parent_segment,
            field_path=f"{field_path}.@{attr_name}",
            data_type=_payload_value_type(scalar),
            required=True,
            cardinality="1..1",
            source_evidence=f"{source_pointer}/@{attr_name}",
            annotations={"xml_attribute": attr_name},
            fields=fields,
            seen_fields=seen_fields,
        )


def _walk_idoc_segment(
    *,
    basic_type: str,
    segment: ET.Element,
    segment_path: str,
    parent_segment: str | None,
    source_pointer: str,
    entities: list[NormalizedSchemaEntity],
    seen_entities: set[str],
    fields: list[NormalizedSchemaField],
    seen_fields: set[tuple[str, str]],
    warnings: list[str],
) -> None:
    segment_name = _xml_local_name(segment.tag)
    _append_idoc_entity(
        basic_type=basic_type,
        segment_name=segment_name,
        parent_segment=parent_segment,
        source_evidence=source_pointer,
        entities=entities,
        seen_entities=seen_entities,
    )

    nested_segments: dict[str, list[ET.Element]] = defaultdict(list)
    for child in segment:
        child_name = _xml_local_name(child.tag)
        if _is_probable_idoc_segment(child):
            nested_segments[child_name].append(child)
            continue
        _walk_idoc_field_element(
            basic_type=basic_type,
            segment_name=segment_name,
            segment_path=segment_path,
            parent_segment=parent_segment,
            element=child,
            field_path=child_name,
            source_pointer=f"{source_pointer}/{child_name}",
            fields=fields,
            seen_fields=seen_fields,
            warnings=warnings,
        )

    for nested_name, grouped in nested_segments.items():
        if len(grouped) > 1:
            warnings.append(
                f"Repeated IDoc segment at {source_pointer}/{nested_name} uses representative "
                "item inference."
            )
        _walk_idoc_segment(
            basic_type=basic_type,
            segment=grouped[0],
            segment_path=f"{segment_path}.{nested_name}",
            parent_segment=segment_name,
            source_pointer=f"{source_pointer}/{nested_name}",
            entities=entities,
            seen_entities=seen_entities,
            fields=fields,
            seen_fields=seen_fields,
            warnings=warnings,
        )


def inspect_idoc_payload(path: Path, document: ET.Element) -> NormalizedSchemaDocument:
    idoc_root = _find_idoc_container(document)
    if idoc_root is None:
        raise SchemaInspectionError("IDoc payload could not be located in XML document.")

    checksum = _file_checksum(path)
    source_identity = _idoc_identity(path, document, idoc_root)
    namespace = _xml_namespace(document.tag) or _xml_namespace(idoc_root.tag) or None
    warnings = [
        "IDoc payload inference uses representative observed segments and does not prove full "
        "optionality or cardinality."
    ]
    entities: list[NormalizedSchemaEntity] = []
    seen_entities: set[str] = set()
    fields: list[NormalizedSchemaField] = []
    seen_fields: set[tuple[str, str]] = set()

    root_pointer = "#" if _xml_local_name(document.tag) == "IDOC" else "#/IDOC"
    root_segment_groups: dict[str, list[ET.Element]] = defaultdict(list)
    for child in idoc_root:
        if _is_probable_idoc_segment(child):
            root_segment_groups[_xml_local_name(child.tag)].append(child)

    for segment_name, grouped in root_segment_groups.items():
        if len(grouped) > 1:
            warnings.append(
                f"Repeated IDoc segment at {root_pointer}/{segment_name} uses representative "
                "item inference."
            )
        _walk_idoc_segment(
            basic_type=source_identity,
            segment=grouped[0],
            segment_path=f"IDOC.{segment_name}",
            parent_segment=None,
            source_pointer=f"{root_pointer}/{segment_name}",
            entities=entities,
            seen_entities=seen_entities,
            fields=fields,
            seen_fields=seen_fields,
            warnings=warnings,
        )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="idoc_payload",
        source_identity=source_identity,
        source_version=None,
        namespace=namespace,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        warnings=warnings,
    )


def inspect_we60_html(path: Path) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []
    seen_entities: set[str] = set()
    seen_fields: set[tuple[str, str]] = set()

    root_document = _parse_html_document(path)
    linked_paths: list[Path] = []
    seen_linked_paths: set[Path] = set()
    for link in root_document.links:
        candidate = _resolve_local_html_link(path, link.href)
        if candidate is None:
            if link.href.lower().endswith((".html", ".htm")):
                warnings.append(f"Linked WE60 page could not be resolved locally: {link.href}")
            continue
        if candidate == path or candidate in seen_linked_paths:
            continue
        seen_linked_paths.add(candidate)
        linked_paths.append(candidate)

    parsed_documents: list[tuple[Path, _HtmlDocument]] = [(path, root_document)]
    for linked_path in linked_paths:
        parsed_documents.append((linked_path, _parse_html_document(linked_path)))

    if not any(_looks_like_we60_html(document) for _, document in parsed_documents):
        raise SchemaInspectionError(
            "Unsupported HTML schema document: expected SAP WE60-style IDoc documentation."
        )

    bundle_text = " ".join(
        filter(
            None,
            [
                root_document.title,
                *root_document.headings,
                *(linked.title or "" for _, linked in parsed_documents[1:]),
            ],
        )
    )
    source_identity = _extract_basic_type_name(bundle_text) or path.stem.upper()

    for document_path, document in parsed_documents:
        document_hint = _extract_segment_name(
            " ".join(filter(None, [document.title, *document.headings]))
        ) or _extract_segment_name(document_path.stem)
        for table_index, table in enumerate(document.tables, start=1):
            if not table.headers:
                continue
            segment_idx = _header_index(table.headers, _WE60_SEGMENT_ALIASES)
            field_idx = _header_index(table.headers, _WE60_FIELD_ALIASES)
            description_idx = _header_index(table.headers, _WE60_DESCRIPTION_ALIASES)
            data_type_idx = _header_index(table.headers, _WE60_DATA_TYPE_ALIASES)
            length_idx = _header_index(table.headers, _WE60_LENGTH_ALIASES)
            required_idx = _header_index(table.headers, _WE60_REQUIRED_ALIASES)
            min_idx = _header_index(table.headers, _WE60_MIN_ALIASES)
            max_idx = _header_index(table.headers, _WE60_MAX_ALIASES)
            parent_idx = _header_index(table.headers, _WE60_PARENT_ALIASES)

            if field_idx is None and segment_idx is None:
                continue

            if field_idx is None and segment_idx is not None:
                for row_index, row in enumerate(table.rows, start=2):
                    if segment_idx >= len(row):
                        continue
                    segment_name = _extract_segment_name(row[segment_idx])
                    if segment_name is None:
                        continue
                    if segment_name in seen_entities:
                        continue
                    seen_entities.add(segment_name)
                    description = (
                        row[description_idx].strip()
                        if description_idx is not None and description_idx < len(row)
                        else None
                    )
                    parent_segment = (
                        _extract_segment_name(row[parent_idx])
                        if parent_idx is not None and parent_idx < len(row)
                        else None
                    )
                    kind = "idoc_control_record" if segment_name == "EDI_DC40" else "idoc_segment"
                    if description is None:
                        if parent_segment is None:
                            description = (
                                f"Observed WE60 segment documentation for {segment_name} in "
                                f"{source_identity}."
                            )
                        else:
                            description = (
                                f"Observed WE60 nested segment documentation for {segment_name} "
                                f"under {parent_segment} in {source_identity}."
                            )
                    entities.append(
                        NormalizedSchemaEntity(
                            name=segment_name,
                            kind=kind,
                            description=description,
                            source_evidence=f"{document_path.name}#table={table_index},row={row_index}",
                        )
                    )
                continue

            for row_index, row in enumerate(table.rows, start=2):
                if field_idx is None or field_idx >= len(row):
                    continue
                field_name = row[field_idx].strip()
                if not field_name:
                    continue
                segment_name = None
                if segment_idx is not None and segment_idx < len(row):
                    segment_name = _extract_segment_name(row[segment_idx])
                segment_name = segment_name or document_hint or _extract_segment_name(table.heading)
                if segment_name is None:
                    warnings.append(
                        f"WE60 field row in {document_path.name} table {table_index} "
                        "has no segment context."
                    )
                    continue

                if segment_name not in seen_entities:
                    seen_entities.add(segment_name)
                    kind = "idoc_control_record" if segment_name == "EDI_DC40" else "idoc_segment"
                    entities.append(
                        NormalizedSchemaEntity(
                            name=segment_name,
                            kind=kind,
                            description=f"Observed WE60 segment documentation for {segment_name}.",
                            source_evidence=f"{document_path.name}#table={table_index}",
                        )
                    )

                required = None
                cardinality = None
                if required_idx is not None and required_idx < len(row):
                    required = _parse_catalog_bool(row[required_idx])
                    if required is not None:
                        cardinality = _cardinality(required, data_type=None)
                occurs_required, occurs_cardinality = _cardinality_from_occurs(
                    row[min_idx] if min_idx is not None and min_idx < len(row) else None,
                    row[max_idx] if max_idx is not None and max_idx < len(row) else None,
                )
                if required is None:
                    required = occurs_required
                if cardinality is None:
                    cardinality = occurs_cardinality

                field_key = (segment_name, field_name)
                if field_key in seen_fields:
                    continue
                seen_fields.add(field_key)
                fields.append(
                    NormalizedSchemaField(
                        entity_name=segment_name,
                        field_path=field_name,
                        parent_path=None,
                        data_type=(
                            row[data_type_idx].strip()
                            if data_type_idx is not None and data_type_idx < len(row)
                            else None
                        ),
                        required=required,
                        cardinality=cardinality,
                        length=_parse_catalog_int(
                            row[length_idx]
                            if length_idx is not None and length_idx < len(row)
                            else None
                        ),
                        precision=None,
                        scale=None,
                        is_key=field_name.lower() == "id",
                        enumerations=[],
                        associations=[],
                        annotations={
                            "we60_basic_type": source_identity,
                            "we60_source_page": document_path.name,
                            "we60_table_heading": table.heading,
                        },
                        description=(
                            row[description_idx].strip()
                            if description_idx is not None and description_idx < len(row)
                            else None
                        ),
                        source_evidence=f"{document_path.name}#table={table_index},row={row_index}",
                        warnings=[],
                    )
                )

    if not fields:
        raise SchemaInspectionError(
            "WE60 HTML documentation did not yield any segment field rows for normalization."
        )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="we60_html_documentation",
        source_identity=source_identity,
        source_version=None,
        namespace=None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        warnings=warnings,
    )


def inspect_integration_flow_artifact(path: Path, document: ET.Element) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    flow_name = _iflow_identity(document, path.stem)
    entities, operations, warnings = _inspect_iflow_document(
        source_identity=flow_name,
        flow_name=flow_name,
        source_evidence=path.name,
        document=document,
    )
    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="integration_flow_artifact",
        source_identity=flow_name,
        source_version=None,
        namespace=None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=[],
        operations=operations,
        warnings=warnings,
    )


def inspect_integration_suite_package(path: Path) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []
    operations: list[NormalizedSchemaOperation] = []
    seen_entities: set[str] = set()
    seen_resource_fields: set[str] = set()
    total_uncompressed = 0
    manifest: dict[str, str] = {}

    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        if len(infos) > _ZIP_ENTRY_LIMIT:
            raise SchemaInspectionError(
                f"Integration Suite package has too many ZIP entries ({len(infos)})."
            )

        for info in infos:
            total_uncompressed += info.file_size
            if total_uncompressed > _ZIP_TOTAL_UNCOMPRESSED_LIMIT:
                raise SchemaInspectionError(
                    "Integration Suite package exceeds safe uncompressed size limit."
                )
            pure_path = PurePosixPath(info.filename)
            if pure_path.is_absolute() or ".." in pure_path.parts:
                raise SchemaInspectionError(
                    f"Integration Suite package contains unsafe path: {info.filename}"
                )

        manifest_info = next(
            (
                info
                for info in infos
                if PurePosixPath(info.filename).as_posix().upper().endswith("META-INF/MANIFEST.MF")
            ),
            None,
        )
        if manifest_info is not None:
            manifest = _parse_manifest_text(
                archive.read(manifest_info).decode("utf-8", errors="replace")
            )
        else:
            warnings.append("Integration Suite package has no MANIFEST.MF metadata file.")

        source_identity = (
            manifest.get("Bundle-SymbolicName") or manifest.get("Bundle-Name") or path.stem
        )
        entities.append(
            NormalizedSchemaEntity(
                name=source_identity,
                kind="integration_package",
                description=f"Imported Integration Suite package from {path.name}.",
                source_evidence=path.name,
            )
        )
        seen_entities.add(source_identity)

        for info in infos:
            if info.is_dir():
                continue
            zip_path = PurePosixPath(info.filename).as_posix()
            if zip_path.endswith(".iflw"):
                try:
                    document = ET.fromstring(archive.read(info))
                except ET.ParseError as exc:
                    warnings.append(f"Skipped invalid iFlow XML at {zip_path}: {exc}.")
                    continue
                flow_name = _iflow_identity(document, PurePosixPath(info.filename).stem)
                flow_entities, flow_operations, flow_warnings = _inspect_iflow_document(
                    source_identity=source_identity,
                    flow_name=flow_name,
                    source_evidence=zip_path,
                    document=document,
                )
                for entity in flow_entities:
                    if entity.name not in seen_entities:
                        seen_entities.add(entity.name)
                        entities.append(entity)
                operations.extend(flow_operations)
                warnings.extend(flow_warnings)
                continue

            resource_type = None
            for directory, candidate_type in _INTEGRATION_RESOURCE_DIRS.items():
                if zip_path.startswith(f"{directory}/"):
                    resource_type = candidate_type
                    break
            if resource_type is None:
                continue

            field_path = zip_path.removeprefix("src/main/resources/")
            if field_path in seen_resource_fields:
                continue
            seen_resource_fields.add(field_path)
            fields.append(
                NormalizedSchemaField(
                    entity_name=source_identity,
                    field_path=field_path,
                    parent_path=None,
                    data_type=resource_type,
                    required=True,
                    cardinality="1..1",
                    is_key=False,
                    annotations={"resource_path": zip_path, "resource_type": resource_type},
                    description=(
                        f"Packaged Integration Suite resource {PurePosixPath(zip_path).name}."
                    ),
                    source_evidence=f"{path.name}:{zip_path}",
                )
            )

    if not operations:
        warnings.append("Integration Suite package contains no .iflw integration flow artifacts.")

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="integration_suite_package",
        source_identity=source_identity,
        source_version=manifest.get("Bundle-Version"),
        namespace=manifest.get("Bundle-Vendor"),
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        operations=operations,
        warnings=warnings,
    )


def inspect_openapi(path: Path, document: dict[str, Any]) -> NormalizedSchemaDocument:
    info = document.get("info", {}) or {}
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []
    operations: list[NormalizedSchemaOperation] = []

    schemas = (document.get("components", {}) or {}).get("schemas", {}) or {}
    for schema_name, schema in schemas.items():
        entities.append(
            NormalizedSchemaEntity(
                name=schema_name,
                kind="message_type",
                description=schema.get("description"),
                source_evidence=f"#/components/schemas/{schema_name}",
            )
        )
        _walk_json_schema(
            entity_name=schema_name,
            schema=schema or {},
            source_pointer=f"#/components/schemas/{schema_name}",
            field_path=None,
            required=None,
            fields=fields,
            warnings=warnings,
        )

    def _openapi_schema_name(schema: dict[str, Any]) -> str:
        ref = schema.get("$ref")
        if ref:
            return str(ref).rsplit("/", 1)[-1]
        return str(schema.get("title") or "inline")

    for path_name, path_item in (document.get("paths", {}) or {}).items():
        for method, operation in (path_item or {}).items():
            if method.lower() not in _HTTP_METHODS:
                continue
            operation = operation or {}
            request_schema: str | None = None
            request_body = (operation.get("requestBody", {}) or {}).get("content", {}) or {}
            for media in request_body.values():
                schema = (media or {}).get("schema", {}) or {}
                request_schema = _openapi_schema_name(schema)
                if "$ref" not in schema and schema:
                    warnings.append(
                        f"Inline request schema for {method.upper()} {path_name} is not promoted "
                        "to a named entity in this slice."
                    )
                break

            response_schemas: list[dict[str, str]] = []
            for status_code, response in (operation.get("responses", {}) or {}).items():
                content = (response or {}).get("content", {}) or {}
                for media in content.values():
                    schema = (media or {}).get("schema", {}) or {}
                    if not schema:
                        continue
                    response_schemas.append(
                        {
                            "status_code": str(status_code),
                            "schema": _openapi_schema_name(schema),
                        }
                    )
                    if "$ref" not in schema:
                        warnings.append(
                            f"Inline response schema for {method.upper()} {path_name} "
                            f"status {status_code} is not promoted to a named entity in this slice."
                        )
                    break

            parameters = [
                str(parameter.get("name"))
                for parameter in operation.get("parameters", []) or []
                if parameter.get("name")
            ]
            operations.append(
                NormalizedSchemaOperation(
                    operation_id=str(
                        operation.get("operationId") or f"{method.upper()} {path_name}"
                    ),
                    method=method.upper(),
                    path=path_name,
                    protocol="openapi",
                    request_body_schema=request_schema,
                    response_schemas=response_schemas,
                    parameters=parameters,
                    description=operation.get("summary") or operation.get("description"),
                    source_evidence=f"#/paths/{path_name}/{method}",
                )
            )

    namespace = None
    servers = document.get("servers", []) or []
    if servers:
        namespace = servers[0].get("url")
    if namespace is None:
        namespace = info.get("title")

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="openapi",
        source_identity=str(info.get("title") or path.stem),
        source_version=str(document.get("openapi") or document.get("swagger") or ""),
        namespace=str(namespace) if namespace else None,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        operations=operations,
        warnings=warnings,
    )


def _xsd_type_name(type_name: str | None) -> str | None:
    if not type_name:
        return None
    return type_name.split(":")[-1]


def _xsd_cardinality(min_occurs: str | None, max_occurs: str | None) -> tuple[bool, str]:
    required = min_occurs != "0"
    if max_occurs == "unbounded":
        return required, "1..*" if required else "0..*"
    if max_occurs and max_occurs.isdigit() and int(max_occurs) > 1:
        return required, "1..*" if required else "0..*"
    return required, "1..1" if required else "0..1"


def _xsd_simple_type_details(simple_type: ET.Element | None) -> dict[str, Any]:
    details: dict[str, Any] = {
        "data_type": None,
        "enumerations": [],
        "length": None,
        "precision": None,
        "scale": None,
    }
    if simple_type is None:
        return details

    restriction = simple_type.find("xs:restriction", _XSD_NAMESPACES)
    if restriction is None:
        return details

    details["data_type"] = _xsd_type_name(restriction.attrib.get("base"))
    details["enumerations"] = [
        value.attrib["value"]
        for value in restriction.findall("xs:enumeration", _XSD_NAMESPACES)
        if "value" in value.attrib
    ]
    max_length = restriction.find("xs:maxLength", _XSD_NAMESPACES)
    if max_length is not None and max_length.attrib.get("value", "").isdigit():
        details["length"] = int(max_length.attrib["value"])
    total_digits = restriction.find("xs:totalDigits", _XSD_NAMESPACES)
    if total_digits is not None and total_digits.attrib.get("value", "").isdigit():
        details["precision"] = int(total_digits.attrib["value"])
    fraction_digits = restriction.find("xs:fractionDigits", _XSD_NAMESPACES)
    if fraction_digits is not None and fraction_digits.attrib.get("value", "").isdigit():
        details["scale"] = int(fraction_digits.attrib["value"])
    return details


def _xsd_named_simple_type(
    type_name: str | None,
    simple_types: dict[str, ET.Element],
) -> ET.Element | None:
    local_name = _xsd_type_name(type_name)
    if local_name is None:
        return None
    return simple_types.get(local_name)


def _xsd_named_complex_type(
    type_name: str | None,
    complex_types: dict[str, ET.Element],
) -> ET.Element | None:
    local_name = _xsd_type_name(type_name)
    if local_name is None:
        return None
    return complex_types.get(local_name)


def _xsd_collect_components(
    schema_roots: list[ET.Element],
) -> tuple[dict[str, ET.Element], dict[str, ET.Element], dict[str, ET.Element]]:
    complex_types: dict[str, ET.Element] = {}
    simple_types: dict[str, ET.Element] = {}
    top_level_elements: dict[str, ET.Element] = {}
    for schema_root in schema_roots:
        complex_types.update(
            {
                element.attrib["name"]: element
                for element in schema_root.findall("xs:complexType", _XSD_NAMESPACES)
                if "name" in element.attrib
            }
        )
        simple_types.update(
            {
                element.attrib["name"]: element
                for element in schema_root.findall("xs:simpleType", _XSD_NAMESPACES)
                if "name" in element.attrib
            }
        )
        top_level_elements.update(
            {
                element.attrib["name"]: element
                for element in schema_root.findall("xs:element", _XSD_NAMESPACES)
                if "name" in element.attrib
            }
        )
    return complex_types, simple_types, top_level_elements


def _xsd_append_field(
    *,
    entity_name: str,
    field_path: str,
    required: bool,
    cardinality: str,
    data_type: str | None,
    enumerations: list[str],
    length: int | None,
    precision: int | None,
    scale: int | None,
    associations: list[str],
    annotations: dict[str, Any],
    source_evidence: str,
    warnings: list[str],
    fields: list[NormalizedSchemaField],
) -> None:
    fields.append(
        NormalizedSchemaField(
            entity_name=entity_name,
            field_path=field_path,
            parent_path=field_path.rsplit(".", 1)[0] if "." in field_path else None,
            data_type=data_type,
            required=required,
            cardinality=cardinality,
            length=length,
            precision=precision,
            scale=scale,
            is_key=field_path.split(".")[-1].lstrip("@").lower() == "id",
            enumerations=enumerations,
            associations=associations,
            annotations=annotations,
            description=None,
            source_evidence=source_evidence,
            warnings=warnings,
        )
    )


def _walk_xsd_container(
    *,
    entity_name: str,
    container: ET.Element,
    parent_path: str | None,
    source_pointer: str,
    fields: list[NormalizedSchemaField],
    warnings: list[str],
    complex_types: dict[str, ET.Element],
    simple_types: dict[str, ET.Element],
    active_types: set[str],
) -> None:
    for compositor_name in ("sequence", "all", "choice"):
        for compositor in container.findall(f"xs:{compositor_name}", _XSD_NAMESPACES):
            if compositor_name == "choice":
                warnings.append(
                    f"XSD choice compositor at {source_pointer} requires manual review."
                )
            for element in compositor.findall("xs:element", _XSD_NAMESPACES):
                _walk_xsd_element(
                    entity_name=entity_name,
                    element=element,
                    parent_path=parent_path,
                    source_pointer=source_pointer,
                    fields=fields,
                    warnings=warnings,
                    complex_types=complex_types,
                    simple_types=simple_types,
                    active_types=active_types,
                )

    for attribute in container.findall("xs:attribute", _XSD_NAMESPACES):
        attribute_name = attribute.attrib.get("name") or _xsd_type_name(attribute.attrib.get("ref"))
        if not attribute_name:
            warnings.append(f"Unnamed XSD attribute at {source_pointer} was skipped.")
            continue
        field_path = f"{parent_path}.@{attribute_name}" if parent_path else f"@{attribute_name}"
        inline_simple_type = attribute.find("xs:simpleType", _XSD_NAMESPACES)
        detail = _xsd_simple_type_details(
            inline_simple_type or _xsd_named_simple_type(attribute.attrib.get("type"), simple_types)
        )
        data_type = detail["data_type"] or _xsd_type_name(attribute.attrib.get("type"))
        required = attribute.attrib.get("use") == "required"
        _xsd_append_field(
            entity_name=entity_name,
            field_path=field_path,
            required=required,
            cardinality="1..1" if required else "0..1",
            data_type=data_type,
            enumerations=detail["enumerations"],
            length=detail["length"],
            precision=detail["precision"],
            scale=detail["scale"],
            associations=[],
            annotations={
                key: value
                for key, value in attribute.attrib.items()
                if key in {"type", "use", "ref"}
            },
            source_evidence=f"{source_pointer}/attribute/{attribute_name}",
            warnings=[],
            fields=fields,
        )


def _walk_xsd_element(
    *,
    entity_name: str,
    element: ET.Element,
    parent_path: str | None,
    source_pointer: str,
    fields: list[NormalizedSchemaField],
    warnings: list[str],
    complex_types: dict[str, ET.Element],
    simple_types: dict[str, ET.Element],
    active_types: set[str],
) -> None:
    element_name = element.attrib.get("name") or _xsd_type_name(element.attrib.get("ref"))
    if not element_name:
        warnings.append(f"Unnamed XSD element at {source_pointer} was skipped.")
        return

    field_path = element_name if parent_path is None else f"{parent_path}.{element_name}"
    required, cardinality = _xsd_cardinality(
        element.attrib.get("minOccurs"),
        element.attrib.get("maxOccurs"),
    )
    inline_simple_type = element.find("xs:simpleType", _XSD_NAMESPACES)
    inline_complex_type = element.find("xs:complexType", _XSD_NAMESPACES)
    named_simple_type = _xsd_named_simple_type(element.attrib.get("type"), simple_types)
    detail = _xsd_simple_type_details(inline_simple_type or named_simple_type)
    data_type = detail["data_type"] or _xsd_type_name(element.attrib.get("type"))
    named_complex_type = _xsd_named_complex_type(element.attrib.get("type"), complex_types)
    complex_type = inline_complex_type or named_complex_type
    associations = [element.attrib["type"]] if named_complex_type is not None else []

    _xsd_append_field(
        entity_name=entity_name,
        field_path=field_path,
        required=required,
        cardinality=cardinality,
        data_type=data_type or ("object" if complex_type is not None else None),
        enumerations=detail["enumerations"],
        length=detail["length"],
        precision=detail["precision"],
        scale=detail["scale"],
        associations=associations,
        annotations={
            key: value
            for key, value in element.attrib.items()
            if key in {"type", "ref", "minOccurs", "maxOccurs"}
        },
        source_evidence=f"{source_pointer}/element/{element_name}",
        warnings=[],
        fields=fields,
    )

    if complex_type is None:
        return

    type_marker = element.attrib.get("type") or f"inline:{source_pointer}/element/{element_name}"
    if type_marker in active_types:
        warnings.append(
            f"Recursive XSD type reference at {source_pointer}/element/{element_name} "
            "requires manual review."
        )
        return

    _walk_xsd_container(
        entity_name=entity_name,
        container=complex_type,
        parent_path=field_path,
        source_pointer=f"{source_pointer}/element/{element_name}",
        fields=fields,
        warnings=warnings,
        complex_types=complex_types,
        simple_types=simple_types,
        active_types=active_types | {type_marker},
    )


def _materialize_xsd_root_element(
    *,
    entity_name: str,
    element: ET.Element,
    source_pointer: str,
    fields: list[NormalizedSchemaField],
    warnings: list[str],
    complex_types: dict[str, ET.Element],
    simple_types: dict[str, ET.Element],
) -> None:
    inline_complex_type = element.find("xs:complexType", _XSD_NAMESPACES)
    named_complex_type = _xsd_named_complex_type(element.attrib.get("type"), complex_types)
    complex_type = inline_complex_type or named_complex_type
    if complex_type is not None:
        _walk_xsd_container(
            entity_name=entity_name,
            container=complex_type,
            parent_path=None,
            source_pointer=source_pointer,
            fields=fields,
            warnings=warnings,
            complex_types=complex_types,
            simple_types=simple_types,
            active_types={element.attrib.get("type", entity_name)},
        )
        return

    _walk_xsd_element(
        entity_name=entity_name,
        element=element,
        parent_path=None,
        source_pointer=source_pointer.rsplit("/", 1)[0],
        fields=fields,
        warnings=warnings,
        complex_types=complex_types,
        simple_types=simple_types,
        active_types=set(),
    )


def inspect_xsd(path: Path, document: ET.Element) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []

    complex_types, simple_types, top_level_elements = _xsd_collect_components([document])

    source_identity = document.attrib.get("targetNamespace") or path.stem
    version = document.attrib.get("version")
    namespace = document.attrib.get("targetNamespace")
    materialized_entities: set[str] = set()

    for entity_name, element in top_level_elements.items():
        if not entity_name:
            continue
        materialized_entities.add(entity_name)
        source_pointer = f"{source_identity}/element/{entity_name}"
        entities.append(
            NormalizedSchemaEntity(
                name=entity_name,
                kind="message_type",
                description=None,
                source_evidence=source_pointer,
            )
        )
        _materialize_xsd_root_element(
            entity_name=entity_name,
            element=element,
            source_pointer=source_pointer,
            fields=fields,
            warnings=warnings,
            complex_types=complex_types,
            simple_types=simple_types,
        )

    if not entities:
        for complex_type_name, complex_type in complex_types.items():
            if complex_type_name in materialized_entities:
                continue
            source_pointer = f"{source_identity}/complexType/{complex_type_name}"
            entities.append(
                NormalizedSchemaEntity(
                    name=complex_type_name,
                    kind="entity_type",
                    description=None,
                    source_evidence=source_pointer,
                )
            )
            _walk_xsd_container(
                entity_name=complex_type_name,
                container=complex_type,
                parent_path=None,
                source_pointer=source_pointer,
                fields=fields,
                warnings=warnings,
                complex_types=complex_types,
                simple_types=simple_types,
                active_types={complex_type_name},
            )

    for wildcard in document.findall(".//xs:any", _XSD_NAMESPACES):
        warnings.append(f"XSD wildcard at {_xml_local_name(wildcard.tag)} requires manual review.")

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="xsd",
        source_identity=source_identity,
        source_version=version,
        namespace=namespace,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        warnings=warnings,
    )


def inspect_wsdl(path: Path, document: ET.Element) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []
    operations: list[NormalizedSchemaOperation] = []

    source_identity = document.attrib.get("name") or path.stem
    namespace = document.attrib.get("targetNamespace")
    version = document.attrib.get("version")

    types_node = document.find("wsdl:types", _WSDL_NAMESPACES)
    schema_roots = (
        types_node.findall("xs:schema", _XSD_NAMESPACES) if types_node is not None else []
    )
    complex_types, simple_types, top_level_elements = _xsd_collect_components(schema_roots)

    for element_name, element in top_level_elements.items():
        source_pointer = f"{source_identity}/schema/{element_name}"
        entities.append(
            NormalizedSchemaEntity(
                name=element_name,
                kind="message_type",
                description=None,
                source_evidence=source_pointer,
            )
        )
        _materialize_xsd_root_element(
            entity_name=element_name,
            element=element,
            source_pointer=source_pointer,
            fields=fields,
            warnings=warnings,
            complex_types=complex_types,
            simple_types=simple_types,
        )

    wsdl_messages: set[str] = set()
    for message in document.findall("wsdl:message", _WSDL_NAMESPACES):
        message_name = message.attrib.get("name")
        if not message_name:
            warnings.append("Unnamed WSDL message was skipped.")
            continue
        wsdl_messages.add(message_name)
        source_pointer = f"{source_identity}/message/{message_name}"
        entities.append(
            NormalizedSchemaEntity(
                name=message_name,
                kind="message_type",
                description=None,
                source_evidence=source_pointer,
            )
        )
        for part in message.findall("wsdl:part", _WSDL_NAMESPACES):
            part_name = part.attrib.get("name") or "part"
            part_element = part.attrib.get("element")
            part_type = part.attrib.get("type")
            referenced_element_name = _xsd_type_name(part_element)
            referenced_type_name = _xsd_type_name(part_type)
            referenced_element = top_level_elements.get(referenced_element_name or "")
            referenced_complex_type = _xsd_named_complex_type(part_type, complex_types)
            part_simple_type = _xsd_named_simple_type(part_type, simple_types)
            part_detail = _xsd_simple_type_details(part_simple_type)
            part_associations = [value for value in (part_element, part_type) if value]
            part_warnings: list[str] = []
            if part_element and referenced_element is None:
                warning = (
                    f"WSDL message {message_name}.{part_name} references unresolved "
                    f"element {part_element}."
                )
                warnings.append(warning)
                part_warnings.append(warning)
            if (
                part_type
                and referenced_complex_type is None
                and part_simple_type is None
                and not part_type.startswith("xs:")
            ):
                warning = (
                    f"WSDL message {message_name}.{part_name} references unresolved "
                    f"type {part_type}."
                )
                warnings.append(warning)
                part_warnings.append(warning)
            _xsd_append_field(
                entity_name=message_name,
                field_path=part_name,
                required=True,
                cardinality="1..1",
                data_type=part_detail["data_type"]
                or referenced_type_name
                or ("object" if referenced_element is not None else None),
                enumerations=part_detail["enumerations"],
                length=part_detail["length"],
                precision=part_detail["precision"],
                scale=part_detail["scale"],
                associations=part_associations,
                annotations={
                    key: value for key, value in part.attrib.items() if key in {"element", "type"}
                },
                source_evidence=f"{source_pointer}/part/{part_name}",
                warnings=part_warnings,
                fields=fields,
            )

            if referenced_element is not None:
                inline_complex_type = referenced_element.find("xs:complexType", _XSD_NAMESPACES)
                named_complex_type = _xsd_named_complex_type(
                    referenced_element.attrib.get("type"),
                    complex_types,
                )
                complex_type = inline_complex_type or named_complex_type
                if complex_type is not None:
                    _walk_xsd_container(
                        entity_name=message_name,
                        container=complex_type,
                        parent_path=part_name,
                        source_pointer=f"{source_pointer}/part/{part_name}",
                        fields=fields,
                        warnings=warnings,
                        complex_types=complex_types,
                        simple_types=simple_types,
                        active_types={
                            referenced_element.attrib.get(
                                "type",
                                referenced_element_name or part_name,
                            )
                        },
                    )
            elif referenced_complex_type is not None:
                _walk_xsd_container(
                    entity_name=message_name,
                    container=referenced_complex_type,
                    parent_path=part_name,
                    source_pointer=f"{source_pointer}/part/{part_name}",
                    fields=fields,
                    warnings=warnings,
                    complex_types=complex_types,
                    simple_types=simple_types,
                    active_types={part_type or part_name},
                )

    binding_protocols: dict[str, str] = {}
    binding_actions: dict[tuple[str, str], str] = {}
    for binding in document.findall("wsdl:binding", _WSDL_NAMESPACES):
        port_type_name = _xsd_type_name(binding.attrib.get("type"))
        if not port_type_name:
            continue
        protocol = "wsdl"
        if binding.find("soap12:binding", _WSDL_NAMESPACES) is not None:
            protocol = "soap12"
        elif binding.find("soap:binding", _WSDL_NAMESPACES) is not None:
            protocol = "soap"
        binding_protocols[port_type_name] = protocol
        for operation in binding.findall("wsdl:operation", _WSDL_NAMESPACES):
            operation_name = operation.attrib.get("name")
            if not operation_name:
                continue
            soap12_operation = operation.find("soap12:operation", _WSDL_NAMESPACES)
            soap_operation = operation.find("soap:operation", _WSDL_NAMESPACES)
            binding_action = None
            if soap12_operation is not None:
                binding_protocols[port_type_name] = "soap12"
                binding_action = soap12_operation.attrib.get("soapAction")
            elif soap_operation is not None:
                binding_action = soap_operation.attrib.get("soapAction")
            if binding_action:
                binding_actions[(port_type_name, operation_name)] = binding_action

    for port_type in document.findall("wsdl:portType", _WSDL_NAMESPACES):
        port_type_name = port_type.attrib.get("name")
        if not port_type_name:
            continue
        protocol = binding_protocols.get(port_type_name, "wsdl")
        for operation in port_type.findall("wsdl:operation", _WSDL_NAMESPACES):
            operation_name = operation.attrib.get("name")
            if not operation_name:
                continue
            input_node = operation.find("wsdl:input", _WSDL_NAMESPACES)
            output_node = operation.find("wsdl:output", _WSDL_NAMESPACES)
            request_message = _xsd_type_name(
                input_node.attrib.get("message") if input_node is not None else None
            )
            response_message = _xsd_type_name(
                output_node.attrib.get("message") if output_node is not None else None
            )
            if request_message and request_message not in wsdl_messages:
                warnings.append(
                    f"WSDL operation {port_type_name}.{operation_name} references missing input "
                    f"message {request_message}."
                )
            if response_message and response_message not in wsdl_messages:
                warnings.append(
                    f"WSDL operation {port_type_name}.{operation_name} references missing output "
                    f"message {response_message}."
                )
            operations.append(
                NormalizedSchemaOperation(
                    operation_id=operation_name,
                    method="CALL",
                    path=binding_actions.get((port_type_name, operation_name), operation_name),
                    protocol=protocol,
                    request_body_schema=request_message,
                    response_schemas=(
                        [{"status_code": "output", "schema": response_message}]
                        if response_message
                        else []
                    ),
                    description=f"{port_type_name}.{operation_name}",
                    source_evidence=f"{source_identity}/portType/{port_type_name}/{operation_name}",
                )
            )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="wsdl",
        source_identity=source_identity,
        source_version=version,
        namespace=namespace,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        operations=operations,
        warnings=warnings,
    )


def inspect_edmx(path: Path, document: ET.Element) -> NormalizedSchemaDocument:
    checksum = _file_checksum(path)
    warnings: list[str] = []
    entities: list[NormalizedSchemaEntity] = []
    fields: list[NormalizedSchemaField] = []

    data_services = document.find("edmx:DataServices", _EDMX_NAMESPACES)
    if data_services is None:
        raise SchemaInspectionError("EDMX document has no DataServices section.")

    schemas = data_services.findall("edm:Schema", _EDMX_NAMESPACES)
    if not schemas:
        raise SchemaInspectionError("EDMX document has no Schema section.")

    namespace = schemas[0].attrib.get("Namespace")
    source_identity = namespace or path.stem
    version = document.attrib.get("Version")

    for schema in schemas:
        schema_namespace = schema.attrib.get("Namespace") or namespace or path.stem
        for entity_type in schema.findall("edm:EntityType", _EDMX_NAMESPACES):
            entity_name = entity_type.attrib.get("Name", "UnknownEntity")
            source_pointer = f"{schema_namespace}/EntityType/{entity_name}"
            entities.append(
                NormalizedSchemaEntity(
                    name=entity_name,
                    kind="entity_type",
                    description=None,
                    source_evidence=source_pointer,
                )
            )
            key_refs = {
                ref.attrib.get("Name")
                for ref in entity_type.findall("edm:Key/edm:PropertyRef", _EDMX_NAMESPACES)
            }
            for prop in entity_type.findall("edm:Property", _EDMX_NAMESPACES):
                field_name = prop.attrib.get("Name", "UnknownProperty")
                data_type = prop.attrib.get("Type")
                nullable = prop.attrib.get("Nullable")
                max_length = prop.attrib.get("MaxLength")
                precision = prop.attrib.get("Precision")
                scale = prop.attrib.get("Scale")
                fields.append(
                    NormalizedSchemaField(
                        entity_name=entity_name,
                        field_path=field_name,
                        parent_path=None,
                        data_type=data_type,
                        required=False if nullable == "true" else True,
                        cardinality=_cardinality(nullable != "true", data_type),
                        length=int(max_length) if max_length and max_length.isdigit() else None,
                        precision=int(precision) if precision and precision.isdigit() else None,
                        scale=int(scale) if scale and scale.isdigit() else None,
                        is_key=field_name in key_refs,
                        annotations=_normalize_annotations(prop.attrib),
                        description=None,
                        source_evidence=f"{source_pointer}/Property/{field_name}",
                    )
                )

            for nav in entity_type.findall("edm:NavigationProperty", _EDMX_NAMESPACES):
                field_name = nav.attrib.get("Name", "UnknownNavigation")
                target = nav.attrib.get("Type", "")
                fields.append(
                    NormalizedSchemaField(
                        entity_name=entity_name,
                        field_path=field_name,
                        parent_path=None,
                        data_type="navigation",
                        required=None,
                        cardinality=None,
                        associations=[target] if target else [],
                        annotations=_normalize_annotations(nav.attrib),
                        description=None,
                        source_evidence=f"{source_pointer}/NavigationProperty/{field_name}",
                        warnings=["Navigation property imported as association metadata only."],
                    )
                )
                warnings.append(
                    f"Navigation property {entity_name}.{field_name} requires manual review."
                )

    return NormalizedSchemaDocument(
        source_path=str(path),
        source_format="edmx",
        source_identity=source_identity,
        source_version=version,
        namespace=namespace,
        parser_version=_PARSER_VERSION,
        checksum=checksum,
        entities=entities,
        fields=fields,
        warnings=warnings,
    )


def inspect_schema_file(path: Path) -> NormalizedSchemaDocument:
    """Inspect a local schema contract or payload evidence file."""
    if not path.is_file():
        raise SchemaInspectionError(f"Schema file not found: {path}")
    if path.suffix.lower() not in {
        ".json",
        ".yaml",
        ".yml",
        ".xml",
        ".edmx",
        ".xsd",
        ".wsdl",
        ".csv",
        ".xlsx",
        ".html",
        ".htm",
        ".iflw",
        ".zip",
    }:
        raise SchemaInspectionError(
            "Schema file must be .json, .yaml, .yml, .xml, .edmx, .xsd, .wsdl, .csv, .xlsx, "
            ".html, .htm, .iflw, or .zip"
        )

    if path.suffix.lower() == ".csv":
        return inspect_field_catalog_csv(path)
    if path.suffix.lower() == ".xlsx":
        return inspect_xlsx_schema_evidence(path)
    if path.suffix.lower() in {".html", ".htm"}:
        return inspect_we60_html(path)
    if path.suffix.lower() == ".zip":
        return inspect_integration_suite_package(path)

    try:
        document = _load_document(path)
    except (json.JSONDecodeError, yaml.YAMLError, ET.ParseError) as exc:
        raise SchemaInspectionError(f"Schema file could not be parsed: {exc}") from exc

    kind = _infer_schema_kind(document, suffix=path.suffix.lower())
    if kind == "cds_metadata_export":
        return inspect_cds_metadata_json(path, document)
    if kind == "integration_flow_artifact":
        return inspect_integration_flow_artifact(path, document)
    if kind == "json_payload":
        return inspect_json_payload(path, document)
    if kind == "idoc_payload":
        return inspect_idoc_payload(path, document)
    if kind == "xml_payload":
        return inspect_xml_payload(path, document)
    if kind == "json_schema":
        return inspect_json_schema(path, document)
    if kind == "openapi":
        return inspect_openapi(path, document)
    if kind == "edmx":
        return inspect_edmx(path, document)
    if kind == "wsdl":
        return inspect_wsdl(path, document)
    if kind == "xsd":
        return inspect_xsd(path, document)
    raise SchemaInspectionError(f"Unsupported schema document type: {kind}")
