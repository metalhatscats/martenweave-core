from __future__ import annotations

import csv
import re
from pathlib import Path

from modelops_core.evidence.models import EvidenceFinding, EvidenceFindingKind

_ID_RE = re.compile(r"^[A-Z][A-Z0-9]*(-[A-Z0-9]+)*$")


def parse_markdown_note(path: Path) -> list[EvidenceFinding]:
    findings: list[EvidenceFinding] = []
    for lineno, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        stripped = line.strip().lstrip("-* ")
        if not stripped:
            continue
        obj_id = _extract_object_id(stripped)
        kind, field = _classify_text(stripped)
        findings.append(
            EvidenceFinding(
                kind=kind,
                object_id=obj_id,
                field=field,
                message=stripped,
                source_line=lineno,
            )
        )
    return findings


def parse_csv_report(path: Path) -> list[EvidenceFinding]:
    findings: list[EvidenceFinding] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for lineno, row in enumerate(reader, start=2):
            message = row.get("message") or row.get("issue") or row.get("description") or ""
            obj_id = row.get("object_id") or _extract_object_id(message)
            kind, field = _classify_text(message)
            findings.append(
                EvidenceFinding(
                    kind=kind,
                    object_id=obj_id,
                    field=field,
                    message=message,
                    severity=row.get("severity"),
                    source_line=lineno,
                )
            )
    return findings


def parse_xlsx_report(path: Path) -> list[EvidenceFinding]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX evidence import.") from exc

    findings: list[EvidenceFinding] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) for h in next(rows)]
        except StopIteration:
            continue
        for lineno, row in enumerate(rows, start=2):
            row_dict = {
                h: str(v) if v is not None else "" for h, v in zip(headers, row, strict=False)
            }
            message = (
                row_dict.get("message")
                or row_dict.get("issue")
                or row_dict.get("description")
                or ""
            )
            if not message:
                continue
            obj_id = row_dict.get("object_id") or _extract_object_id(message)
            kind, field = _classify_text(message)
            findings.append(
                EvidenceFinding(
                    kind=kind,
                    object_id=obj_id,
                    field=field,
                    message=message,
                    severity=row_dict.get("severity"),
                    source_line=lineno,
                )
            )
    wb.close()
    return findings


def _extract_object_id(text: str) -> str | None:
    for token in text.replace(",", " ").replace("'", " ").split():
        if _ID_RE.match(token):
            return token
    return None


_KEYWORDS: list[tuple[list[str], EvidenceFindingKind, str | None]] = [
    (["missing owner", "no owner", "owner missing"], EvidenceFindingKind.MISSING_OWNER, "owner"),
    (["missing mapping", "no mapping"], EvidenceFindingKind.MISSING_MAPPING, "mapping"),
    (["validation", "invalid", "broken"], EvidenceFindingKind.VALIDATION_ISSUE, None),
    (["decision", "agreed"], EvidenceFindingKind.DECISION_NOTE, None),
    (["rename", "renaming"], EvidenceFindingKind.RENAME_SUGGESTION, "name"),
    (["question", "unclear"], EvidenceFindingKind.FIELD_QUESTION, None),
]


def _classify_text(text: str) -> tuple[EvidenceFindingKind, str | None]:
    lower = text.lower()
    for keywords, kind, field in _KEYWORDS:
        if any(kw in lower for kw in keywords):
            return kind, field
    return EvidenceFindingKind.FIELD_QUESTION, None
