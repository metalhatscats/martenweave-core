"""Governed workbook suggestion review workbook export and feedback import."""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from modelops_core.pilot.workbook_suggestions import WorkbookSuggestionSet

REVIEW_SHEET_NAME = "Suggestions"
REVIEW_DECISIONS = ("accepted", "rejected", "unresolved")
_REVIEW_COLUMNS = [
    "suggestion_id",
    "suggestion_type",
    "suggestion_value",
    "confidence_label",
    "confidence_score",
    "status",
    "reviewer_decision",
    "reviewer_notes",
    "evidence_sheet",
    "evidence_table_id",
    "evidence_header_row",
    "evidence_column_name",
    "explanation",
    "input_fingerprint",
    "agent_identity",
    "model_identity",
    "evidence_json",
    "deterministic_context_json",
]
_EDITABLE_COLUMNS = {"reviewer_decision", "reviewer_notes"}


class WorkbookSuggestionReviewError(ValueError):
    """Raised when a suggestion review workbook cannot be safely imported."""


@dataclass
class WorkbookSuggestionFeedbackRecord:
    """Normalized reviewer feedback for a single workbook suggestion."""

    suggestion_id: str
    suggestion_type: str
    suggestion_value: str
    original_status: str
    reviewed_status: str
    confidence_label: str
    confidence_score: float
    explanation: str
    evidence: dict[str, Any]
    deterministic_context: dict[str, Any]
    input_fingerprint: str
    reviewer_notes: str = ""
    agent_identity: str | None = None
    model_identity: str | None = None


@dataclass
class WorkbookSuggestionFeedbackSet:
    """Machine-readable workbook suggestion review feedback."""

    generated_at: str
    source_workbook: str
    input_fingerprint: str
    generator: str = "human_review"
    feedback_records: list[WorkbookSuggestionFeedbackRecord] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def decision_counts(self) -> dict[str, int]:
        counts = {decision: 0 for decision in REVIEW_DECISIONS}
        for record in self.feedback_records:
            counts[record.reviewed_status] = counts.get(record.reviewed_status, 0) + 1
        return counts

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["decision_counts"] = self.decision_counts()
        return payload


def _review_row(suggestion: dict[str, Any]) -> list[Any]:
    evidence = suggestion.get("evidence", {})
    deterministic_context = suggestion.get("deterministic_context", {})
    return [
        suggestion["suggestion_id"],
        suggestion["suggestion_type"],
        suggestion["suggestion_value"],
        suggestion["confidence_label"],
        suggestion["confidence_score"],
        suggestion.get("status", "unresolved"),
        "",
        "",
        evidence.get("sheet", ""),
        evidence.get("table_id", ""),
        evidence.get("header_row", ""),
        evidence.get("column_name", ""),
        suggestion["explanation"],
        suggestion["input_fingerprint"],
        suggestion.get("agent_identity") or "",
        suggestion.get("model_identity") or "",
        json.dumps(evidence, sort_keys=True),
        json.dumps(deterministic_context, sort_keys=True),
    ]


def write_workbook_suggestion_review_xlsx(
    suggestion_set: WorkbookSuggestionSet,
    path: Path,
    *,
    generated_at: datetime | None = None,
) -> Path:
    """Write a protected XLSX workbook for reviewing workbook suggestions."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for workbook suggestion review export."
        ) from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    if generated_at is not None:
        workbook.properties.created = generated_at
        workbook.properties.modified = generated_at

    readme = workbook.active
    readme.title = "Read Me"
    readme.column_dimensions["A"].width = 90
    readme_rows = [
        "Workbook suggestion review",
        "",
        "This workbook contains deterministic workbook interpretation suggestions.",
        "It is a review artifact only and never changes canonical model files directly.",
        "",
        "Edit only these columns on the Suggestions sheet:",
        "  - reviewer_decision: accepted, rejected, or unresolved",
        "  - reviewer_notes: optional reviewer context",
        "",
        "Technical identity columns remain protected so suggestion IDs and evidence stay stable.",
        "Import reviewed decisions with:",
        "  martenweave import-workbook-suggestion-review --from <workbook.xlsx> --out <dir>",
        "",
        f"Input fingerprint: {suggestion_set.input_fingerprint}",
        f"Suggestion count: {len(suggestion_set.suggestions)}",
    ]
    for row_index, value in enumerate(readme_rows, start=1):
        cell = readme.cell(row=row_index, column=1, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row_index == 1:
            cell.font = Font(bold=True, size=14)

    sheet = workbook.create_sheet(REVIEW_SHEET_NAME)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    readonly_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    alt_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_index, column_name in enumerate(_REVIEW_COLUMNS, start=1):
        header = sheet.cell(row=1, column=col_index, value=column_name)
        header.font = Font(bold=True, color="FFFFFF")
        header.fill = header_fill
        header.alignment = wrap

    for row_index, suggestion in enumerate(suggestion_set.to_dict()["suggestions"], start=2):
        for col_index, value in enumerate(_review_row(suggestion), start=1):
            column_name = _REVIEW_COLUMNS[col_index - 1]
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = wrap
            cell.border = thin_border
            if row_index % 2 == 0:
                cell.fill = alt_fill
            if column_name in _EDITABLE_COLUMNS:
                cell.fill = review_fill
                cell.protection = Protection(locked=False)
            else:
                cell.fill = readonly_fill
                cell.font = Font(italic=True, color="666666")

    decision_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(REVIEW_DECISIONS)}"',
        allow_blank=True,
    )
    decision_validation.error = "Use accepted, rejected, or unresolved."
    decision_validation.errorTitle = "Invalid reviewer decision"
    sheet.add_data_validation(decision_validation)
    decision_column_index = _REVIEW_COLUMNS.index("reviewer_decision") + 1
    decision_column_letter = sheet.cell(row=1, column=decision_column_index).column_letter
    decision_validation.add(
        f"{decision_column_letter}2:{decision_column_letter}{len(suggestion_set.suggestions) + 1}"
    )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:R{len(suggestion_set.suggestions) + 1}"
    sheet.protection.sheet = True
    sheet.protection.sort = False
    sheet.protection.autoFilter = True

    for col_index, column_name in enumerate(_REVIEW_COLUMNS, start=1):
        letter = sheet.cell(row=1, column=col_index).column_letter
        if column_name in {"explanation", "reviewer_notes"}:
            width = 42
        elif column_name.endswith("_json"):
            width = 30
        elif column_name in {"suggestion_id", "input_fingerprint"}:
            width = 24
        else:
            width = 18
        sheet.column_dimensions[letter].width = width

    workbook.save(path)
    workbook.close()
    return path


def import_workbook_suggestion_review_xlsx(
    path: Path,
) -> WorkbookSuggestionFeedbackSet:
    """Import a reviewed workbook-suggestion XLSX workbook and normalize feedback."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for workbook suggestion review import."
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if REVIEW_SHEET_NAME not in workbook.sheetnames:
            raise WorkbookSuggestionReviewError(
                f"Workbook must contain a '{REVIEW_SHEET_NAME}' sheet."
            )
        sheet = workbook[REVIEW_SHEET_NAME]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows)]
        header_index = {header: index for index, header in enumerate(headers)}
        missing_headers = {
            "suggestion_id",
            "status",
            "reviewer_decision",
            "reviewer_notes",
            "evidence_json",
            "deterministic_context_json",
            "input_fingerprint",
        } - set(header_index)
        if missing_headers:
            missing = ", ".join(sorted(missing_headers))
            raise WorkbookSuggestionReviewError(
                f"Suggestion review workbook is missing required column(s): {missing}."
            )

        feedback_records: list[WorkbookSuggestionFeedbackRecord] = []
        warnings: list[str] = []
        seen_ids: set[str] = set()
        input_fingerprint = ""

        for row_number, row in enumerate(rows, start=2):
            row_values = ["" if value is None else str(value) for value in row]
            if not any(value.strip() for value in row_values):
                continue

            suggestion_id = row_values[header_index["suggestion_id"]].strip()
            if not suggestion_id:
                raise WorkbookSuggestionReviewError(
                    f"Every non-empty review row must include a stable suggestion_id; "
                    f"missing value at row {row_number}."
                )
            if suggestion_id in seen_ids:
                warnings.append(
                    f"Duplicate suggestion_id ignored at row {row_number}: {suggestion_id}"
                )
                continue
            seen_ids.add(suggestion_id)

            original_status = row_values[header_index["status"]].strip() or "unresolved"
            reviewer_decision = row_values[header_index["reviewer_decision"]].strip().lower()
            reviewer_notes = row_values[header_index["reviewer_notes"]].strip()
            reviewed_status = reviewer_decision or original_status
            if reviewed_status not in REVIEW_DECISIONS:
                raise WorkbookSuggestionReviewError(
                    f"Row {row_number} has invalid reviewer_decision/status '{reviewed_status}'. "
                    f"Expected one of: {', '.join(REVIEW_DECISIONS)}."
                )

            row_fingerprint = row_values[header_index["input_fingerprint"]].strip()
            if not input_fingerprint:
                input_fingerprint = row_fingerprint
            elif row_fingerprint and row_fingerprint != input_fingerprint:
                warnings.append(
                    f"Row {row_number} uses a different input_fingerprint; kept for audit."
                )

            feedback_records.append(
                WorkbookSuggestionFeedbackRecord(
                    suggestion_id=suggestion_id,
                    suggestion_type=row_values[header_index["suggestion_type"]].strip(),
                    suggestion_value=row_values[header_index["suggestion_value"]].strip(),
                    original_status=original_status,
                    reviewed_status=reviewed_status,
                    confidence_label=row_values[header_index["confidence_label"]].strip(),
                    confidence_score=float(
                        row_values[header_index["confidence_score"]].strip() or "0"
                    ),
                    explanation=row_values[header_index["explanation"]].strip(),
                    evidence=json.loads(row_values[header_index["evidence_json"]].strip() or "{}"),
                    deterministic_context=json.loads(
                        row_values[header_index["deterministic_context_json"]].strip() or "{}"
                    ),
                    input_fingerprint=row_fingerprint,
                    reviewer_notes=reviewer_notes,
                    agent_identity=row_values[header_index["agent_identity"]].strip() or None,
                    model_identity=row_values[header_index["model_identity"]].strip() or None,
                )
            )
    finally:
        workbook.close()

    return WorkbookSuggestionFeedbackSet(
        generated_at=datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        source_workbook=str(path),
        input_fingerprint=input_fingerprint,
        feedback_records=feedback_records,
        warnings=warnings,
    )


def write_workbook_suggestion_feedback_json(
    feedback_set: WorkbookSuggestionFeedbackSet,
    path: Path,
) -> Path:
    """Write normalized suggestion review feedback as JSON."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(feedback_set.to_dict(), indent=2) + "\n", encoding="utf-8")
    return path


def write_workbook_suggestion_feedback_markdown(
    feedback_set: WorkbookSuggestionFeedbackSet,
    path: Path,
) -> Path:
    """Write normalized suggestion review feedback as Markdown."""
    lines = [
        "# Workbook suggestion review feedback",
        "",
        f"- Source workbook: `{Path(feedback_set.source_workbook).name}`",
        f"- Input fingerprint: `{feedback_set.input_fingerprint}`",
        f"- Generated: {feedback_set.generated_at}",
        "",
        "## Decision counts",
        "",
    ]
    for decision, count in feedback_set.decision_counts().items():
        lines.append(f"- `{decision}`: {count}")
    if feedback_set.warnings:
        lines.extend(["", "## Warnings", ""])
        for warning in feedback_set.warnings:
            lines.append(f"- {warning}")
    lines.extend(["", "## Reviewed suggestions", ""])
    if not feedback_set.feedback_records:
        lines.append("No reviewed suggestions were imported.")
    for record in feedback_set.feedback_records:
        lines.append(f"### {record.suggestion_id}")
        lines.append("")
        lines.append(f"- Type: `{record.suggestion_type}`")
        lines.append(f"- Value: `{record.suggestion_value}`")
        lines.append(f"- Original status: `{record.original_status}`")
        lines.append(f"- Reviewed status: `{record.reviewed_status}`")
        lines.append(
            f"- Confidence: `{record.confidence_label}` ({record.confidence_score:.2f})"
        )
        lines.append(f"- Explanation: {record.explanation}")
        if record.reviewer_notes:
            lines.append(f"- Reviewer notes: {record.reviewer_notes}")
        lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path
