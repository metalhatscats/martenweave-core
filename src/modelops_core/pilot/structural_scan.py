"""Deterministic structural scan for irregular workbook evidence.

The scanner stays metadata-first: it detects likely table sections, probable
header rows, workbook structure warnings, and column-role candidates without
promoting workbook values to canonical truth.
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

SCANNER_VERSION = "1.0"
MAX_COLUMNS = 50
MAX_TITLE_ROWS = 2
MAX_SCAN_ROWS_PER_SHEET = 50_000

_ROLE_PATTERNS: dict[str, tuple[str, ...]] = {
    "source_field": (
        "source_field",
        "source field",
        "legacy field",
        "old field",
        "source",
        "поле источника",
        "исходное поле",
    ),
    "target_table": ("target_table", "target table", "sap table", "таблица цели"),
    "target_field": ("target_field", "target field", "new field", "target", "поле цели"),
    "owner": ("owner", "steward", "responsible", "владелец"),
    "status": ("status", "state", "disposition", "статус"),
    "rule": ("rule", "transform", "transformation", "mapping type"),
    "comment": ("comment", "note", "remarks", "reviewer comment"),
    "decision": ("decision", "topic", "approval"),
    "source_system": ("source system", "source_system", "legacy system", "система источника"),
    "target_system": ("target system", "target_system", "sap system"),
    "description": ("description", "label", "business name"),
    "data_type": ("data type", "datatype", "type", "format", "тип данных"),
}

_MAPPING_KEYWORDS = {
    "source_field",
    "target_field",
    "target_table",
    "source_system",
    "target_system",
    "rule",
}


@dataclass
class DetectedColumn:
    name: str
    normalized_name: str
    role: str
    confidence: str


@dataclass
class DetectedTable:
    table_id: str
    title_rows: list[int]
    header_row: int
    start_row: int
    end_row: int
    row_count: int
    column_count: int
    confidence: str
    fingerprint: str
    repeated_header: bool = False
    warnings: list[str] = field(default_factory=list)
    detected_columns: list[DetectedColumn] = field(default_factory=list)


@dataclass
class SheetScanResult:
    name: str
    hidden: bool
    included: bool
    purpose: str
    purpose_confidence: str
    row_count: int
    scanned_row_count: int
    probable_header_rows: list[int]
    exclusions: list[str]
    warnings: list[str]
    fingerprint: str
    tables: list[DetectedTable] = field(default_factory=list)


@dataclass
class WorkbookStructuralManifest:
    file_path: str
    file_hash: str
    scanner_version: str
    workbook_warnings: list[str]
    sheets: list[SheetScanResult]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8192), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_label(value: Any) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).strip().split())
    return text.lower()


def _string_cells(values: tuple[Any, ...]) -> list[str]:
    return [str(value).strip() if value is not None else "" for value in values[:MAX_COLUMNS]]


def _non_empty_cells(values: list[str]) -> list[str]:
    return [value for value in values if value]


def _header_score(values: list[str]) -> tuple[float, list[DetectedColumn]]:
    non_empty = _non_empty_cells(values)
    if len(non_empty) < 2:
        return 0.0, []

    unique_ratio = len({value.lower() for value in non_empty}) / len(non_empty)
    keyword_hits = 0
    detected_columns: list[DetectedColumn] = []
    for value in non_empty:
        normalized = _normalize_label(value)
        role = "unknown"
        confidence = "low"
        for candidate_role, patterns in _ROLE_PATTERNS.items():
            if any(pattern in normalized for pattern in patterns):
                role = candidate_role
                confidence = "high" if candidate_role in _MAPPING_KEYWORDS else "medium"
                keyword_hits += 1
                break
        detected_columns.append(
            DetectedColumn(
                name=value,
                normalized_name=normalized.replace(" ", "_"),
                role=role,
                confidence=confidence,
            )
        )

    string_ratio = sum(1 for value in non_empty if not value.isdigit()) / len(non_empty)
    keyword_ratio = min(keyword_hits / len(non_empty), 1.0)
    score = min((unique_ratio * 0.35) + (string_ratio * 0.25) + (keyword_ratio * 0.4), 1.0)
    return score, detected_columns


def _confidence_from_score(score: float) -> str:
    if score >= 0.8:
        return "high"
    if score >= 0.55:
        return "medium"
    if score >= 0.3:
        return "low"
    return "unknown"


def _is_blank_row(values: list[str]) -> bool:
    return not any(values)


def _looks_like_title_row(values: list[str]) -> bool:
    non_empty = _non_empty_cells(values)
    return bool(non_empty) and len(non_empty) <= 2 and len(" ".join(non_empty)) <= 80


def _same_header(values: list[str], detected_columns: list[DetectedColumn]) -> bool:
    expected = [_normalize_label(column.name) for column in detected_columns]
    candidate = [_normalize_label(value) for value in values if value]
    expected = [value for value in expected if value]
    return bool(expected) and candidate == expected


def _sheet_purpose(sheet_name: str, tables: list[DetectedTable]) -> tuple[str, str]:
    name = _normalize_label(sheet_name)
    if "decision" in name:
        return "decision_register", "high"
    if "validation" in name or "error" in name:
        return "validation_results", "high"
    if "mapping" in name:
        return "mapping", "high"
    if tables:
        roles = {column.role for table in tables for column in table.detected_columns}
        if roles & _MAPPING_KEYWORDS:
            return "mapping", "medium"
        if {"status", "comment", "decision"} & roles:
            return "review_register", "medium"
    return "unknown", "low"


def _table_fingerprint(
    sheet_name: str,
    header_row: int,
    start_row: int,
    end_row: int,
    columns: list[DetectedColumn],
) -> str:
    payload = {
        "sheet": sheet_name,
        "header_row": header_row,
        "start_row": start_row,
        "end_row": end_row,
        "columns": [column.normalized_name for column in columns],
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()


def _scan_sheet(sheet: Any) -> SheetScanResult:
    probable_header_rows: list[int] = []
    exclusions: list[str] = []
    warnings: list[str] = []
    tables: list[DetectedTable] = []
    buffered_non_empty_rows: list[tuple[int, list[str]]] = []

    current_table: dict[str, Any] | None = None
    last_non_empty_row = 0
    scanned_row_count = 0
    repeated_header_start = False

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        scanned_row_count = row_index
        values = _string_cells(row)
        if _is_blank_row(values):
            buffered_non_empty_rows.clear()
            if current_table is not None and current_table["data_row_seen"]:
                current_table["end_row"] = last_non_empty_row
                tables.append(_finalize_table(sheet.title, current_table))
                current_table = None
            continue

        last_non_empty_row = row_index
        score, detected_columns = _header_score(values)
        is_header_candidate = score >= 0.55

        if current_table is not None and _same_header(values, current_table["detected_columns"]):
            current_table["end_row"] = row_index - 1
            tables.append(_finalize_table(sheet.title, current_table))
            current_table = None
            repeated_header_start = True

        if current_table is None and is_header_candidate:
            probable_header_rows.append(row_index)
            title_rows = [
                index
                for index, buffered_values in buffered_non_empty_rows[-MAX_TITLE_ROWS:]
                if _looks_like_title_row(buffered_values)
            ]
            current_table = {
                "title_rows": title_rows,
                "header_row": row_index,
                "start_row": min(title_rows) if title_rows else row_index,
                "end_row": row_index,
                "detected_columns": detected_columns,
                "score": score,
                "data_row_seen": False,
                "repeated_header": repeated_header_start,
            }
            buffered_non_empty_rows.clear()
            repeated_header_start = False
            continue

        if current_table is not None:
            current_table["end_row"] = row_index
            current_table["data_row_seen"] = True
        else:
            buffered_non_empty_rows.append((row_index, values))
            buffered_non_empty_rows = buffered_non_empty_rows[-MAX_TITLE_ROWS:]

        if row_index >= MAX_SCAN_ROWS_PER_SHEET:
            warnings.append(
                f"Scan capped at {MAX_SCAN_ROWS_PER_SHEET} rows for bounded memory review."
            )
            break

    if current_table is not None:
        tables.append(_finalize_table(sheet.title, current_table))

    if len(tables) > 1:
        warnings.append(f"{len(tables)} table-like section(s) detected.")
    if not tables:
        exclusions.append("No reliable table header detected.")

    purpose, purpose_confidence = _sheet_purpose(sheet.title, tables)
    if sheet.sheet_state != "visible":
        exclusions.append("Hidden worksheet excluded from initial interpretation.")

    sheet_payload = {
        "sheet": sheet.title,
        "purpose": purpose,
        "headers": probable_header_rows,
        "tables": [table.fingerprint for table in tables],
    }
    fingerprint = hashlib.sha256(
        json.dumps(sheet_payload, sort_keys=True).encode("utf-8")
    ).hexdigest()

    return SheetScanResult(
        name=sheet.title,
        hidden=sheet.sheet_state != "visible",
        included=sheet.sheet_state == "visible",
        purpose=purpose,
        purpose_confidence=purpose_confidence,
        row_count=sheet.max_row or scanned_row_count,
        scanned_row_count=min(scanned_row_count, MAX_SCAN_ROWS_PER_SHEET),
        probable_header_rows=probable_header_rows,
        exclusions=exclusions,
        warnings=warnings,
        fingerprint=fingerprint,
        tables=tables,
    )


def _finalize_table(
    sheet_name: str,
    current_table: dict[str, Any],
) -> DetectedTable:
    header_row = current_table["header_row"]
    start_row = current_table["start_row"]
    end_row = max(current_table["end_row"], header_row)
    row_count = max(end_row - header_row, 0)
    score = current_table["score"]
    detected_columns = current_table["detected_columns"]
    repeated_header = current_table.get("repeated_header", False)
    warnings: list[str] = []
    if repeated_header:
        warnings.append("Repeated header row detected; new table section split here.")
    if score < 0.8:
        warnings.append("Table header confidence below high.")
    fingerprint = _table_fingerprint(sheet_name, header_row, start_row, end_row, detected_columns)
    return DetectedTable(
        table_id=f"{sheet_name}:{header_row}",
        title_rows=current_table["title_rows"],
        header_row=header_row,
        start_row=start_row,
        end_row=end_row,
        row_count=row_count,
        column_count=len(detected_columns),
        confidence=_confidence_from_score(score),
        fingerprint=fingerprint,
        repeated_header=repeated_header,
        warnings=warnings,
        detected_columns=detected_columns,
    )


def scan_workbook_structure(path: Path) -> WorkbookStructuralManifest:
    """Generate a deterministic structural manifest for a workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for workbook structural scanning.") from exc

    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=True)
    try:
        sheets = [_scan_sheet(workbook[sheet_name]) for sheet_name in workbook.sheetnames]
    finally:
        workbook.close()

    workbook_warnings: list[str] = []
    hidden_sheet_count = sum(1 for sheet in sheets if sheet.hidden)
    if hidden_sheet_count:
        workbook_warnings.append(f"{hidden_sheet_count} hidden sheet(s) present.")
    if any(sheet.tables for sheet in sheets):
        table_count = sum(len(sheet.tables) for sheet in sheets)
        workbook_warnings.append(f"{table_count} table-like section(s) detected across workbook.")

    return WorkbookStructuralManifest(
        file_path=str(path),
        file_hash=_file_hash(path),
        scanner_version=SCANNER_VERSION,
        workbook_warnings=workbook_warnings,
        sheets=sheets,
    )
