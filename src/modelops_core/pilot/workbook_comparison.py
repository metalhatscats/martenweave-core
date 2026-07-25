"""Semantic, evidence-only comparison for two SAP mapping workbooks."""

from __future__ import annotations

import hashlib
import html
import json
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from modelops_core.pilot.structural_scan import scan_workbook_structure
from modelops_core.repository import parse_file, scan_repository

_ALIASES = {
    "source_field": ("source field", "source_field", "legacy field", "поле источника"),
    "source_table": ("source table", "source_table", "legacy table", "таблица источника"),
    "source_system": ("source system", "source_system", "legacy system", "система источника"),
    "target_field": ("target field", "target_field", "new field", "поле цели"),
    "target_table": ("target table", "target_table", "sap table", "таблица цели"),
    "condition": ("condition", "when", "if", "условие"),
    "transformation": ("transformation", "transform", "rule", "mapping type"),
    "owner": ("owner", "steward", "responsible", "владелец"),
    "status": ("status", "state", "disposition", "статус"),
    "validation_rule": ("validation rule", "validation_rule", "правило валидации"),
    "decision_id": ("decision id", "decision_id", "идентификатор решения"),
    "topic": ("topic", "decision topic", "тема", "решение"),
    "decision": ("decision", "resolution", "решение"),
}
_COMPARABLE_FIELDS = (
    "source_field",
    "source_table",
    "source_system",
    "target_field",
    "target_table",
    "condition",
    "transformation",
    "owner",
    "status",
    "validation_rule",
)


def _norm(value: object) -> str:
    return re.sub(r"[^\w]+", "", unicodedata.normalize("NFKC", str(value)).casefold())


def _roles(headers: list[str]) -> dict[str, str]:
    values = {_norm(header): header for header in headers if header}
    return {
        role: values[_norm(alias)]
        for role, aliases in _ALIASES.items()
        for alias in aliases
        if _norm(alias) in values
    }


def _value(row: dict[str, str], roles: dict[str, str], role: str) -> str:
    return row.get(roles.get(role, ""), "")


def _read_records(path: Path) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    from openpyxl import load_workbook

    mappings: dict[str, dict[str, str]] = {}
    decisions: dict[str, dict[str, str]] = {}
    manifest = scan_workbook_structure(path)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        for sheet in manifest.sheets:
            if not sheet.included:
                continue
            for table in sheet.tables:
                worksheet = workbook[sheet.name]
                header = next(
                    worksheet.iter_rows(
                        min_row=table.header_row, max_row=table.header_row, values_only=True
                    ),
                    (),
                )
                headers = [str(value).strip() if value is not None else "" for value in header]
                roles = _roles(headers)
                if not roles:
                    continue
                for row_number, values in enumerate(
                    worksheet.iter_rows(
                        min_row=table.header_row + 1, max_row=table.end_row, values_only=True
                    ),
                    start=table.header_row + 1,
                ):
                    row = {
                        header: "" if value is None else str(value).strip()
                        for header, value in zip(headers, values, strict=False)
                        if header
                    }
                    if not any(row.values()):
                        continue
                    location = {"sheet": sheet.name, "row": row_number}
                    if _value(row, roles, "source_field") and (
                        _value(row, roles, "target_field") or _value(row, roles, "target_table")
                    ):
                        record = {field: _value(row, roles, field) for field in _COMPARABLE_FIELDS}
                        record["location"] = location  # type: ignore[assignment]
                        key = "|".join(
                            _norm(record[field])
                            for field in ("source_system", "source_table", "source_field")
                        )
                        mappings.setdefault(key, record)
                    if _value(row, roles, "decision_id") or _value(row, roles, "topic"):
                        record = {
                            "decision_id": _value(row, roles, "decision_id"),
                            "topic": _value(row, roles, "topic"),
                            "decision": _value(row, roles, "decision"),
                            "owner": _value(row, roles, "owner"),
                            "status": _value(row, roles, "status"),
                            "location": location,
                        }
                        key = _norm(record["decision_id"] or record["topic"])
                        decisions.setdefault(key, record)
    finally:
        workbook.close()
    return mappings, decisions


def _hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _model_endpoints(repo_root: Path | None) -> dict[tuple[str, str], list[str]]:
    if repo_root is None:
        return {}
    result: dict[tuple[str, str], list[str]] = {}
    for path in scan_repository(repo_root / "model"):
        parsed = parse_file(path)
        data = parsed.frontmatter
        if data.get("type") != "FieldEndpoint":
            continue
        table = str(data.get("sap_table") or data.get("table") or "").upper()
        field = str(data.get("sap_field") or data.get("field") or "").upper()
        if table and field:
            result.setdefault((table, field), []).append(str(data["id"]))
    return result


@dataclass
class WorkbookComparison:
    base_sha256: str
    head_sha256: str
    mapping_changes: list[dict[str, Any]] = field(default_factory=list)
    decision_changes: list[dict[str, Any]] = field(default_factory=list)
    model_impact: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        counts: dict[str, int] = {}
        for item in [*self.mapping_changes, *self.decision_changes]:
            counts[item["change_type"]] = counts.get(item["change_type"], 0) + 1
        return {**asdict(self), "counts": dict(sorted(counts.items()))}


def compare_mapping_workbooks(
    base_path: Path, head_path: Path, repo_root: Path | None = None
) -> WorkbookComparison:
    """Compare mapping semantics and model impact without changing canonical files."""
    base_mappings, base_decisions = _read_records(base_path)
    head_mappings, head_decisions = _read_records(head_path)
    endpoints = _model_endpoints(repo_root)
    mapping_changes: list[dict[str, Any]] = []
    impact: list[dict[str, Any]] = []
    for key in sorted(set(base_mappings) | set(head_mappings)):
        previous, current = base_mappings.get(key), head_mappings.get(key)
        if previous is None:
            change_type, changed = (
                "added",
                {field: {"before": "", "after": current[field]} for field in _COMPARABLE_FIELDS},
            )
        elif current is None:
            change_type, changed = (
                "removed",
                {field: {"before": previous[field], "after": ""} for field in _COMPARABLE_FIELDS},
            )
        else:
            changed = {
                field: {"before": previous[field], "after": current[field]}
                for field in _COMPARABLE_FIELDS
                if previous[field] != current[field]
            }
            if not changed:
                continue
            change_type = "changed"
        item = {
            "change_type": change_type,
            "mapping_key": key,
            "changes": changed,
            "before": previous,
            "after": current,
        }
        mapping_changes.append(item)
        for record in (previous, current):
            if not record:
                continue
            endpoint_key = (record["target_table"].upper(), record["target_field"].upper())
            for object_id in endpoints.get(endpoint_key, []):
                impact.append(
                    {"mapping_key": key, "object_id": object_id, "reason": "target_field_endpoint"}
                )

    decision_changes: list[dict[str, Any]] = []
    for key in sorted(set(base_decisions) | set(head_decisions)):
        previous, current = base_decisions.get(key), head_decisions.get(key)
        if previous is None:
            change_type, changed = (
                "decision_added",
                {
                    name: {"before": "", "after": current[name]}
                    for name in ("decision", "owner", "status")
                },
            )
        elif current is None:
            change_type, changed = (
                "decision_removed",
                {
                    name: {"before": previous[name], "after": ""}
                    for name in ("decision", "owner", "status")
                },
            )
        else:
            changed = {
                name: {"before": previous[name], "after": current[name]}
                for name in ("decision", "owner", "status")
                if previous[name] != current[name]
            }
            if not changed:
                continue
            change_type = "decision_changed"
        decision_changes.append(
            {
                "change_type": change_type,
                "decision_key": key,
                "changes": changed,
                "before": previous,
                "after": current,
            }
        )
    return WorkbookComparison(
        _hash(base_path),
        _hash(head_path),
        mapping_changes,
        decision_changes,
        sorted(impact, key=lambda item: (item["mapping_key"], item["object_id"])),
    )


def write_workbook_comparison(report: WorkbookComparison, out_dir: Path) -> dict[str, Path]:
    """Write JSON, HTML, and an XLSX review register for a comparison."""
    from openpyxl import Workbook

    out_dir.mkdir(parents=True, exist_ok=True)
    payload = report.to_dict()
    json_path = out_dir / "mapping-workbook-comparison.json"
    html_path = out_dir / "mapping-workbook-comparison.html"
    xlsx_path = out_dir / "mapping-workbook-comparison.xlsx"
    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    rows = ["<tr><th>Type</th><th>Key</th><th>Changes</th></tr>"]
    for item in [*report.mapping_changes, *report.decision_changes]:
        key = item.get("mapping_key", item.get("decision_key", ""))
        changes = html.escape(json.dumps(item["changes"], sort_keys=True))
        rows.append(
            f"<tr><td>{html.escape(item['change_type'])}</td><td>{html.escape(key)}</td>"
            f"<td>{changes}</td></tr>"
        )
    document = (
        "<!doctype html><title>Mapping workbook comparison</title>"
        "<h1>Mapping workbook comparison</h1>"
        "<p>Evidence only: review before any PatchProposal.</p><table>" + "".join(rows) + "</table>"
    )
    html_path.write_text(document, encoding="utf-8")
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Changes"
    sheet.append(["Change type", "Key", "Changed fields", "Before location", "After location"])
    for item in [*report.mapping_changes, *report.decision_changes]:
        before = item.get("before") or {}
        after = item.get("after") or {}
        sheet.append(
            [
                item["change_type"],
                item.get("mapping_key", item.get("decision_key", "")),
                ", ".join(sorted(item["changes"])),
                json.dumps(before.get("location", {})),
                json.dumps(after.get("location", {})),
            ]
        )
    impact_sheet = workbook.create_sheet("Model impact")
    impact_sheet.append(["Mapping key", "Canonical object", "Reason"])
    for item in report.model_impact:
        impact_sheet.append([item["mapping_key"], item["object_id"], item["reason"]])
    workbook.save(xlsx_path)
    workbook.close()
    return {"json": json_path, "html": html_path, "xlsx": xlsx_path}
