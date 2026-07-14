"""Bootstrap a Martenweave repository from a SAP mapping workbook.

Produces a valid repository skeleton and a reviewable PatchProposal with
inferred model objects. Canonical files are never mutated until the proposal
is explicitly reviewed and applied.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import yaml

from modelops_core import __version__
from modelops_core.config import RepoConfig
from modelops_core.domain_packs.sap import _SAP_CONTEXT_RULES
from modelops_core.patching.patch_model import PatchOperation
from modelops_core.patching.patch_proposal_service import (
    build_patch_proposal,
    write_patch_proposal,
)

_REQUIRED_COLUMNS = {"source_field", "target_table", "target_field"}
_MAPPING_SHEET_SUFFIX = "_Mappings"


@dataclass
class BootstrapResult:
    """Result of a bootstrap assessment."""

    repo_root: Path
    repo_name: str
    proposal_path: Path
    report_md_path: Path
    report_json_path: Path
    inferred_objects_count: int
    warnings: list[str] = field(default_factory=list)
    assumptions: list[str] = field(default_factory=list)
    next_commands: list[str] = field(default_factory=list)


def _clean_id(text: str) -> str:
    """Normalize a free-text label into a canonical object ID fragment."""
    text = str(text).upper().strip()
    text = re.sub(r"[^A-Z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text


def _object_id(prefix: str, *parts: str) -> str:
    """Build a deterministic canonical object ID from prefix and parts."""
    cleaned = [_clean_id(p) for p in parts if p]
    if not cleaned:
        raise ValueError("Cannot build object ID from empty parts")
    return f"{prefix}-{'-'.join(cleaned)}"


def _read_workbook(path: Path) -> dict[str, list[dict[str, str]]]:
    """Read all sheets from an XLSX workbook into row dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for bootstrap-assessment") from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, list[dict[str, str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets[sheet_name] = []
            continue
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data: list[dict[str, str]] = []
        for row in rows[1:]:
            data.append(
                {
                    h: (str(v).strip() if v is not None else "")
                    for h, v in zip(headers, row, strict=False)
                }
            )
        sheets[sheet_name] = data
    wb.close()
    return sheets


def _find_column(row: dict[str, str], *candidates: str) -> str | None:
    """Return the first matching column name from candidates."""
    for key in row:
        if key.lower() in candidates:
            return key
    return None


def _sheet_domain_id(sheet_name: str) -> str:
    """Derive a domain ID from a mapping sheet name."""
    base = sheet_name
    if base.endswith(_MAPPING_SHEET_SUFFIX):
        base = base[: -len(_MAPPING_SHEET_SUFFIX)]
    return _object_id("DOMAIN", base)


def _sheet_domain_name(sheet_name: str) -> str:
    base = sheet_name
    if base.endswith(_MAPPING_SHEET_SUFFIX):
        base = base[: -len(_MAPPING_SHEET_SUFFIX)]
    return base.replace("_", " ").strip() or sheet_name


def _infer_context_category(sap_table: str) -> str | None:
    """Look up the SAP context category for a known SAP table."""
    rules = {r.sap_table: r.required_context_category for r in _SAP_CONTEXT_RULES}
    return rules.get(sap_table)


def _infer_from_mapping_sheets(
    sheets: dict[str, list[dict[str, str]]],
) -> tuple[dict[str, dict[str, Any]], list[str], list[str]]:
    """Infer canonical objects from mapping sheets and return ops-ready objects."""
    objects: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    assumptions: list[str] = ["Target system is assumed to be SAP S/4HANA (S4)."]

    for sheet_name, rows in sheets.items():
        if not sheet_name.endswith(_MAPPING_SHEET_SUFFIX):
            continue
        if not rows:
            warnings.append(f"Mapping sheet '{sheet_name}' is empty.")
            continue

        headers = set(rows[0].keys())
        missing = _REQUIRED_COLUMNS - headers
        if missing:
            warnings.append(
                f"Sheet '{sheet_name}' is missing required columns: {', '.join(sorted(missing))}."
            )
            continue

        domain_id = _sheet_domain_id(sheet_name)
        if domain_id not in objects:
            objects[domain_id] = {
                "id": domain_id,
                "type": "MasterDataDomain",
                "status": "draft",
                "name": _sheet_domain_name(sheet_name),
            }

        for idx, row in enumerate(rows, start=2):
            source_field_col = _find_column(row, "source_field")
            source_system_col = _find_column(row, "source_system")
            target_table_col = _find_column(row, "target_table")
            target_field_col = _find_column(row, "target_field")
            owner_col = _find_column(row, "owner")
            status_col = _find_column(row, "status")
            condition_col = _find_column(row, "condition")
            validation_rule_col = _find_column(row, "validation_rule")

            source_field = row.get(source_field_col or "", "") if source_field_col else ""
            source_system = row.get(source_system_col or "", "") if source_system_col else ""
            target_table = row.get(target_table_col or "", "") if target_table_col else ""
            target_field = row.get(target_field_col or "", "") if target_field_col else ""
            owner = row.get(owner_col or "", "") if owner_col else ""
            status = row.get(status_col or "", "").strip().lower() if status_col else ""
            condition = row.get(condition_col or "", "") if condition_col else ""
            validation_rule = row.get(validation_rule_col or "", "") if validation_rule_col else ""

            if not source_field:
                warnings.append(f"Sheet '{sheet_name}' row {idx} has no source_field.")
                continue

            if not owner:
                warnings.append(
                    f"Sheet '{sheet_name}' row {idx} ({source_field}) has no owner."
                )

            if source_system:
                system_id = _object_id("SYSTEM", source_system)
                if system_id not in objects:
                    objects[system_id] = {
                        "id": system_id,
                        "type": "System",
                        "status": "active",
                        "name": source_system,
                        "domain": domain_id,
                    }

            attribute_id = _object_id("ATTR", _clean_id(sheet_name.split("_")[0]), source_field)
            if attribute_id not in objects:
                objects[attribute_id] = {
                    "id": attribute_id,
                    "type": "Attribute",
                    "status": "draft",
                    "name": source_field.replace("_", " ").title(),
                    "domain": domain_id,
                }
                if owner:
                    objects[attribute_id]["accountable_team"] = owner

            if target_table:
                entity_id = _object_id("ENTITY", target_table)
                if entity_id not in objects:
                    objects[entity_id] = {
                        "id": entity_id,
                        "type": "BusinessEntity",
                        "status": "draft",
                        "name": target_table,
                        "domain": domain_id,
                    }

                context_id = _object_id("EC", target_table)
                if context_id not in objects:
                    context_payload: dict[str, Any] = {
                        "id": context_id,
                        "type": "EntityContext",
                        "status": "draft",
                        "name": f"SAP {target_table} context",
                        "domain": domain_id,
                        "entity": entity_id,
                        "sap_table": target_table,
                    }
                    context_category = _infer_context_category(target_table)
                    if context_category:
                        context_payload["context_category"] = context_category
                    objects[context_id] = context_payload

                if target_field and status != "obsolete":
                    field_id = _object_id("FEP", "S4", target_table, target_field)
                    if field_id not in objects:
                        objects[field_id] = {
                            "id": field_id,
                            "type": "FieldEndpoint",
                            "status": "draft",
                            "name": target_field,
                            "domain": domain_id,
                            "entity": entity_id,
                            "entity_context": context_id,
                            "system": "SYSTEM-S4",
                            "endpoint_type": "sap_table_field",
                            "sap_table": target_table,
                            "sap_field": target_field,
                            "business_attribute": attribute_id,
                        }
                        if owner:
                            objects[field_id]["accountable_team"] = owner

                    mapping_id = _object_id(
                        "MAP",
                        _clean_id(sheet_name.split("_")[0]),
                        source_field,
                        target_table,
                        target_field,
                    )
                    if mapping_id not in objects:
                        objects[mapping_id] = {
                            "id": mapping_id,
                            "type": "Mapping",
                            "status": "draft",
                            "name": f"{source_field} → {target_table}.{target_field}",
                            "domain": domain_id,
                            "source_endpoint": field_id,
                            "target_endpoint": field_id,
                        }
                        if owner:
                            objects[mapping_id]["accountable_team"] = owner
                elif status == "obsolete":
                    warnings.append(
                        f"Sheet '{sheet_name}' row {idx} ({source_field}) is marked obsolete."
                    )

            if status != "obsolete" and condition and not validation_rule:
                warnings.append(
                    f"Sheet '{sheet_name}' row {idx} ({source_field}) has a condition "
                    "without a validation_rule."
                )

    if not any(s.endswith(_MAPPING_SHEET_SUFFIX) for s in sheets):
        warnings.append("No mapping sheets found. Expected sheets ending with '_Mappings'.")

    return objects, warnings, assumptions


def _build_proposal(
    objects: dict[str, dict[str, Any]],
    source_evidence: str,
) -> dict[str, Any]:
    """Turn inferred objects into a PatchProposal."""
    operations: list[PatchOperation] = []
    for obj_id in sorted(objects):
        obj = objects[obj_id]
        operations.append(
            PatchOperation(
                op="create_object",
                object_id=obj_id,
                object_type=obj["type"],
                reason="Inferred from mapping workbook",
                after=obj,
            )
        )
    proposal_id = f"PP-BOOTSTRAP-{datetime.now(UTC).strftime('%Y%m%d')}"
    return build_patch_proposal(
        proposal_id=proposal_id,
        operations=operations,
        affected_objects=[],
        source_evidence=source_evidence,
        created_by="bootstrap-assessment",
        generated_by="bootstrap-assessment",
    )


def _render_report(result: BootstrapResult) -> str:
    """Render bootstrap report as Markdown."""
    lines: list[str] = [
        "# Bootstrap Assessment Report",
        "",
        f"**Repository**: {result.repo_name}",
        f"**Path**: {result.repo_root}",
        f"**Inferred objects**: {result.inferred_objects_count}",
        "",
        "## Warnings",
        "",
    ]
    if result.warnings:
        for warning in result.warnings:
            lines.append(f"- {warning}")
    else:
        lines.append("_No warnings._")
    lines.append("")

    lines.extend(["## Assumptions", ""])
    for assumption in result.assumptions:
        lines.append(f"- {assumption}")
    lines.append("")

    lines.extend(["## Next steps", ""])
    for cmd in result.next_commands:
        lines.append(f"- `{cmd}`")
    lines.append("")

    return "\n".join(lines)


def bootstrap_assessment(
    mapping_path: Path,
    out_repo: Path,
    name: str,
    dataset_path: Path | None = None,
    evidence_paths: list[Path] | None = None,
) -> BootstrapResult:
    """Bootstrap a repository from a mapping workbook.

    Args:
        mapping_path: Path to the XLSX mapping workbook.
        out_repo: Directory to create as the new model repository.
        name: Repository name.
        dataset_path: Optional sample dataset path to reference.
        evidence_paths: Optional evidence file paths to reference.

    Returns:
        ``BootstrapResult`` with proposal and report paths.

    Raises:
        ValueError: If the workbook layout is unsupported or no objects could be inferred.
    """
    mapping_path = mapping_path.resolve()
    out_repo = out_repo.resolve()
    if out_repo.exists():
        raise ValueError(f"Output repository already exists: {out_repo}")

    sheets = _read_workbook(mapping_path)
    objects, warnings, assumptions = _infer_from_mapping_sheets(sheets)

    if not objects:
        raise ValueError(
            "Unsupported workbook layout: could not infer any model objects. "
            "Ensure sheets end with '_Mappings' and include source_field, target_table, "
            "and target_field columns."
        )

    # Scaffold repository
    out_repo.mkdir(parents=True, exist_ok=True)
    (out_repo / "model").mkdir(exist_ok=True)
    (out_repo / "generated").mkdir(exist_ok=True)
    (out_repo / "data" / "samples").mkdir(parents=True, exist_ok=True)

    config = RepoConfig(name=name)
    config_path = out_repo / "modelops.config.yaml"
    config_path.write_text(
        yaml.safe_dump(config.model_dump(), default_flow_style=False, sort_keys=False),
        encoding="utf-8",
    )

    source_evidence = f"Bootstrap from {mapping_path.name}"
    if dataset_path:
        source_evidence += f" with dataset {dataset_path.name}"
    if evidence_paths:
        source_evidence += " and evidence " + ", ".join(p.name for p in evidence_paths)

    proposal = _build_proposal(objects, source_evidence)
    proposal_path = write_patch_proposal(proposal, out_repo / "model")

    report_md_path = out_repo / "bootstrap-report.md"
    report_json_path = out_repo / "bootstrap-report.json"

    next_commands = [
        f"martenweave validate --repo {out_repo}",
        f"martenweave proposal validate --repo {out_repo} --proposal {proposal['id']}",
        "martenweave proposal diff --repo {out_repo} --proposal {id}",
    ]

    result = BootstrapResult(
        repo_root=out_repo,
        repo_name=name,
        proposal_path=proposal_path,
        report_md_path=report_md_path,
        report_json_path=report_json_path,
        inferred_objects_count=len(objects),
        warnings=warnings,
        assumptions=assumptions,
        next_commands=next_commands,
    )

    report_md_path.write_text(_render_report(result), encoding="utf-8")
    report_json_path.write_text(
        json.dumps(
            {
                "tool": "martenweave",
                "version": __version__,
                "repo_name": name,
                "repo_path": str(out_repo),
                "inferred_objects_count": len(objects),
                "proposal_id": proposal["id"],
                "proposal_path": str(proposal_path),
                "warnings": warnings,
                "assumptions": assumptions,
                "next_commands": next_commands,
            },
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )

    return result
