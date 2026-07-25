"""Pilot input privacy preflight checks.

Inspects mapping workbooks, datasets, evidence notes, and validation reports
before they are used in a pilot assessment. The default mode is metadata-only:
raw row values are not written to the report unless explicitly requested.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from modelops_core.guardrails.secrets import scan_file, scan_text
from modelops_core.imports.dataset_profiler import profile_csv, profile_xlsx
from modelops_core.pilot.structural_scan import scan_workbook_structure
from modelops_core.pilot.workbook_suggestion_review import write_workbook_suggestion_review_xlsx
from modelops_core.pilot.workbook_suggestions import (
    generate_workbook_suggestions,
    write_workbook_suggestions_json,
    write_workbook_suggestions_markdown,
)

MAX_SCAN_CELLS = 5_000
MAX_STRUCTURAL_SAMPLE = 25

SENSITIVE_COLUMN_TERMS = {
    "account",
    "address",
    "card",
    "credential",
    "email",
    "iban",
    "name",
    "password",
    "phone",
    "secret",
    "ssn",
    "token",
}


def _sensitive_column_warnings(columns: list[str]) -> list[str]:
    warnings: list[str] = []
    seen: set[str] = set()
    for col in columns:
        low = col.lower()
        for term in SENSITIVE_COLUMN_TERMS:
            if term in low and col not in seen:
                warnings.append(f"Column '{col}' may contain sensitive data ('{term}').")
                seen.add(col)
                break
    return warnings


def _inspect_text(path: Path) -> dict[str, Any]:
    size = path.stat().st_size
    findings = scan_file(path)
    warnings: list[str] = []
    if findings:
        warnings.append(f"{len(findings)} potential secret/pattern match(es).")
    return {
        "path": str(path),
        "file_type": path.suffix.lstrip(".").lower() or "text",
        "size_bytes": size,
        "status": "warning" if warnings else "allowed",
        "warnings": warnings,
        "secret_findings_count": len(findings),
    }


def _inspect_csv(path: Path) -> dict[str, Any]:
    profile = profile_csv(path, dataset_id=path.name)
    size = profile.status.file_size_bytes or path.stat().st_size
    columns = [c.name for c in profile.columns]
    warnings = _sensitive_column_warnings(columns)
    findings = scan_file(path)
    if findings:
        warnings.append(f"{len(findings)} potential secret/pattern match(es).")

    status = "blocked"
    if profile.status.success:
        status = "warning" if warnings else "allowed"

    result: dict[str, Any] = {
        "path": str(path),
        "file_type": "csv",
        "size_bytes": size,
        "status": status,
        "warnings": warnings,
        "secret_findings_count": len(findings),
        "row_count": profile.row_count,
        "column_count": profile.column_count,
        "columns": columns,
    }
    if not profile.status.success:
        result["reason"] = profile.status.reason
    return result


def _inspect_xlsx(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX preflight.") from exc

    profile = profile_xlsx(path, dataset_id=path.name)
    if not profile.status.success:
        return {
            "path": str(path),
            "file_type": "xlsx",
            "size_bytes": profile.status.file_size_bytes or path.stat().st_size,
            "status": "blocked",
            "warnings": [],
            "secret_findings_count": 0,
            "reason": profile.status.reason,
        }

    # Read/write mode is required for merged ranges and cell comments.  This
    # metadata pass never persists or changes the workbook.
    wb_meta = load_workbook(path, data_only=True, read_only=False, keep_links=True)
    sheet_names = list(wb_meta.sheetnames)
    hidden_sheets = [name for name in sheet_names if wb_meta[name].sheet_state != "visible"]
    merged_ranges = {
        name: [str(item) for item in wb_meta[name].merged_cells.ranges]
        for name in sheet_names
        if wb_meta[name].merged_cells.ranges
    }
    comment_count = 0
    hyperlink_count = 0
    coloured_cell_count = 0
    hyperlink_cells: dict[str, list[str]] = {}
    hidden_rows: dict[str, list[int]] = {}
    hidden_columns: dict[str, list[str]] = {}
    hidden_row_counts: dict[str, int] = {}
    hidden_column_counts: dict[str, int] = {}
    hyperlink_counts: dict[str, int] = {}
    excel_tables: list[dict[str, str]] = []
    for name in sheet_names:
        ws = wb_meta[name]
        sheet_hidden_rows = [
            index
            for index, dimension in ws.row_dimensions.items()
            if getattr(dimension, "hidden", False)
        ]
        hidden_row_counts[name] = len(sheet_hidden_rows)
        if sheet_hidden_rows:
            hidden_rows[name] = sheet_hidden_rows[:MAX_STRUCTURAL_SAMPLE]
        sheet_hidden_columns = [
            str(index)
            for index, dimension in ws.column_dimensions.items()
            if getattr(dimension, "hidden", False)
        ]
        hidden_column_counts[name] = len(sheet_hidden_columns)
        if sheet_hidden_columns:
            hidden_columns[name] = sheet_hidden_columns[:MAX_STRUCTURAL_SAMPLE]
        for table in ws.tables.values():
            excel_tables.append({"sheet": name, "name": table.name, "ref": table.ref})
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    comment_count += 1
                if cell.hyperlink is not None:
                    hyperlink_count += 1
                    hyperlink_counts[name] = hyperlink_counts.get(name, 0) + 1
                    hyperlink_cells.setdefault(name, [])
                    if len(hyperlink_cells[name]) < MAX_STRUCTURAL_SAMPLE:
                        hyperlink_cells[name].append(cell.coordinate)
                if cell.fill.fill_type and cell.fill.fill_type != "none":
                    coloured_cell_count += 1
    external_links: list[str] = []
    for link in getattr(wb_meta, "external_links", []) or []:
        target = getattr(link, "Target", None) or str(link)
        if target:
            external_links.append(target)
    defined_names: list[dict[str, str]] = []
    for name, defined_name in wb_meta.defined_names.items():
        target = getattr(defined_name, "attr_text", "") or ""
        defined_names.append({"name": name, "target": target, "type": defined_name.type or ""})
    wb_meta.close()

    warnings: list[str] = []
    if hidden_sheets:
        warnings.append(f"Hidden sheet(s): {', '.join(hidden_sheets)}.")
    if external_links:
        warnings.append(f"External link(s): {', '.join(external_links)}.")
    if merged_ranges:
        warnings.append(
            "Merged cell range(s) present: "
            + "; ".join(f"{name} ({len(ranges)})" for name, ranges in merged_ranges.items())
            + "."
        )
    if hidden_rows:
        warnings.append(
            "Hidden row(s): "
            + "; ".join(f"{name} ({hidden_row_counts[name]})" for name in hidden_rows)
            + "."
        )
    if hidden_columns:
        warnings.append(
            "Hidden column(s): "
            + "; ".join(f"{name} ({hidden_column_counts[name]})" for name in hidden_columns)
            + "."
        )
    if comment_count:
        warnings.append(f"{comment_count} cell comment(s) present for reviewer context.")
    if hyperlink_count:
        warnings.append(f"{hyperlink_count} hyperlink cell(s) present.")
    if coloured_cell_count:
        warnings.append(
            f"{coloured_cell_count} colour-filled cell(s) present; colour-only statuses are "
            "evidence for human review and are not interpreted as data."
        )
    if defined_names:
        warnings.append(f"{len(defined_names)} defined name(s) present.")
    if excel_tables:
        warnings.append(f"{len(excel_tables)} Excel table definition(s) present.")

    all_columns: list[str] = []
    sheet_metadata: list[dict[str, Any]] = []
    for sheet in profile.sheets:
        cols = [c.name for c in sheet.columns]
        all_columns.extend(cols)
        sheet_metadata.append(
            {
                "name": sheet.sheet_name,
                "row_count": sheet.row_count,
                "column_count": sheet.column_count,
                "columns": cols,
                "hidden_row_count": hidden_row_counts.get(sheet.sheet_name, 0),
                "hidden_rows": hidden_rows.get(sheet.sheet_name, []),
                "hidden_column_count": hidden_column_counts.get(sheet.sheet_name, 0),
                "hidden_columns": hidden_columns.get(sheet.sheet_name, []),
                "hyperlink_count": hyperlink_counts.get(sheet.sheet_name, 0),
                "hyperlink_cells": hyperlink_cells.get(sheet.sheet_name, []),
                "tables": [table for table in excel_tables if table["sheet"] == sheet.sheet_name],
                "included": sheet.sheet_name not in hidden_sheets,
                "exclusion_reason": (
                    "Hidden worksheet; retained as evidence but excluded from interpretation."
                    if sheet.sheet_name in hidden_sheets
                    else None
                ),
            }
        )
    warnings.extend(_sensitive_column_warnings(all_columns))

    formula_count = 0
    try:
        wb_formula = load_workbook(path, data_only=False, read_only=True)
        for sheet_name in wb_formula.sheetnames:
            ws = wb_formula[sheet_name]
            for row in ws.iter_rows(min_row=1):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb_formula.close()
    except Exception:  # pragma: no cover - best effort
        pass
    if formula_count:
        warnings.append(f"{formula_count} formula cell(s) present.")

    secret_findings: list[Any] = []
    cells_scanned = 0
    try:
        wb_values = load_workbook(path, data_only=True, read_only=True)
        value_lines: list[str] = []
        for sheet_name in wb_values.sheetnames:
            ws = wb_values[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue
                    text = str(value).strip()
                    if text:
                        value_lines.append(text)
                        cells_scanned += 1
                    if cells_scanned >= MAX_SCAN_CELLS:
                        break
                if cells_scanned >= MAX_SCAN_CELLS:
                    break
            if cells_scanned >= MAX_SCAN_CELLS:
                break
        wb_values.close()
        if value_lines:
            secret_findings = scan_text("\n".join(value_lines))
    except Exception:  # pragma: no cover - best effort
        pass
    if secret_findings:
        warnings.append(f"{len(secret_findings)} potential secret/pattern match(es).")

    status = "warning" if warnings else "allowed"
    return {
        "path": str(path),
        "file_type": "xlsx",
        "size_bytes": profile.status.file_size_bytes or path.stat().st_size,
        "status": status,
        "warnings": warnings,
        "secret_findings_count": len(secret_findings),
        "sheet_names": sheet_names,
        "sheets": sheet_metadata,
        "hidden_sheets": hidden_sheets,
        "merged_ranges": merged_ranges,
        "comment_count": comment_count,
        "hyperlink_count": hyperlink_count,
        "coloured_cell_count": coloured_cell_count,
        "hyperlink_cells": hyperlink_cells,
        "external_links": external_links,
        "hidden_rows": hidden_rows,
        "hidden_columns": hidden_columns,
        "defined_names": defined_names,
        "excel_tables": excel_tables,
        "formula_count": formula_count,
        "assumptions": [
            "Only visible worksheets are included in the initial interpretation.",
            (
                "Merged cells and comments are reported for human review, not converted into "
                "model data."
            ),
            (
                "External links and formulas remain source evidence and are never executed by "
                "Martenweave."
            ),
            (
                "Hidden rows, hidden columns, hyperlinks, named ranges, and Excel tables "
                "remain structural evidence only."
            ),
        ],
    }


def inspect_file(path: Path) -> dict[str, Any]:
    """Inspect a single pilot input file and return a metadata result."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _inspect_xlsx(path)
    if suffix == ".csv":
        return _inspect_csv(path)
    if suffix in {".md", ".txt"}:
        return _inspect_text(path)
    return {
        "path": str(path),
        "file_type": suffix.lstrip(".") or "unknown",
        "size_bytes": path.stat().st_size,
        "status": "blocked",
        "warnings": [],
        "secret_findings_count": 0,
        "reason": "Unsupported file type for pilot preflight.",
    }


@dataclass
class PreflightReport:
    """Container for a completed preflight run."""

    generated_at: str
    overall_status: str
    files: list[dict[str, Any]]
    include_raw_samples: bool
    generated_artifacts: list[str]

    def to_dict(self) -> dict[str, Any]:
        return {
            "generated_at": self.generated_at,
            "overall_status": self.overall_status,
            "include_raw_samples": self.include_raw_samples,
            "files": self.files,
            "generated_artifacts": self.generated_artifacts,
        }


def run_preflight(
    mapping_path: Path,
    dataset_paths: list[Path],
    evidence_paths: list[Path],
    validation_report_paths: list[Path],
    out_dir: Path,
    include_raw_samples: bool = False,
) -> PreflightReport:
    """Inspect all pilot inputs and write JSON/Markdown preflight reports."""
    out_dir.mkdir(parents=True, exist_ok=True)

    files: list[dict[str, Any]] = []
    generated_artifacts: list[str] = []
    files.append(inspect_file(mapping_path))
    for p in dataset_paths:
        files.append(inspect_file(p))
    for p in evidence_paths:
        files.append(inspect_file(p))
    for p in validation_report_paths:
        files.append(inspect_file(p))

    if any(f["status"] == "blocked" for f in files):
        overall = "blocked"
    elif any(f["status"] == "warning" for f in files):
        overall = "warning"
    else:
        overall = "allowed"

    report = PreflightReport(
        generated_at=datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        overall_status=overall,
        files=files,
        include_raw_samples=include_raw_samples,
        generated_artifacts=generated_artifacts,
    )

    if mapping_path.suffix.lower() == ".xlsx":
        workbook_manifest = scan_workbook_structure(mapping_path)
        workbook_manifest_path = out_dir / "workbook_manifest.json"
        workbook_manifest_path.write_text(
            json.dumps(workbook_manifest.to_dict(), indent=2, default=str),
            encoding="utf-8",
        )
        workbook_suggestions = generate_workbook_suggestions(workbook_manifest)
        workbook_suggestions_json_path = write_workbook_suggestions_json(
            workbook_suggestions,
            out_dir / "workbook_suggestions.json",
        )
        workbook_suggestions_markdown_path = write_workbook_suggestions_markdown(
            workbook_suggestions,
            out_dir / "workbook_suggestions.md",
        )
        workbook_suggestion_review_path = write_workbook_suggestion_review_xlsx(
            workbook_suggestions,
            out_dir / "workbook_suggestion_review.xlsx",
        )
        report.generated_artifacts.append(str(workbook_manifest_path))
        report.generated_artifacts.append(str(workbook_suggestions_json_path))
        report.generated_artifacts.append(str(workbook_suggestions_markdown_path))
        report.generated_artifacts.append(str(workbook_suggestion_review_path))

    json_path = out_dir / "preflight_report.json"
    json_path.write_text(
        json.dumps(report.to_dict(), indent=2, default=str),
        encoding="utf-8",
    )
    md_path = out_dir / "preflight_report.md"
    md_path.write_text(_render_markdown(report), encoding="utf-8")
    return report


def _render_markdown(report: PreflightReport) -> str:
    lines = [
        "# Pilot Input Preflight Report",
        "",
        f"Generated: {report.generated_at}",
        f"Overall status: **{report.overall_status.upper()}**",
        "",
        "## Files",
        "",
        "| File | Type | Size | Status | Warnings |",
        "|---|---|---|---|---|",
    ]
    for f in report.files:
        warnings = "; ".join(f.get("warnings", [])) or "-"
        lines.append(
            f"| `{Path(f['path']).name}` | {f['file_type']} | "
            f"{f['size_bytes']} | {f['status']} | {warnings} |"
        )
    lines.append("")
    lines.append("## Details")
    lines.append("")
    if report.generated_artifacts:
        lines.append("## Generated artifacts")
        lines.append("")
        for artifact in report.generated_artifacts:
            lines.append(f"- `{Path(artifact).name}`")
        lines.append("")
    for f in report.files:
        lines.append(f"### {Path(f['path']).name}")
        lines.append("")
        lines.append(f"- **Path**: {f['path']}")
        lines.append(f"- **Status**: {f['status']}")
        if f.get("reason"):
            lines.append(f"- **Reason**: {f['reason']}")
        if f.get("sheet_names"):
            lines.append(f"- **Sheets**: {', '.join(f['sheet_names'])}")
        if f.get("hidden_sheets"):
            lines.append(f"- **Hidden sheets**: {', '.join(f['hidden_sheets'])}")
        if f.get("row_count") is not None:
            lines.append(f"- **Rows**: {f['row_count']}")
        if f.get("column_count") is not None:
            lines.append(f"- **Columns**: {f['column_count']}")
        if f.get("columns"):
            lines.append(f"- **Column names**: {', '.join(f['columns'])}")
        if f.get("hidden_rows"):
            lines.append(
                "- **Hidden rows**: "
                + "; ".join(
                    f"{sheet} ({', '.join(str(row) for row in rows)})"
                    for sheet, rows in f["hidden_rows"].items()
                )
            )
        if f.get("hidden_columns"):
            lines.append(
                "- **Hidden columns**: "
                + "; ".join(
                    f"{sheet} ({', '.join(columns)})"
                    for sheet, columns in f["hidden_columns"].items()
                )
            )
        if f.get("defined_names"):
            lines.append(
                "- **Defined names**: "
                + "; ".join(f"{item['name']} -> {item['target']}" for item in f["defined_names"])
            )
        if f.get("excel_tables"):
            lines.append(
                "- **Excel tables**: "
                + "; ".join(
                    f"{item['sheet']}.{item['name']} ({item['ref']})" for item in f["excel_tables"]
                )
            )
        if f.get("hyperlink_count"):
            lines.append(f"- **Hyperlink cells**: {f['hyperlink_count']}")
        if f.get("warnings"):
            lines.append("- **Warnings**:")
            for w in f["warnings"]:
                lines.append(f"  - {w}")
        lines.append("")
    return "\n".join(lines)
